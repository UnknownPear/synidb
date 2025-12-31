import os
import io
import re
import csv
import json
import time
import base64
import hashlib
from fastapi.encoders import jsonable_encoder
from fastapi.responses import JSONResponse
import tempfile, os, json as _json
from typing import Optional, List, Dict, Any, Tuple
from urllib.parse import quote
from contextlib import contextmanager
import mimetypes
from dotenv import load_dotenv
from pathlib import Path
import json as _json_mod
try:
    import redis, rq  # background queue deps
    HAVE_RQ = True
except Exception:      # ModuleNotFoundError or anything similar
    redis = None
    rq = None
    HAVE_RQ = False
from fastapi.responses import StreamingResponse

ENV_PATH = Path(__file__).resolve().parent.parent / ".env"

from datetime import datetime, date, timedelta, timezone

load_dotenv(dotenv_path=ENV_PATH) 
# main.py
from pydantic import BaseModel, HttpUrl
from typing import List, Optional, Dict, Any
from fastapi import UploadFile, File, Form
EBAY_TRADING_ENDPOINT = os.getenv("EBAY_TRADING_ENDPOINT", "https://api.ebay.com/ws/api.dll")
EBAY_BROWSE_ENDPOINT  = os.getenv("EBAY_BROWSE_ENDPOINT",  "https://api.ebay.com/buy/browse/v1")
EBAY_OAUTH_TOKEN_URL = os.getenv("EBAY_OAUTH_TOKEN_URL", "https://api.ebay.com/identity/v1/oauth2/token")
EBAY_CLIENT_ID = os.getenv("EBAY_CLIENT_ID")
EBAY_CLIENT_SECRET = os.getenv("EBAY_CLIENT_SECRET")
EBAY_NS = {"e": "urn:ebay:apis:eBLBaseComponents"}
assert "e" in EBAY_NS and EBAY_NS["e"].endswith("eBLBaseComponents"), "EBAY_NS malformed"


class HintedPreviewHints(BaseModel):
    header_row: int
    column_roles: Dict[int, str]             # e.g., {0:"product_name_raw", 3:"qty", 5:"unit_cost"}
    selection_rows: Optional[List[int]] = None
    selection_cols: Optional[List[int]] = None
    examples: Optional[List[Dict[str, Any]]] = None   # optional {input_row:{}, expected:{}}
    expand_units: Optional[bool] = False
    vendor_id: Optional[str] = None
    vendor_name: Optional[str] = None

class UploadPreviewResponse(BaseModel):
    ok: bool
    vendor_id: str
    file_name: str
    new_po_lines: List[Dict[str, Any]]
    existing_pos_summary: List[Dict[str, Any]] = []
    ai_notes: Optional[str] = None
    ai_model: Optional[str] = None




# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€ third-party
import requests
from bs4 import BeautifulSoup, FeatureNotFound
from fastapi import (
    FastAPI,
    UploadFile,
    File,
    Form,
    Body,
    HTTPException,
    Request,
    Query,
    Response,
    Path,
)
from fastapi import APIRouter
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import psycopg2
import psycopg2.extras
from psycopg2.extras import register_uuid as _pg_register_uuid
try:
    _pg_register_uuid()
except Exception:
    pass
from psycopg2.extras import RealDictCursor
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
from requests.utils import dict_from_cookiejar
from uuid import UUID
import os, uuid, json, asyncio
# near the top with other imports
from fastapi.responses import JSONResponse

import asyncio
from starlette.responses import StreamingResponse

import redis as redis_lib
import rq

REDIS_URL = os.getenv("REDIS_URL", "redis://redis:6379/0")

def get_redis():
    # decode_responses=False so we get bytes from Pub/Sub (safer)
    return redis_lib.from_url(REDIS_URL, decode_responses=False)

# Global connection + queue (safe to create at import time)
rconn = get_redis()
q = rq.Queue("aiq", connection=rconn, default_timeout=1800)

_subscribers: list[asyncio.Queue] = []

from routes_label_inventory import router as label_inventory_router


def _redis_conn():
    # reuse the global connection if it exists, else create one
    try:
        return rconn  # created via get_redis()
    except NameError:
        import redis as _redis_mod
        return _redis_mod.from_url(REDIS_URL)

def _broadcast(evt_type: str, data: dict | None = None):
    payload = {"type": evt_type, "data": data or {}}
    for q in list(_subscribers):
        try:
            q.put_nowait(payload)
        except Exception:
            pass

# --- STREAMING AI PREVIEW WITH PROGRESS (SSE) ---------------------------------
from fastapi.responses import StreamingResponse

def _sse(data: dict) -> str:
    # Server-Sent Events frame
    return f"data: {json.dumps(data, default=str)}\n\n"


def _heartbeat() -> str:
    # comment line is valid SSE; prevents buffering/timeouts
    return ": keep-alive\n\n"
def _sse_heartbeat() -> str:
    return ": keep-alive\n\n"


REDIS_URL = os.getenv("REDIS_URL", "redis://localhost:6379/0")
if HAVE_RQ:
    rds = redis.Redis.from_url(REDIS_URL)
    q = rq.Queue("aiq", connection=rds, default_timeout=900)
else:
    rds = None
    q = None

def _sse_frame(data: dict) -> str:
    return f"data: {json.dumps(data, default=str)}\n\n"

# Map the common sheet/doc types we upload
_EXT_TO_MIME = {
    "csv": "text/csv",
    "tsv": "text/tab-separated-values",
    "xls": "application/vnd.ms-excel",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}

def _mime_for_ext(ext: str) -> str:
    return _EXT_TO_MIME.get((ext or "").lower(), "application/octet-stream")

def _gemini_parse_inline(
    file_bytes: bytes,
    filename: str,
    ext: str,
    expand_units: bool,
    model=None,
) -> dict:
    """
    Send the file to Gemini WITHOUT the files API.
    - Spreadsheets are converted to CSV and sent as inline bytes (text/csv).
    - Other types (e.g., PDF) are sent as inline bytes with the right mime.
    Returns a dict with at least {"lines":[...]}.
    """
    if model is None:
        model, _, _ = make_gemini_model()

    # Force spreadsheets â†’ CSV
    e = (ext or "").lower().lstrip(".")
    if e in ("xlsx", "xls"):
        file_bytes = _xlsx_to_csv_bytes(file_bytes, max_rows=2000)
        filename = re.sub(r"\.(xlsx|xls)$", ".csv", filename, flags=re.I)
        e = "csv"

    prompt = make_ai_parser_prompt(expand_units=expand_units)

    parts = [
        {"text": prompt + "\n\nReturn ONLY the JSON object matching the schema."}
    ]

    # Always send the file bytes as inline_data to avoid the bad {'mime_type','text'} shape
    mime = "text/csv" if e in ("csv", "tsv", "txt") else _mime_for_ext(e)
    parts.append({"inline_data": {"mime_type": mime, "data": file_bytes}})

    resp = model.generate_content(parts)
    data = parse_gemini_json(resp)
    if not data or "lines" not in data:
        raise RuntimeError("Gemini did not return 'lines' in JSON")
    return data

    

def _state_str(x):
    if x is None:
        return ""
    name = getattr(x, "name", None)
    return (name if name is not None else str(x)).upper()

def _uuid_list(ids):
    """Validate & normalize incoming UUID strings for ANY(%s::uuid[])."""
    out = []
    for x in ids or []:
        try:
            out.append(str(UUID(str(x))))
        except Exception:
            pass
    return out

def _ebay_app_access_token(scopes: list[str] | None = None) -> str:
    """
    Get an application access token (client_credentials) for Buy/Browse.
    Requires EBAY_CLIENT_ID / EBAY_CLIENT_SECRET in env.
    """
    if not EBAY_CLIENT_ID or not EBAY_CLIENT_SECRET:
        raise RuntimeError("EBAY_CLIENT_ID / EBAY_CLIENT_SECRET not set")

    scope_str = " ".join(scopes or ["https://api.ebay.com/oauth/api_scope/buy.browse.readonly"])
    auth = base64.b64encode(f"{EBAY_CLIENT_ID}:{EBAY_CLIENT_SECRET}".encode()).decode()
    headers = {
        "Authorization": f"Basic {auth}",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    data = {
        "grant_type": "client_credentials",
        "scope": scope_str,
    }
    r = requests.post(EBAY_OAUTH_TOKEN_URL, headers=headers, data=data, timeout=20)
    r.raise_for_status()
    j = r.json()
    return j["access_token"]
def _row_val(row, key, idx=0, default=0):
    """
    Safely get a column from a DB row that might be a dict or tuple.
    key: column name (when row is a dict)
    idx: index (when row is a tuple)
    """
    if row is None:
        return default
    # dict-like
    if isinstance(row, dict):
        return row.get(key, default)
    # tuple-like
    try:
        return row[idx]
    except Exception:
        return default


def wait_for_files_active(files, timeout_sec: float = 30.0, poll: float = 0.5):
    if not _genai_loaded:
        return
    deadline = time.time() + timeout_sec
    for f in files or []:
        name = getattr(f, "name", None) or getattr(f, "uri", None)
        if not name:
            continue
        while True:
            meta = genai.get_file(name=name)
            st = _state_str(getattr(meta, "state", None))
            if st == "ACTIVE":
                break
            if st == "FAILED":
                raise RuntimeError(f"Gemini file processing failed: {name}")
            if time.time() > deadline:
                raise TimeoutError(f"Gemini file processing timed out: {name} (last state={st})")
            time.sleep(poll)

def get_redis():
    """Return a connected Redis client or raise."""
    global _rconn
    if _rconn is None:
        _rconn = redis_lib.from_url(REDIS_URL, decode_responses=False)
        # Force a DNS/connect check early so we fail fast & clearly
        _rconn.ping()
    return _rconn

def get_queue():
    """Return an RQ queue if Redis is available, else None."""
    global _q
    if _q is not None:
        return _q
    try:
        conn = get_redis()
        _q = rq.Queue("aiq", connection=conn, default_timeout=1800)
        return _q
    except Exception as e:
        # Donâ€™t crash import; callers can decide what to do
        print(f"[WARN] Redis unavailable: {e} (REDIS_URL={REDIS_URL})")
        return None

def _el_text(el):
    if el is None:
        return None
    # Standard ElementTree puts text in .text
    t = el.text
    return t.strip() if isinstance(t, str) else t

def _price_from_el(el):
    # <SomePrice currencyID="USD">123.45</SomePrice>
    if el is None:
        return (None, None)
    cur = el.attrib.get("currencyID")
    raw = _el_text(el)
    try:
        return (float(raw), cur)
    except (TypeError, ValueError):
        return (None, cur)

def _first_non_null(*vals):
    for v in vals:
        if v is not None:
            return v
    return None

def _first_non_null(*vals):
    for v in vals:
        if v is not None:
            return v
    return None


def _xlsx_to_csv_bytes(xlsx_bytes: bytes, max_rows: int | None = 2000) -> bytes:
    """
    Convert the first worksheet of an .xlsx/.xls file into UTF-8 CSV bytes.
    Requires `openpyxl` for .xlsx/.xls.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl is required for XLS/XLSX conversion") from e

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    out = io.StringIO()
    w = csv.writer(out, lineterminator="\n")
    rows = 0
    for row in ws.iter_rows(values_only=True):
        w.writerow(["" if v is None else str(v) for v in row])
        rows += 1
        if max_rows and rows >= max_rows:
            break
    return out.getvalue().encode("utf-8")
def normalize_spreadsheet_upload(filename: str, ext: str, content: bytes) -> tuple[str, str, bytes]:
    """
    If the incoming file is XLS/XLSX, convert it to CSV for Gemini.
    Returns (new_filename, new_ext, new_content_bytes).
    """
    clean_ext = (ext or "").lower().lstrip(".")
    if clean_ext in ("xlsx", "xls"):
        print("[AI Preview] Converting spreadsheet to CSV for Gemini", flush=True)
        try:
            csv_bytes = _xlsx_to_csv_bytes(content)
        except Exception as conv_err:
            # Surface the error to trigger your existing local-fallback path
            print("[AI Preview] XLSXâ†’CSV conversion failed; using local preview:", repr(conv_err), flush=True)
            raise
        new_filename = re.sub(r"\.(xlsx|xls)$", ".csv", filename, flags=re.I)
        return new_filename, "csv", csv_bytes
    return filename, ext, content

def _strip_code_fences(s: str) -> str:
    s = s.strip()
    # ```json ... ``` or ``` ... ```
    if s.startswith("```"):
        s = re.sub(r"^```[a-zA-Z0-9_-]*\s*", "", s, flags=re.DOTALL)
        s = re.sub(r"\s*```$", "", s, flags=re.DOTALL)
    return s.strip()

def _find_balanced_json(s: str) -> str | None:
    """Return the first balanced {...} or [...] block from s, or None."""
    starts = [i for i in (s.find("{"), s.find("[")) if i != -1]
    if not starts:
        return None
    start = min(starts)
    stack = []
    open_to_close = {"{": "}", "[": "]"}
    for i, ch in enumerate(s[start:], start):
        if ch in "{[":
            stack.append(open_to_close[ch])
        elif stack and ch == stack[-1]:
            stack.pop()
            if not stack:
                return s[start:i + 1]
    return None

def parse_gemini_json(resp) -> dict:
    """
    Robustly parse JSON from a Gemini response object across SDK variants:
    - prefer resp.parsed (structured output)
    - else try resp.text (strip code fences)
    - else gather all text parts from candidates and parse
    - finally extract the first balanced JSON block if needed
    """
    # 1) Structured output path
    if hasattr(resp, "parsed") and resp.parsed:
        return resp.parsed

    candidates_text: list[str] = []

    # 2) Direct text
    t = getattr(resp, "text", None)
    if isinstance(t, str) and t.strip():
        candidates_text.append(t)

    # 3) Aggregate text from parts (some SDKs only fill parts)
    for cand in getattr(resp, "candidates", []) or []:
        content = getattr(cand, "content", None)
        if content and getattr(content, "parts", None):
            for p in content.parts:
                txt = getattr(p, "text", None)
                if isinstance(txt, str) and txt.strip():
                    candidates_text.append(txt)

    # Try strict JSON first, then with code-fence strip, then balanced find
    for s in candidates_text:
        try:
            return json.loads(s)
        except Exception:
            pass
        s2 = _strip_code_fences(s)
        if s2 and s2 != s:
            try:
                return json.loads(s2)
            except Exception:
                pass
        s3 = _find_balanced_json(s)
        if s3:
            try:
                return json.loads(s3)
            except Exception:
                pass

    raise ValueError("model_returned_no_json")


class CreatePOBody(BaseModel):
    po_number: str
    vendor_id: str | None = None
    vendor_name: str | None = None  # allow creating/selecting vendor by name


# Early router to allow defining routes before the FastAPI app is created
early_router = APIRouter()

# --- Synergy: preview & mint ---------------------------------------------------
from typing import Dict, Any, List, Optional
from pydantic import BaseModel
from uuid import UUID

class SynergyReq(BaseModel):
    line_ids: Optional[List[str]] = None
    use_line_categories: bool = True
    for_all_missing: bool = False

def max_seq_for_prefix(prefix: str) -> int:
    if not prefix:
        return 0
    if prefix in seq_by_prefix:
        return seq_by_prefix[prefix]

    max_seq = 0

    # items.synergy_id
    cur.execute(
        """
        SELECT COALESCE(MAX(CAST(SPLIT_PART(synergy_id, '-', 2) AS INTEGER)), 0) AS max_seq
        FROM items
        WHERE synergy_id LIKE %s
        """,
        (prefix + "-%",),
    )
    r1 = cur.fetchone()
    max_seq = max(max_seq, int(_row_val(r1, "max_seq", 0, 0) or 0))

    # po_lines.synergy_id
    cur.execute(
        """
        SELECT COALESCE(MAX(CAST(SPLIT_PART(synergy_id, '-', 2) AS INTEGER)), 0) AS max_seq
        FROM po_lines
        WHERE synergy_id LIKE %s
        """,
        (prefix + "-%",),
    )
    r2 = cur.fetchone()
    max_seq = max(max_seq, int(_row_val(r2, "max_seq", 0, 0) or 0))

    seq_by_prefix[prefix] = max_seq
    return max_seq

def _gemini_preview_inline_bytes(csv_bytes: bytes, expand_units: bool) -> dict:
    """
    Ask Gemini to parse CSV bytes and return {"lines":[...], "notes": "..."}.
    Avoids any RAG/FileStore and passes bytes inline as text/csv.
    """
    model, ai_model, _structured = make_gemini_model()

    # Build instructions (same schema as AI_PARSER_PROMPT)
    flags = f"EXPAND_UNITS = {'true' if expand_units else 'false'}"
    system_msg = f"{AI_PARSER_PROMPT}\n\nMODE FLAGS:\n{flags}\n\nReturn only JSON."

    file_part = {
        "inline_data": {
            "mime_type": "text/csv",
            "data": csv_bytes,
        }
    }

    resp = model.generate_content([system_msg, file_part])
    data = parse_gemini_json(resp)  # âœ… pass the response object

    return {
        "ok": True,
        "ai": True,
        "model": ai_model,
        **(data or {}),
    }


# ---------- Synergy: preview (no DB writes) ----------
@early_router.post("/pos/{po_id}/synergy_preview")
def synergy_preview(po_id: str, payload: Dict[str, Any] = Body(...)):
    """
    payload: { line_ids: string[], use_line_categories: bool }
    Returns suggested codes per line based on category prefix + current max seq
    across items.synergy_id and po_lines.synergy_id.
    """
    line_ids = _uuid_list(payload.get("line_ids") or [])
    if not line_ids:
        return {"previews": [], "ai_notes": "No line_ids provided."}

    with db() as (con, cur):
        cur.execute(
            """
            SELECT pl.id,
                   pl.qty,
                   pl.product_name_raw,
                   pl.category_guess AS category_id,
                   c.prefix
            FROM po_lines pl
            LEFT JOIN categories c ON c.id = pl.category_guess
            WHERE pl.purchase_order_id = %s
              AND pl.id = ANY(%s::uuid[])
            ORDER BY COALESCE(pl.created_at, NOW()) ASC, pl.id ASC
            """,
            (po_id, line_ids),
        )
        rows = list(cur.fetchall())

        seq_by_prefix: Dict[str, int] = {}

        # nested so it can close over cur/seq_by_prefix
        def max_seq_for_prefix(prefix: str) -> int:
            
            if not prefix:
                return 0
            if prefix in seq_by_prefix:
                return seq_by_prefix[prefix]

            max_seq = 0
            cur.execute(
                """
                SELECT COALESCE(MAX(CAST(SPLIT_PART(synergy_id, '-', 2) AS INTEGER)), 0) AS max_seq
                FROM items
                WHERE synergy_id LIKE %s
                """,
                (prefix + "-%",),
            )
            r1 = cur.fetchone()
            max_seq = max(max_seq, int(_row_val(r1, "max_seq", 0, 0) or 0))

            cur.execute(
                """
                SELECT COALESCE(MAX(CAST(SPLIT_PART(synergy_id, '-', 2) AS INTEGER)), 0) AS max_seq
                FROM po_lines
                WHERE synergy_id LIKE %s
                """,
                (prefix + "-%",),
            )
            r2 = cur.fetchone()
            max_seq = max(max_seq, int(_row_val(r2, "max_seq", 0, 0) or 0))

            seq_by_prefix[prefix] = max_seq
            return max_seq

        previews = []
        for r in rows:
            # support dict or tuple rows
            if isinstance(r, dict):
                line_id = r.get("id")
                qty = r.get("qty")
                name_raw = r.get("product_name_raw")
                prefix = (r.get("prefix") or "").strip()
            else:
                line_id, qty, name_raw, _category_id, prefix = r
                prefix = (prefix or "").strip()

            try:
                qty = int(qty or 0) or 1
            except Exception:
                qty = 1

            if not prefix:
                previews.append({
                    "line_id": str(line_id),
                    "product_name_raw": name_raw,
                    "prefix": "",
                    "qty": qty,
                    "codes": [],
                    "note": "No category prefix; cannot preview.",
                })
                continue

            start = max_seq_for_prefix(prefix) + 1
            codes = [f"{prefix}-{str(start + i).zfill(5)}" for i in range(qty)]
            seq_by_prefix[prefix] = start + qty - 1

            previews.append({
                "line_id": str(line_id),
                "product_name_raw": name_raw,
                "prefix": prefix,
                "qty": qty,
                "codes": codes,
            })

        return {"previews": previews, "ai_notes": "Local preview generated."}



# ---------- Synergy: mint (DB writes, one code per line) ----------
@early_router.post("/pos/{po_id}/mint_synergy")
def mint_synergy(po_id: str, payload: dict = Body(default={})):
    """
    Overwrite behavior:
      - If payload.overwrite is True:
          * If line_ids is provided: clear synergy_id only on those lines
          * Else: clear synergy_id on ALL lines in the PO
      - Then run your existing mint logic which mints for rows where synergy_id is NULL.
    """
    line_ids = payload.get("line_ids") or []
    overwrite = bool(payload.get("overwrite", False))

    if overwrite:
        with db() as (con, cur):
            if line_ids:
                # Coerce to UUID, then cast array to uuid[] in SQL (prevents uuid=text error)
                uuid_ids = [UUID(str(x)) for x in line_ids]
                cur.execute(
                    "UPDATE po_lines SET synergy_id = NULL WHERE id = ANY(%s::uuid[]);",
                    (uuid_ids,),
                )
            else:
                # Clear all lines in this PO (if your mint step processes whole PO)
                cur.execute(
                    "UPDATE po_lines SET synergy_id = NULL WHERE purchase_order_id = %s;",
                    (po_id,),
                )
            con.commit()

    with db() as (con, cur):
        cur.execute(
            """
            SELECT pl.id,
                   pl.qty,
                   pl.category_guess AS category_id,
                   c.prefix
            FROM po_lines pl
            LEFT JOIN categories c ON c.id = pl.category_guess
            WHERE pl.purchase_order_id = %s
              AND pl.id = ANY(%s::uuid[])
            ORDER BY COALESCE(pl.created_at, NOW()) ASC, pl.id ASC
            """,
            (po_id, line_ids),
        )
        rows = list(cur.fetchall())

        seq_by_prefix: Dict[str, int] = {}

        def max_seq_for_prefix(prefix: str) -> int:
            if not prefix:
                return 0
            if prefix in seq_by_prefix:
                return seq_by_prefix[prefix]

            max_seq = 0
            cur.execute(
                """
                SELECT COALESCE(MAX(CAST(SPLIT_PART(synergy_id, '-', 2) AS INTEGER)), 0) AS max_seq
                FROM items
                WHERE synergy_id LIKE %s
                """,
                (prefix + "-%",),
            )
            r1 = cur.fetchone()
            max_seq = max(max_seq, int(_row_val(r1, "max_seq", 0, 0) or 0))

            cur.execute(
                """
                SELECT COALESCE(MAX(CAST(SPLIT_PART(synergy_id, '-', 2) AS INTEGER)), 0) AS max_seq
                FROM po_lines
                WHERE synergy_id LIKE %s
                """,
                (prefix + "-%",),
            )
            r2 = cur.fetchone()
            max_seq = max(max_seq, int(_row_val(r2, "max_seq", 0, 0) or 0))

            seq_by_prefix[prefix] = max_seq
            return max_seq

        updated = 0
        for r in rows:
            if isinstance(r, dict):
                line_id = r.get("id")
                qty = r.get("qty")
                prefix = (r.get("prefix") or "").strip()
            else:
                line_id, qty, _category_id, prefix = r
                prefix = (prefix or "").strip()

            try:
                qty = int(qty or 0) or 1
            except Exception:
                qty = 1

            if not prefix:
                continue  # cannot mint without a category prefix

            start = max_seq_for_prefix(prefix) + 1
            code = f"{prefix}-{str(start).zfill(5)}"
            seq_by_prefix[prefix] = start + (qty - 1)  # advance block if you want to reserve qty

            cur.execute(
                "UPDATE po_lines SET synergy_id = %s WHERE id = %s",
                (code, line_id),
            )
            updated += 1

        con.commit()
        _broadcast("row.bulk_upserted", {"poId": po_id})
        return {"updated": updated, "ai_notes": "Synergy IDs minted."}


@early_router.get("/pos/{po_id}/lines")
def get_po_lines(
    po_id: str,
    q: Optional[str] = Query(None, description="Search in name/upc/asin"),
    category_ids: Optional[str] = Query(None, description="Comma-separated UUIDs"),
    uncategorized: bool = Query(False),
    has_upc: Optional[bool] = Query(None),
    has_asin: Optional[bool] = Query(None),
    min_qty: Optional[int] = Query(None),
    max_qty: Optional[int] = Query(None),
    min_cost: Optional[float] = Query(None),
    max_cost: Optional[float] = Query(None),
    sort: str = Query("id", pattern="^(id|name|qty|unit_cost)$"),
    dir: str = Query("asc", pattern="^(asc|desc)$"),
):
    where = ["pl.purchase_order_id = %s"]
    vals: list[object] = [po_id]

    if q:
        like = f"%{q.lower()}%"
        where.append("(LOWER(pl.product_name_raw) LIKE %s OR pl.upc ILIKE %s OR pl.asin ILIKE %s)")
        vals.extend([like, f"%{q}%", f"%{q}%"])

    if category_ids:
        ids = [UUID(x) for x in category_ids.split(",") if x.strip()]
        if ids:
            where.append("pl.category_guess = ANY(%s)")
            vals.append(ids)

    if uncategorized:
        where.append("pl.category_guess IS NULL")

    if has_upc is not None:
        where.append("pl.upc IS NOT NULL AND pl.upc <> ''" if has_upc
                     else "(pl.upc IS NULL OR pl.upc = '')")

    if has_asin is not None:
        where.append("pl.asin IS NOT NULL AND pl.asin <> ''" if has_asin
                     else "(pl.asin IS NULL OR pl.asin = '')")

    if min_qty is not None:
        where.append("COALESCE(pl.qty,0) >= %s"); vals.append(min_qty)
    if max_qty is not None:
        where.append("COALESCE(pl.qty,0) <= %s"); vals.append(max_qty)
    if min_cost is not None:
        where.append("COALESCE(pl.unit_cost,0) >= %s"); vals.append(min_cost)
    if max_cost is not None:
        where.append("COALESCE(pl.unit_cost,0) <= %s"); vals.append(max_cost)

    order_by_map = {"id": "pl.id", "name": "pl.product_name_raw", "qty": "pl.qty", "unit_cost": "pl.unit_cost"}
    order_by = order_by_map.get(sort, "pl.id")
    order_dir = "DESC" if dir.lower() == "desc" else "ASC"

    q_sql = f"""
        SELECT pl.id,
               pl.product_name_raw,
               pl.upc,
               pl.asin,
               pl.qty,
               pl.unit_cost,
               pl.msrp,
               pl.category_guess AS category_id,
               pl.synergy_id        -- â† added so the UI can display the minted code
        FROM po_lines pl
        WHERE {" AND ".join(where)}
        ORDER BY {order_by} {order_dir};
    """
    with db() as (con, cur):
        cur.execute(q_sql, tuple(vals))
        return [dict(r) for r in cur.fetchall()]



class BulkCategoryBody(BaseModel):
    line_ids: List[str]             # list of po_lines.id (UUIDs as strings)
    category_id: Optional[str] = None           # if provided, assign this UUID
    category_label: Optional[str] = None        # or resolve by human label/prefix
    create_if_missing: bool = False             # optional: create category if not found
    new_category_prefix: Optional[str] = None   # used only when creating

@early_router.post("/pos/lines/bulk_category")
def bulk_set_category(payload: dict = Body(...)):
    """
    Body:
      {
        "line_ids": [<uuid|string>...],
        "category_id": <uuid|string|null>
      }

    Behavior:
      - If category_id is NULL => category becomes NULL and synergy_id is cleared.
      - If category_id is a UUID => category is set; existing synergy_id is left as-is
        (user can press "Reassign Synergy IDs" to regenerate with the new prefix).
    """
    raw_ids = payload.get("line_ids") or []
    if not raw_ids:
        return {"updated": 0}

    line_ids = [UUID(str(x)) for x in raw_ids]  # ensure uuid[]
    cat = payload.get("category_id")
    category_id = UUID(str(cat)) if cat is not None else None

    with db() as (con, cur):
        if category_id is None:
            # Unassign: clear category AND synergy_id immediately
            cur.execute(
                """
                UPDATE po_lines
                   SET category_guess = NULL,
                       synergy_id     = NULL
                 WHERE id = ANY(%s::uuid[]);
                """,
                (line_ids,),
            )
        else:
            # Assign: set category only (user can reassign IDs later if desired)
            cur.execute(
                """
                UPDATE po_lines
                   SET category_guess = %s
                 WHERE id = ANY(%s::uuid[]);
                """,
                (str(category_id), line_ids),
            )
        updated = cur.rowcount
        con.commit()

    return {"updated": updated}



@early_router.patch("/purchase_orders/{po_id}")
def update_purchase_order(po_id: str, body: dict = Body(...)):
    """
    Accepts partial updates:
      - vendor_id: uuid string
      - po_number: string
    Dynamically builds SET clause; does NOT reference columns that may not exist.
    """
    vendor_id = body.get("vendor_id")
    po_number = body.get("po_number")

    set_parts = []
    values = []

    if vendor_id is not None:
      # keep as text if your DB column is uuid; psycopg2 will cast fine
      set_parts.append("vendor_id = %s")
      values.append(str(vendor_id))

    if po_number is not None:
      set_parts.append("po_number = %s")
      values.append(po_number)

    if not set_parts:
      return {"updated": 0}

    sql = f"UPDATE purchase_orders SET {', '.join(set_parts)} WHERE id = %s"
    values.append(po_id)

    with db() as (con, cur):
      cur.execute(sql, tuple(values))
      con.commit()

    return {"updated": cur.rowcount}



# --- STUB: Create a PO line --------------------------------------------------
@early_router.post("/pos/{po_id}/lines")
def create_po_line(po_id: str, body: Dict[str, Any] = Body(...)):
    """
    Create a new po_lines row (minimal fields).
    body: {
      product_name_raw: str,
      qty?: int,
      upc?: str|null,
      unit_cost?: float|null,   # price paid
      msrp?: float|null
    }
    """
    name = (body.get("product_name_raw") or "").strip()
    if not name:
      return JSONResponse({"detail": "product_name_raw required"}, status_code=400)

    qty = int(body.get("qty") or 1)
    upc = body.get("upc")
    unit_cost = body.get("unit_cost")  # may be None
    msrp = body.get("msrp")            # may be None

    with db() as (con, cur):
        try:
            # Try with unit_cost & msrp columns
            cur.execute(
                """
                INSERT INTO po_lines (purchase_order_id, product_name_raw, qty, upc, unit_cost, msrp)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (po_id, name, qty, upc, unit_cost, msrp),
            )
            new_id = cur.fetchone()["id"]
            con.commit()
            return {"ok": True, "id": str(new_id)}
        except Exception as e:
            con.rollback()
            # Fallback if schema doesn't have unit_cost/msrp yet
            try:
                with db() as (con2, cur2):
                    cur2.execute(
                        """
                        INSERT INTO po_lines (purchase_order_id, product_name_raw, qty, upc)
                        VALUES (%s, %s, %s, %s)
                        RETURNING id
                        """,
                        (po_id, name, qty, upc),
                    )
                    new_id = cur2.fetchone()["id"]
                    con2.commit()
                    return {"ok": True, "id": str(new_id), "note": "inserted without unit_cost/msrp", "error": str(e)}
            except Exception as e2:
                return JSONResponse({"ok": False, "error": str(e2)}, status_code=400)


# --- STUB: Delete a PO line --------------------------------------------------
@early_router.delete("/pos/lines/{line_id}")
def delete_po_line(line_id: str):
    with db() as (con, cur):
        try:
            cur.execute("DELETE FROM po_lines WHERE id = %s", (line_id,))
            con.commit()
            return {"ok": True, "deleted": line_id}
        except Exception as e:
            con.rollback()
            return JSONResponse({"ok": False, "error": str(e)}, status_code=400)

# --- STUB: AI Category Draft -------------------------------------------------
@early_router.post("/pos/{po_id}/ai_category_draft")
def ai_category_draft(po_id: str, body: Dict[str, Any] = Body(...)):
    """
    Returns naive suggestions so the UI can proceed. Upgrade later.
    body: { line_ids: string[] }
    Response: { suggestions: [{ line_id, category_id|null }] }
    """
    line_ids = body.get("line_ids") or []
    if not isinstance(line_ids, list) or not line_ids:
        return {"suggestions": []}

    # Simple heuristic: if a category label word appears in product_name_raw, pick that category.
    suggestions = []
    with db() as (con, cur):
        # load lines
        cur.execute(
            """
            SELECT pl.id, pl.product_name_raw
            FROM po_lines pl
            WHERE pl.purchase_order_id = %s AND pl.id = ANY(%s::uuid[])
            """,
            (po_id, _uuid_list(line_ids)),
        )
        line_rows = {str(r["id"]): (r.get("product_name_raw") or "").lower() for r in cur.fetchall()}

        # load categories
        cur.execute("SELECT id, label FROM categories")
        cats = [(str(r["id"]), (r["label"] or "").lower()) for r in cur.fetchall()]

    for lid, name in line_rows.items():
        chosen = None
        for cid, label in cats:
          # super naive token presence
          if label and any(tok and tok in name for tok in label.split()):
              chosen = cid
              break
        suggestions.append({"line_id": lid, "category_id": chosen})

    return {"suggestions": suggestions}

def guess_mime(filename: str | None, ext: str | None) -> str:
    ext = (ext or "").lower().lstrip(".")
    if ext in _EXT_TO_MIME:
        return _EXT_TO_MIME[ext]
    mt, _ = mimetypes.guess_type(filename or "")
    return mt or "application/octet-stream"

def make_gemini_model():
    # Require plain JSON output; skip response_schema to avoid TypeError on 2.5 flash
    cfg = {
        "response_mime_type": "application/json",
        # You can add temperature/topP etc. here if you like
    }
    model = genai.GenerativeModel(
        model_name="models/gemini-2.5-flash",
        generation_config=cfg,
    )
    ai_model = "gemini-2.5-flash"
    structured = False  # weâ€™re not using the schema feature here
    return model, ai_model, structured

    

def _tag_ai_response(payload: dict, via: str, model: str | None = None) -> JSONResponse:
    """
    Wraps a JSON payload and annotates whether Gemini was used.
    Adds both a field in the JSON (`via`) and HTTP headers (`X-AI-Used`, `X-AI-Model`).
    """
    out = dict(payload)
    out["via"] = via  # "gemini" or "local"
    resp = JSONResponse(out)
    resp.headers["X-AI-Used"] = "1" if via == "gemini" else "0"
    if model:
        resp.headers["X-AI-Model"] = model
    return resp

    # helper: parse file locally into "lines" (same shape as your upload/preview uses)
def _local_preview_lines_from_bytes(content: bytes, ext: str):
    rows = []
    if ext in ("csv", "tsv", "txt"):
        text = content.decode("utf-8", errors="ignore")
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
            reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        except Exception:
            reader = csv.DictReader(io.StringIO(text))
        for raw in reader:
            r = {}
            for k, v in raw.items():
                r[map_header(k)] = v
            rows.append(r)
    elif ext in ("xlsx", "xlsm", "xltx", "xltm"):
        if load_workbook is None:
            raise HTTPException(400, "openpyxl not installed (needed for .xlsx preview)")
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        hdr = [
            map_header(str(c) if c is not None else "")
            for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        for rr in ws.iter_rows(min_row=2, values_only=True):
            obj = {}
            for idx, val in enumerate(rr):
                key = hdr[idx] if idx < len(hdr) else f"col_{idx}"
                obj[key] = val
            rows.append(obj)
    else:
        raise HTTPException(400, "Unsupported file type. Use CSV or XLSX.")

    # shape into preview lines
    preview_lines = []
    for r in rows:
        if not any(filter(None, r.values())):
            continue
        preview_lines.append({
            "product_name_raw": (r.get("product_name_raw") or r.get("product") or r.get("name") or ""),
            "upc": r.get("upc"),
            "asin": r.get("asin"),
            "qty": int(re.sub(r"[^0-9]", "", str(r.get("qty") or 1)) or 1),
            "unit_cost": to_num(r.get("unit_cost")),
            "msrp": to_num(r.get("msrp")),
        })
    return preview_lines


# -----------------------------------------------------------------------------
# Config / env
# -----------------------------------------------------------------------------
PORT                = int(os.getenv("PORT", "3000"))  # default 3000 to match frontend
CONNECT_TIMEOUT     = float(os.getenv("CONNECT_TIMEOUT", "5"))
READ_TIMEOUT        = float(os.getenv("READ_TIMEOUT", "45"))

DIRECTUS_URL        = (os.getenv("DIRECTUS_URL") or "").rstrip("/")
DIRECTUS_TOKEN      = os.getenv("DIRECTUS_TOKEN", "")

DATABASE_URL        = os.getenv("DATABASE_URL")  # e.g. postgresql://app:app@localhost:5432/synergy_fastapi
if not DATABASE_URL:
    print("[api] WARNING: DATABASE_URL not set; DB routes will fail", flush=True)

EBAY_CLIENT_ID      = os.getenv("EBAY_CLIENT_ID", "")
EBAY_CLIENT_SECRET  = os.getenv("EBAY_CLIENT_SECRET", "")
EBAY_REFRESH_TOKEN  = os.getenv("EBAY_REFRESH_TOKEN", "")
EBAY_MARKETPLACE_ID = os.getenv("VITE_EBAY_MARKETPLACE_ID", "EBAY_US")

AVG_CACHE_TTL_MS    = int(os.getenv("AVG_CACHE_TTL_MS", str(6 * 60 * 60 * 1000)))  # 6h default
AI_FIRST            = (os.getenv("AI_FIRST", "0") == "1")

# -----------------------------------------------------------------------------
# HTTP session with retries
# -----------------------------------------------------------------------------
UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome Safari"
)
session = requests.Session()
session.headers.update({
    "User-Agent":      UA,
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
})
retry = Retry(
    total=3,
    backoff_factor=0.7,
    status_forcelist=(429, 500, 502, 503, 504),
    allowed_methods=frozenset({"GET", "POST", "PUT", "DELETE"}),
    raise_on_status=False,
)
session.mount("https://", HTTPAdapter(max_retries=retry))
session.mount("http://",  HTTPAdapter(max_retries=retry))

def http_get(
    url: str,
    cookie: str = "",
    timeout: Tuple[float, float] = (CONNECT_TIMEOUT, READ_TIMEOUT),
) -> requests.Response:
    headers = {}
    if cookie:
        headers["Cookie"] = cookie
    return session.get(url, headers=headers, timeout=timeout, allow_redirects=True)

# -----------------------------------------------------------------------------
# DB helpers
# -----------------------------------------------------------------------------
def db_conn():
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL not set")
    return psycopg2.connect(DATABASE_URL, sslmode=os.getenv("PGSSLMODE", "prefer"))

@contextmanager
def db():
    con = db_conn()
    cur = con.cursor(cursor_factory=RealDictCursor)
    try:
        yield con, cur
    finally:
        cur.close()
        con.close()

def to_num(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(re.sub(r"[^0-9.\-]", "", str(v)))
    except Exception:
        return None

def map_header(h: str) -> str:
    s = (h or "").strip().lower()

    # quantity
    if s == "qty" or "quantity" in s or re.search(r"\b(qty|quant)\b", s):
        return "qty"

    # prices
    if "stock image" in s or "price paid" in s or s == "cost" or "unit cost" in s or "our cost" in s:
        return "unit_cost"
    if "orig" in s and "retail" in s:
        return "msrp"
    if s == "msrp" or "list" in s:
        return "msrp"

    # identifiers
    if "upc" in s:
        return "upc"
    if "asin" in s:
        return "asin"

    # names / descriptions
    if "product" in s or "desc" in s or "name" in s or "listing title" in s or "title" in s:
        return "product_name_raw"

    # category
    if "category" in s:
        return "category_guess"

    # totals we explicitly ignore downstream (kept in case you want them later)
    if "total" in s and ("ret" in s or "ext" in s or "extended" in s):
        return "total_ignored"

    return h


def fmt_synergy(prefix: str, seq: int) -> str:
    return f"{prefix}-{str(int(seq)).zfill(4)}"

# -----------------------------------------------------------------------------
# eBay OAuth (user refresh)
# -----------------------------------------------------------------------------
_access_token: Optional[str] = None
_expires_at_ms: int = 0

EBAY_TOKEN_CACHE = {"token": None, "exp": 0}

# fallback marketplace (use VITE_* if EBAY_MARKETPLACE_ID not set)
EBAY_MARKETPLACE_ID = os.getenv("EBAY_MARKETPLACE_ID") or os.getenv("VITE_EBAY_MARKETPLACE_ID") or "EBAY_US"

def _collapse_scopes(s: str) -> str:
    # turn any multi-line / extra spaces into a single space-separated string
    return " ".join((s or "").split())
def get_ebay_token(force: bool = False) -> str | None:
    now = time.time()
    if not force and EBAY_TOKEN_CACHE["token"] and EBAY_TOKEN_CACHE["exp"] - 60 > now:
        return EBAY_TOKEN_CACHE["token"]

    cid = os.getenv("EBAY_CLIENT_ID")
    cs  = os.getenv("EBAY_CLIENT_SECRET")
    rt  = os.getenv("EBAY_REFRESH_TOKEN")
    scopes = _collapse_scopes(os.getenv("EBAY_USER_SCOPES", ""))

    if not (cid and cs and rt):
        print("[EBAY] missing EBAY_CLIENT_ID/EBAY_CLIENT_SECRET/EBAY_REFRESH_TOKEN", flush=True)
        return None

    auth = base64.b64encode(f"{cid}:{cs}".encode()).decode()
    resp = session.post(
        "https://api.ebay.com/identity/v1/oauth2/token",
        headers={
            "Authorization": f"Basic {auth}",
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
        },
        data={
            "grant_type": "refresh_token",
            "refresh_token": rt,
            "scope": scopes,
        },
        timeout=(CONNECT_TIMEOUT, READ_TIMEOUT),
    )
    if resp.status_code != 200:
        print("[EBAY] oauth refresh failed:", resp.status_code, resp.text[:400], flush=True)
        return None

    data = resp.json()
    tok = data.get("access_token")
    ttl = int(data.get("expires_in", 7200))
    if not tok:
        print("[EBAY] oauth refresh returned no access_token", flush=True)
        return None

    EBAY_TOKEN_CACHE["token"] = tok
    EBAY_TOKEN_CACHE["exp"]   = now + ttl
    return tok
def _ebay_access_token_from_refresh() -> str:
    tok = get_ebay_token(force=True)
    if not tok:
        raise HTTPException(status_code=502, detail="eBay token error: no_token")
    return tok

def _parse_legacy_item_id(url_or_id: str) -> str | None:
    s = (url_or_id or "").strip()
    # Already an ID?
    if s.isdigit() and len(s) >= 9:
        return s
    # Common URL shapes: .../itm/<id> or ...?item=<id>
    import re
    m = re.search(r"/itm/(\d{9,})", s)
    if m: return m.group(1)
    m = re.search(r"[?&](?:item|iid|itemId)=(\d{9,})", s, re.I)
    if m: return m.group(1)
    return None
def _parse_legacy_from_url(url: str) -> str | None:
    if not url:
        return None
    # supports .../itm/326698728632 or ...?item=326698728632
    m = re.search(r"/itm/(\d+)", url)
    if m:
        return m.group(1)
    m = re.search(r"[?&](?:item|itemid|ItemID)=(\d+)", url, re.I)
    return m.group(1) if m else None

def _legacy_for_synergy(cur, synergy_id: str) -> str | None:
    # (1) ebay_links
    cur.execute("SELECT legacy_item_id FROM ebay_links WHERE synergy_id=%s", (synergy_id,))
    row = cur.fetchone()
    if row and row.get("legacy_item_id"):
        return str(row["legacy_item_id"])

    # (2) external_listings (preferred) â€” assumes platform='ebay'
    try:
        cur.execute("""
            SELECT ebay_legacy_id, url
              FROM external_listings
             WHERE synergy_id=%s AND platform='ebay'
             ORDER BY updated_at DESC NULLS LAST
             LIMIT 1
        """, (synergy_id,))
        row = cur.fetchone()
        if row:
            if row.get("ebay_legacy_id"):
                return str(row["ebay_legacy_id"])
            if row.get("url"):
                lid = _parse_legacy_item_id(row["url"])
                if lid:
                    return lid
    except Exception:
        pass

    # (3) inventory_items.ebay_item_url
    try:
        cur.execute("""
            SELECT ebay_item_url
              FROM inventory_items
             WHERE synergy_id=%s
             LIMIT 1
        """, (synergy_id,))
        row = cur.fetchone()
        if row and row.get("ebay_item_url"):
            lid = _parse_legacy_item_id(row["ebay_item_url"])
            if lid:
                return lid
    except Exception:
        pass

    return None
    
# -----------------------------------------------------------------------------
# FastAPI app + CORS
# -----------------------------------------------------------------------------
app = FastAPI(title="Synergy API")

origins = [
    "http://localhost:8081",
    "http://127.0.0.1:8081",
    "http://10.28.1.37:8081",
    "http://synergy.lan:8081",
    "http://synergy.local:8081",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.include_router(early_router)
app.include_router(label_inventory_router)

# -----------------------------------------------------------------------------
# Health
# -----------------------------------------------------------------------------
@app.get("/health")
def health():
    return {"ok": True, "ts": time.time()}

# -----------------------------------------------------------------------------
# Minimal users (auth-lite)
# -----------------------------------------------------------------------------
class NewUser(BaseModel):
    name: str
    initials: str | None = None
    role: str = "Tester"
    active: bool = True

class LinkEbayBody(BaseModel):
    synergyId: str
    ebayUrl: str
    # optional hints; kept for compatibility
    inventoryItemId: int | None = None
    poLineSynergyId: str | None = None
    synergyCode: str | None = None
    price: float | None = None
    currency: str | None = None
    postedBy: str | None = None
    sku: str | None = None

class EbaySyncBody(BaseModel):
    synergyId: str
    ebayUrl: HttpUrl

EBAY_ITEM_RE = re.compile(
    r"(?:/itm/(?:[^/?#]*/)?(?P<itm>\d{9,14}))|"
    r"(?:[?&#](?:item|nid)=(?P<q>\d{9,14}))",
    re.IGNORECASE,
)

def _parse_ebay_legacy_id(url: str) -> str | None:
    if not url:
        return None
    m = EBAY_ITEM_RE.search(url)
    return (m.group("itm") or m.group("q")) if m else None

# Alias the name you call later:
def _parse_legacy_item_id(url: str) -> str | None:
    return _parse_ebay_legacy_id(url)

@app.get("/auth/users")
def list_users(role: Optional[str] = Query(None), active: bool = Query(True)):
    # return role as Title-Case so the UI's equality check passes
    sql = ["SELECT id, name, initials, initcap(role) AS role, active FROM app_users WHERE 1=1"]
    params: List[Any] = []
    if role:
        # case-insensitive filter
        sql.append("AND lower(role) = lower(%s)")
        params.append(role)
    if active:
        sql.append("AND active = TRUE")
    sql.append("ORDER BY name ASC")
    with db() as (con, cur):
        cur.execute(" ".join(sql), params)
        rows = cur.fetchall()  # RealDictCursor -> list[dict]
        return rows

@app.post("/auth/users")
def create_user(body: dict):
    name = (body.get("name") or "").strip()
    initials = (body.get("initials") or "").strip().upper() or "".join([p[:1] for p in name.split()])[:4].upper()
    role = (body.get("role") or "poster").strip().lower()  # store lowercase
    if role == "manager":
        raise HTTPException(status_code=403, detail="Managers must be created by an administrator")
    with db() as (con, cur):
        cur.execute(
            "INSERT INTO public.app_users(name, initials, role, active) VALUES(%s,%s,%s,TRUE) "
            "RETURNING id, name, initials, initcap(role) AS role, active",
            (name, initials, role),
        )
        row = cur.fetchone()
        con.commit()
    return {"id": row[0], "name": row[1], "initials": row[2], "role": row[3], "active": row[4]}


# -----------------------------------------------------------------------------
# Cookie + Warmup
# -----------------------------------------------------------------------------
# 1) Start an AI preview job. Returns {job_id} immediately (202 Accepted).

@app.post("/imports/ai-preview-jobs")
async def start_ai_preview_job(
    vendor_id: str = Form(...),
    file: UploadFile = File(...),
    expand_units: bool = Form(False),
    require_ai: bool = Form(False),
    limit_rows: int = Form(500),
):
    if not vendor_id:
        raise HTTPException(400, "vendor_id is required")
    raw = await file.read()
    filename = file.filename or "upload"
    ext = (filename.rsplit(".", 1)[-1] or "").lower()

    # RQ sometimes chokes on large raw bytes; base64 them to be safe
    raw_b64 = base64.b64encode(raw).decode("ascii")

    job_id = str(uuid.uuid4())
    q = rq.Queue("ai", connection=_redis_conn())
    # IMPORTANT: import path for jobs.ai_preview_job must be real on worker
    q.enqueue(
        "jobs.ai_preview_job",
        job_id,
        vendor_id,
        raw_b64,        # <-- base64 string
        filename,
        ext,
        expand_units,
        limit_rows,
        require_ai,
        os.getenv("DATABASE_URL"),
        os.getenv("REDIS_URL"),
        job_id=job_id,
        description=f"AI preview for vendor {vendor_id}",
    )
    return {"ok": True, "job_id": job_id}

def _ebay_get_orders_range(start: datetime, end: datetime, limit: int = 200):
    """Generator over /sell/fulfillment/v1/order within a time window."""
    token = get_ebay_token()
    if not token:
        raise HTTPException(status_code=401, detail="no_token")

    url = "https://api.ebay.com/sell/fulfillment/v1/order"
    headers = {
        "Authorization": f"Bearer {token}",
        "X-EBAY-C-MARKETPLACE-ID": EBAY_MARKETPLACE_ID,
        "Accept": "application/json",
    }
    offset = 0
    while True:
        params = {
            "filter": f"creationdate:[{_iso(start)}..{_iso(end)}]",
            "limit": str(limit),
            "offset": str(offset),
        }
        resp = session.get(url, headers=headers, params=params, timeout=(CONNECT_TIMEOUT, READ_TIMEOUT))
        if resp.status_code != 200:
            raise HTTPException(status_code=resp.status_code, detail=f"ebay_fulfillment_error: {resp.text[:300]}")
        data = resp.json()
        for o in data.get("orders", []) or []:
            yield o
        if not data.get("next"):
            break
        offset += limit
def _fulfillment_call(access: str, url: str):
    r = requests.get(
        url,
        headers={
            "Authorization": f"Bearer {access}",
            "X-EBAY-C-MARKETPLACE-ID": EBAY_MARKETPLACE_ID,
        },
        timeout=30,
    )
    r.raise_for_status()
    return r.json()

def _fulfillment_orders_in_window(access: str, start_iso: str, end_iso: str):
    base = "https://api.ebay.com/sell/fulfillment/v1/order"
    orders = []
    for filt in (f"creationdate:[{start_iso}..{end_iso}]",
                 f"lastmodifieddate:[{start_iso}..{end_iso}]"):
        offset = 0
        while True:
            url = f"{base}?filter={quote(filt, safe='[]:,')}&limit=200&offset={offset}"
            j = _fulfillment_call(access, url)
            orders.extend(j.get("orders", []) or [])
            total = j.get("total", 0) or 0
            if offset + 200 >= total:
                break
            offset += 200
    return orders

def _extract_matches(orders, legacy_or_item_id: str):
    id_str = str(legacy_or_item_id)
    hits = []
    for o in orders:
        created = o.get("creationDate")
        for li in (o.get("lineItems") or []):
            if str(li.get("legacyItemId") or "") == id_str or str(li.get("itemId") or "") == id_str:
                qty = li.get("quantity", 1) or 1
                hits.append({"orderId": o.get("orderId"), "creationDate": created, "quantity": qty})
    hits.sort(key=lambda s: s["creationDate"] or "", reverse=True)
    return hits

def _trading_get_item(user_access_token: str, legacy_item_id: str) -> dict:
    xml = f"""<?xml version="1.0" encoding="utf-8"?>
<GetItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">
  <ErrorLanguage>en_US</ErrorLanguage>
  <WarningLevel>High</WarningLevel>
  <ItemID>{legacy_item_id}</ItemID>
  <DetailLevel>ReturnAll</DetailLevel>
  <IncludeSelector>Details,ItemSpecifics,ShippingCosts,Variations</IncludeSelector>
</GetItemRequest>"""

    headers = {
        "Content-Type": "text/xml",
        "X-EBAY-API-CALL-NAME": "GetItem",
        "X-EBAY-API-SITEID": "0",
        "X-EBAY-API-COMPATIBILITY-LEVEL": "967",
        "X-EBAY-API-IAF-TOKEN": user_access_token,
    }

    r = requests.post(EBAY_TRADING_ENDPOINT, data=xml.encode("utf-8"), headers=headers, timeout=25)
    r.raise_for_status()

    root = ET.fromstring(r.text)
    item = root.find(f".//{EBAY_NS}Item")
    if item is None:
        raise RuntimeError("Trading.GetItem: <Item> not found")

    seller = _el_text(item.find(f"./{EBAY_NS}Seller/{EBAY_NS}UserID"))
    status = _el_text(item.find(f"./{EBAY_NS}SellingStatus/{EBAY_NS}ListingStatus")) \
          or _el_text(item.find(f"./{EBAY_NS}ListingStatus"))

    qty_sold_txt = _first_non_null(
        _el_text(item.find(f"./{EBAY_NS}QuantitySold")),
        _el_text(item.find(f"./{EBAY_NS}SellingStatus/{EBAY_NS}QuantitySold")),
        "0",
    )
    try:
        sold_lifetime = int(qty_sold_txt or "0")
    except ValueError:
        sold_lifetime = 0

    # --- price candidates (Trading) ---
    # Primary
    p1, c1 = _price_from_el(item.find(f"./{EBAY_NS}SellingStatus/{EBAY_NS}CurrentPrice"))
    # Converted (shows up often)
    p1c, c1c = _price_from_el(item.find(f"./{EBAY_NS}SellingStatus/{EBAY_NS}ConvertedCurrentPrice"))
    # Listing-level
    p2, c2 = _price_from_el(item.find(f"./{EBAY_NS}StartPrice"))
    p2c, c2c = _price_from_el(item.find(f"./{EBAY_NS}ConvertedStartPrice"))
    p3, c3 = _price_from_el(item.find(f"./{EBAY_NS}BuyItNowPrice"))
    p3c, c3c = _price_from_el(item.find(f"./{EBAY_NS}ConvertedBuyItNowPrice"))

    price, currency = None, None
    for p, c in ((p1, c1), (p1c, c1c), (p2, c2), (p2c, c2c), (p3, c3), (p3c, c3c)):
        if p is not None:
            price, currency = p, c
            break

    # Variations: choose MIN StartPrice/CurrentPrice as a single â€œfromâ€ price
    if price is None:
        variations = item.find(f"./{EBAY_NS}Variations")
        if variations is not None:
            prices = []
            cur = None
            for var in variations.findall(f"./{EBAY_NS}Variation"):
                vp, vc = _price_from_el(var.find(f"./{EBAY_NS}StartPrice"))
                if vp is None:
                    vp, vc = _price_from_el(var.find(f"./{EBAY_NS}SellingStatus/{EBAY_NS}CurrentPrice"))
                if vp is not None:
                    prices.append(vp)
                    if cur is None:
                        cur = vc
            if prices:
                price, currency = (min(prices), cur)

    return {
        "sold_lifetime": sold_lifetime,
        "seller": seller,
        "status": status,
        "price": price,
        "currency": currency,
    }
def _browse_get_price(user_access_token: str, legacy_item_id: str) -> tuple[float | None, str | None]:
    # REST item id format: v1|{legacy}|0
    rest_id = f"v1|{legacy_item_id}|0"
    url = f"{EBAY_BROWSE_ENDPOINT}/item/{requests.utils.quote(rest_id, safe='')}"
    headers = {
        "Authorization": f"Bearer {user_access_token}",
        "Accept": "application/json",
    }
    r = requests.get(url, headers=headers, timeout=15)
    if r.status_code == 404:
        return (None, None)
    r.raise_for_status()
    j = r.json()
    # Expect: { "price": { "value": "199.99", "currency": "USD" }, ... }
    price_obj = (j or {}).get("price") or {}
    val = price_obj.get("value")
    cur = price_obj.get("currency")
    try:
        return (float(val), cur)
    except (TypeError, ValueError):
        return (None, cur)

def _browse_get_price_app(legacy_item_id: str) -> tuple[float | None, str | None, str]:
    """
    Use Buy Browse with app token.
    Tries get_item_by_legacy_id (simpler than v1|â€¦|0 form).
    Returns (price, currency, price_source).
    """
    token = _ebay_app_access_token(["https://api.ebay.com/oauth/api_scope/buy.browse.readonly"])
    url = f"{EBAY_BROWSE_ENDPOINT}/item/get_item_by_legacy_id"
    params = {"legacy_item_id": legacy_item_id}
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "X-EBAY-C-MARKETPLACE-ID": "EBAY_US",
    }
    r = requests.get(url, params=params, headers=headers, timeout=20)
    if r.status_code == 403:
        # Insufficient permissions (app not approved for Buy APIs)
        return (None, None, "browse.denied")
    if r.status_code == 404:
        return (None, None, "browse.404")
    r.raise_for_status()
    j = r.json()
    pr = (j or {}).get("price") or {}
    val, cur = pr.get("value"), pr.get("currency")
    try:
        return (float(val), cur, "browse.item.price")
    except (TypeError, ValueError):
        return (None, cur, "browse.item.price.invalid")
    
def summarize_for_item(access: str, legacy_item_id: str, months_back: int = 24):
    # 1) lifetime sold via Trading (works even if very old)
    ti = _trading_get_item(access, legacy_item_id)  # Seller / Status / QuantitySold

    # 2) sweep Fulfillment in monthly slices up to N months to find "last sold"
    end = datetime.now(timezone.utc).replace(microsecond=0)
    start = end - timedelta(days=months_back * 30)
    cursor = start
    newest_hit = None
    total_qty = 0

    while cursor < end:
        slice_end = min(cursor + timedelta(days=30), end)
        orders = _fulfillment_orders_in_window(
            access,
            cursor.isoformat().replace("+00:00", "Z"),
            slice_end.isoformat().replace("+00:00", "Z"),
        )
        hits = _extract_matches(orders, legacy_item_id)
        if hits:
            total_qty += sum(h["quantity"] for h in hits)
            if not newest_hit or hits[0]["creationDate"] > newest_hit["creationDate"]:
                newest_hit = hits[0]
        cursor = slice_end + timedelta(seconds=1)

    return {
        "seller": ti["seller"],
        "listingStatus": ti["status"],
        "soldCount": ti["quantitySold"],                # lifetime
        "lastSoldAt": newest_hit["creationDate"] if newest_hit else None,  # recent window
        "recentQtyInWindow": total_qty,
        "note": None if newest_hit else "No orders in Fulfillment retention window (~2y max). Lifetime sold from Trading.",
    }

def _ebay_user_access_token_from_refresh() -> str:
    cid = os.getenv("EBAY_CLIENT_ID")
    csec = os.getenv("EBAY_CLIENT_SECRET")
    rt = os.getenv("EBAY_REFRESH_TOKEN")
    scopes = os.getenv(
        "EBAY_USER_SCOPES",
        "https://api.ebay.com/oauth/api_scope/sell.inventory.readonly https://api.ebay.com/oauth/api_scope/sell.fulfillment.readonly"
    )
    if not (cid and csec and rt):
        raise RuntimeError("EBAY oauth env missing (CLIENT_ID/SECRET/REFRESH_TOKEN)")

    basic = base64.b64encode(f"{cid}:{csec}".encode()).decode()
    r = requests.post(
        "https://api.ebay.com/identity/v1/oauth2/token",
        headers={"Authorization": f"Basic {basic}",
                 "Content-Type": "application/x-www-form-urlencoded"},
        data={"grant_type": "refresh_token",
              "refresh_token": rt,
              "scope": scopes},
        timeout=30,
    )
    r.raise_for_status()
    tok = r.json().get("access_token")
    if not tok:
        raise RuntimeError(f"oauth refresh failed: {r.text}")
    return tok

def _trading_get_item(access_token: str, legacy_item_id: str) -> dict:
    xml = f"""<?xml version="1.0" encoding="utf-8"?>
<GetItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">
  <ItemID>{legacy_item_id}</ItemID>
  <DetailLevel>ReturnAll</DetailLevel>
</GetItemRequest>"""
    r = requests.post(
        "https://api.ebay.com/ws/api.dll",
        headers={
            "X-EBAY-API-CALL-NAME": "GetItem",
            "X-EBAY-API-SITEID": "0",
            "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
            "X-EBAY-API-IAF-TOKEN": access_token,
            "Content-Type": "text/xml",
        },
        data=xml, timeout=30
    )
    r.raise_for_status()
    t = r.text
    def tag(name):
        m = re.search(fr"<{name}>(.*?)</{name}>", t)
        return m.group(1) if m else None
    return {
        "seller": tag("UserID"),
        "status": tag("ListingStatus"),
        "item_id": tag("ItemID"),
        "sold_lifetime": int((tag("QuantitySold") or "0")),
    }

def _trading_get_last_sold(access_token: str, legacy_item_id: str, months_back: int = 24) -> str | None:
    end = datetime.now(timezone.utc).replace(microsecond=0)
    start = end - timedelta(days=months_back * 30)
    newest = None
    cur = start
    while cur < end:
        nxt = min(cur + timedelta(days=30), end)
        xml = f"""<?xml version="1.0" encoding="utf-8"?>
<GetItemTransactionsRequest xmlns="urn:ebay:apis:eBLBaseComponents">
  <ItemID>{legacy_item_id}</ItemID>
  <DetailLevel>ReturnAll</DetailLevel>
  <ModTimeFrom>{cur.strftime('%Y-%m-%dT%H:%M:%SZ')}</ModTimeFrom>
  <ModTimeTo>{nxt.strftime('%Y-%m-%dT%H:%M:%SZ')}</ModTimeTo>
</GetItemTransactionsRequest>"""
        r = requests.post(
            "https://api.ebay.com/ws/api.dll",
            headers={
                "X-EBAY-API-CALL-NAME": "GetItemTransactions",
                "X-EBAY-API-SITEID": "0",
                "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
                "X-EBAY-API-IAF-TOKEN": access_token,
                "Content-Type": "text/xml",
            },
            data=xml, timeout=30
        )
        r.raise_for_status()
        for m in re.finditer(r"<CreatedDate>(.*?)</CreatedDate>", r.text):
            cd = m.group(1)
            if not newest or cd > newest:
                newest = cd
        cur = nxt + timedelta(seconds=1)
    return newest

def _parse_legacy_from_url(url: str) -> str | None:
    m = re.search(r"/itm/(?:[^/]+/)?(\d{9,})", (url or ""))
    return m.group(1) if m else None
def _summarize_sales_for_legacy_id(legacy_item_id: str, days_back: int = 365) -> dict:
    """
    Look up orders in Sell Fulfillment and summarize sales for a legacy item id.
    Always returns JSON-serializable types.
    """
    token = get_ebay_token()  # uses your env scopes
    if not token:
        raise HTTPException(status_code=502, detail="no_token")

    start = (datetime.now(timezone.utc) - timedelta(days=days_back)).isoformat(timespec="seconds").replace("+00:00", "Z")
    end   = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")

    base = "https://api.ebay.com/sell/fulfillment/v1/order"
    url = f"{base}?filter=creationdate:[{start}..{end}]&limit=100"

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "X-EBAY-C-MARKETPLACE-ID": os.getenv("EBAY_MARKETPLACE_ID") or os.getenv("VITE_EBAY_MARKETPLACE_ID") or "EBAY_US",
    }

    sales = []
    s = requests.Session()

    while True:
        r = s.get(url, headers=headers, timeout=30)
        if r.status_code != 200:
            # Return a clean JSON error to the client (so jq doesn't choke)
            try:
                body = r.json()
            except Exception:
                body = {"raw": r.text[:400]}
            raise HTTPException(status_code=502, detail={"ebay_status": r.status_code, "ebay_error": body})

        data = r.json() or {}
        for order in data.get("orders", []):
            # payment date (if any)
            pay_date = None
            ps = (order.get("paymentSummary") or {}).get("payments") or []
            if ps and isinstance(ps, list):
                pay_date = (ps[0].get("paymentDate") or None)

            for li in order.get("lineItems", []):
                if str(li.get("legacyItemId") or "") == str(legacy_item_id):
                    sales.append({
                        "orderId": order.get("orderId"),
                        "creationDate": order.get("creationDate"),
                        "paymentDate": pay_date,
                        "quantity": li.get("quantity") or 0,
                    })

        # pagination
        next_href = None
        for link in data.get("links", []):
            if link.get("rel") == "next" and link.get("href"):
                next_href = link["href"]
                break
        if not next_href:
            break
        url = next_href

    # summarize
    def _pick_dt(sale):
        return sale.get("paymentDate") or sale.get("creationDate")

    last_sold_at = None
    if sales:
        last_sold_at = max((_pick_dt(s) for s in sales if _pick_dt(s)), default=None)

    sold_count = sum(int(s.get("quantity") or 0) for s in sales)

    return {
        "ok": True,
        "legacyItemId": str(legacy_item_id),
        "soldCount": sold_count,
        "lastSoldAt": last_sold_at,   # ISO string from eBay (already JSON-safe)
        "sales": sales,               # array of simple dicts
    }

class AssocBody(BaseModel):
    synergyId: str
    ebayUrl: str

@app.post("/integrations/ebay/associate-url")
def ebay_associate_url(body: AssocBody):
    legacy = _parse_legacy_item_id(body.ebayUrl)  # you already have this helper
    if not legacy:
        raise HTTPException(status_code=400, detail="Could not parse legacy item id from URL")
    with db() as (con, cur):
        cur.execute("""
            INSERT INTO ebay_links (synergy_id, ebay_url, legacy_item_id, updated_at)
            VALUES (%s, %s, %s, now())
            ON CONFLICT (synergy_id)
            DO UPDATE SET ebay_url = EXCLUDED.ebay_url,
                          legacy_item_id = EXCLUDED.legacy_item_id,
                          updated_at = now();
        """, (body.synergyId, body.ebayUrl, legacy))
        con.commit()
    return {"ok": True, "synergyId": body.synergyId, "legacyItemId": legacy}


def _xml_text(root, path):
    el = root.find(path, EBAY_NS)
    return el.text if el is not None else None

def _fetch_price_via_trading(access_token: str, legacy_item_id: str) -> tuple[float | None, str | None, str | None]:
    """
    GetItem-based price resolver:
      - Multi-variation: min(Variation.StartPrice)
      - Fixed price: Item.StartPrice
      - Auction/fallback: SellingStatus.CurrentPrice
    Returns: (price, currency, source_tag)
    """
    import requests
    import xml.etree.ElementTree as ET

    xml = f"""<?xml version="1.0" encoding="utf-8"?>
<GetItemRequest xmlns="urn:ebay:apis:eBLBaseComponents">
  <ItemID>{legacy_item_id}</ItemID>
  <DetailLevel>ReturnAll</DetailLevel>
</GetItemRequest>"""

    r = requests.post(
        "https://api.ebay.com/ws/api.dll",
        headers={
            "X-EBAY-API-CALL-NAME": "GetItem",
            "X-EBAY-API-SITEID": "0",
            "X-EBAY-API-COMPATIBILITY-LEVEL": "1193",
            "X-EBAY-API-IAF-TOKEN": access_token,  # OAuth2 user token
            "Content-Type": "text/xml",
        },
        data=xml.encode("utf-8"),
        timeout=25,
    )
    r.raise_for_status()
    root = ET.fromstring(r.text)

    # 1) Variations â†’ choose MIN(StartPrice)
    var_prices: list[tuple[float, str | None]] = []
    for sp in root.findall(".//e:Variations/e:Variation/e:StartPrice", EBAY_NS):
        txt = (sp.text or "").strip()
        if not txt:
            continue
        try:
            var_prices.append((float(txt), sp.attrib.get("currencyID")))
        except Exception:
            pass
    if var_prices:
        price, currency = sorted(var_prices, key=lambda t: t[0])[0]
        return price, currency, "trading.getItem.variations.startPrice"

    # Listing type (FixedPriceItem vs Chinese, etc.)
    listing_type_el = root.find(".//e:ListingType", EBAY_NS)
    listing_type = (listing_type_el.text or "").strip() if listing_type_el is not None else ""

    # 2) Fixed price listing â†’ Item.StartPrice
    if listing_type == "FixedPriceItem":
        sp = root.find(".//e:StartPrice", EBAY_NS)
        if sp is not None and (sp.text or "").strip():
            try:
                return float(sp.text), sp.attrib.get("currencyID"), "trading.getItem.startPrice"
            except Exception:
                pass

    # 3) Auction / fallback â†’ SellingStatus.CurrentPrice
    cp = root.find(".//e:SellingStatus/e:CurrentPrice", EBAY_NS)
    if cp is not None and (cp.text or "").strip():
        try:
            return float(cp.text), cp.attrib.get("currencyID"), "trading.getItem.sellingStatus.currentPrice"
        except Exception:
            pass

    # No price fields found
    return None, None, None

def _safe_float(v):
    try:
        return float(str(v).replace(",", ""))
    except Exception:
        return None

def _html_get_price_from_item_url(ebay_url: str):
    """
    Best-effort: scrape public item page for price.
    Returns (price_value, currency or None, source_str) or (None, None, "page.none").
    """
    if not ebay_url:
        return (None, None, "page.no_url")
    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }
        r = requests.get(ebay_url, headers=headers, timeout=12)
        if r.status_code >= 400:
            return (None, None, f"page.{r.status_code}")
        html = r.text or ""

        # 1) Look for JSON with currentPrice like: "currentPrice":{"value":123.45,"currency":"USD"}
        m = re.search(r'"currentPrice"\s*:\s*\{\s*"value"\s*:\s*([0-9\.,]+)\s*,\s*"currency"\s*:\s*"([A-Z]{3})"', html)
        if m:
            val = _safe_float(m.group(1))
            cur = m.group(2)
            if val is not None:
                return (val, cur, "page.currentPrice")

        # 2) Try "price":{"value":...,"currency":"USD"}
        m = re.search(r'"price"\s*:\s*\{\s*"value"\s*:\s*([0-9\.,]+)\s*,\s*"currency"\s*:\s*"([A-Z]{3})"', html)
        if m:
            val = _safe_float(m.group(1))
            cur = m.group(2)
            if val is not None:
                return (val, cur, "page.price_obj")

        # 3) OG meta tags (less reliable on eBay, but try):
        m = re.search(r'<meta\s+property="og:price:amount"\s+content="([0-9\.,]+)"', html)
        if m:
            val = _safe_float(m.group(1))
            if val is not None:
                cur = None
                m2 = re.search(r'<meta\s+property="og:price:currency"\s+content="([A-Z]{3})"', html)
                if m2:
                    cur = m2.group(1)
                return (val, cur, "page.og")

        # 4) JSON-LD block with "price": "123.45"
        m = re.search(r'<script type="application/ld\+json">(.+?)</script>', html, re.S)
        if m:
            try:
                data = json.loads(m.group(1))
                # handle list or dict
                cands = data if isinstance(data, list) else [data]
                for d in cands:
                    offer = d.get("offers") if isinstance(d, dict) else None
                    if isinstance(offer, dict):
                        val = _safe_float(offer.get("price"))
                        cur = offer.get("priceCurrency")
                        if val is not None:
                            return (val, cur, "page.ldjson")
            except Exception:
                pass

        return (None, None, "page.none")
    except Exception:
        return (None, None, "page.error")


def _ebay_app_access_token(scopes=None) -> str:
    """
    App token for Buy/Browse (client_credentials). Only used as a *last* resort.
    Any error here is swallowed by the caller.
    """
    scope_str = " ".join(scopes or ["https://api.ebay.com/oauth/api_scope/buy.browse.readonly"])
    auth = base64.b64encode(f"{EBAY_CLIENT_ID}:{EBAY_CLIENT_SECRET}".encode()).decode()
    headers = {"Authorization": f"Basic {auth}", "Content-Type": "application/x-www-form-urlencoded"}
    data = {"grant_type": "client_credentials", "scope": scope_str}
    r = requests.post(EBAY_OAUTH_TOKEN_URL, headers=headers, data=data, timeout=12)
    r.raise_for_status()
    return r.json()["access_token"]


def _browse_get_price_app_safe(legacy_item_id: str):
    """
    Safe wrapper around Buy/Browse with APP token â€” NEVER raises; returns (price, currency, source).
    """
    if not EBAY_CLIENT_ID or not EBAY_CLIENT_SECRET:
        return (None, None, "browse.app.unset")
    try:
        token = _ebay_app_access_token(["https://api.ebay.com/oauth/api_scope/buy.browse.readonly"])
        url = f"{EBAY_BROWSE_ENDPOINT}/item/get_item_by_legacy_id"
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "X-EBAY-C-MARKETPLACE-ID": "EBAY_US",
        }
        r = requests.get(url, params={"legacy_item_id": legacy_item_id}, headers=headers, timeout=12)
        if r.status_code == 403:
            return (None, None, "browse.denied")
        if r.status_code == 404:
            return (None, None, "browse.404")
        if r.status_code >= 400:
            return (None, None, f"browse.{r.status_code}")
        j = r.json()
        pr = (j or {}).get("price") or {}
        val, cur = pr.get("value"), pr.get("currency")
        val = _safe_float(val)
        if val is not None:
            return (val, cur, "browse.item.price")
        return (None, cur, "browse.item.missing")
    except Exception:
        return (None, None, "browse.app.error")
@app.post("/integrations/ebay/refresh-sold-when/{synergy_id}")
def ebay_refresh_sold_when(synergy_id: str, days: int = 730):
    try:
        # 1) Ensure we have legacy item id
        with psycopg2.connect(DATABASE_URL) as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    "SELECT synergy_id, ebay_url, legacy_item_id FROM ebay_links WHERE synergy_id=%s",
                    (synergy_id,),
                )
                link = cur.fetchone()
                if not link:
                    return JSONResponse(status_code=400, content={"ok": False, "error": "no_link_for_synergy"})

                ebay_url = link["ebay_url"]
                legacy = (link["legacy_item_id"] or "")
                if not legacy:
                    legacy = _parse_legacy_from_url(ebay_url or "")
                    if not legacy:
                        return JSONResponse(status_code=400, content={"ok": False, "error": "no_legacy_item_for_synergy"})
                    cur.execute(
                        "UPDATE ebay_links SET legacy_item_id=%s, updated_at=now() WHERE synergy_id=%s RETURNING synergy_id, ebay_url, legacy_item_id",
                        (legacy, synergy_id),
                    )
                    link = cur.fetchone()
                    conn.commit()

        # 2) Trading calls (same approach you already use)
        access = _ebay_user_access_token_from_refresh()

        ti = _trading_get_item(access, legacy)  # your existing helper
        sold_lifetime = ti.get("sold_lifetime") or ti.get("quantitySold") or 0
        seller = ti.get("seller")
        status = ti.get("status")

        # Price via Trading/GetItem (no Browse)
        price, currency, price_src = None, None, None
        try:
            price, currency, price_src = _fetch_price_via_trading(access, legacy)
        except Exception:
            price, currency, price_src = None, None, None

        # 3) Last sold within retention window (unchanged)
        months_back = min(max(days // 30, 1), 24)
        try:
            last_sold = _trading_get_last_sold(access, legacy, months_back=months_back)
        except Exception:
            last_sold = None

        # 4) Persist counters/timestamps (leave price persistence to your row patcher if you want)
        with psycopg2.connect(DATABASE_URL) as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    UPDATE ebay_links
                       SET sold_count = %s,
                           last_sold_at = %s,
                           updated_at = now()
                     WHERE synergy_id = %s
                 RETURNING synergy_id, ebay_url, legacy_item_id, last_sold_at, sold_count
                """, (sold_lifetime, last_sold, synergy_id))
                saved = cur.fetchone()
                conn.commit()

        src_map = {
            "soldCount": "trading.getItem.quantitySold",
            "lastSoldAt": "trading.getItemTransactions" if last_sold else None,
        }
        if price_src:
            src_map["price"] = price_src

        return {
            "ok": True,
            "synergyId": synergy_id,
            "legacyItemId": legacy,
            "soldCount": sold_lifetime,
            "lastSoldAt": last_sold,
            "sales": [],
            "link": saved,
            "seller": seller,
            "listingStatus": status,
            "price": price,
            "currency": currency or "USD",
            "source": src_map,
        }

    except requests.HTTPError as e:
        return JSONResponse(
            status_code=502,
            content={"ok": False, "error": "http_error",
                     "detail": {"status": e.response.status_code, "body": e.response.text[:2000]}},
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": "server_error", "detail": str(e)})

class BulkRefreshBody(BaseModel):
    synergyIds: List[str]
    days: int = 365

@app.post("/integrations/ebay/refresh-sold-when/bulk")
def ebay_refresh_bulk(body: BulkRefreshBody):
    out = {}
    for sid in body.synergyIds:
        try:
            res = ebay_refresh_sold_when.__wrapped__(sid, days=body.days)  # call underlying fn
            out[sid] = {"ok": True, "lastSoldAt": res["lastSoldAt"], "soldCount": res["soldCount"]}
        except HTTPException as e:
            out[sid] = {"ok": False, "error": e.detail}
    return out


@app.get("/integrations/ebay/status/{synergy_id}")
def ebay_status_one(synergy_id: str):
    with db() as (con, cur):
        cur.execute("""
          SELECT synergy_id, ebay_url, legacy_item_id, last_sold_at, sold_count, updated_at
            FROM ebay_links WHERE synergy_id = %s
        """, (synergy_id,))
        row = cur.fetchone()
        return row or {}

@app.get("/integrations/ebay/status")
def ebay_status_bulk(ids: str = Query(..., description="comma-separated synergyIds")):
    wanted = [s.strip() for s in ids.split(",") if s.strip()]
    with db() as (con, cur):
        cur.execute("""
          SELECT synergy_id, ebay_url, legacy_item_id, last_sold_at, sold_count, updated_at
            FROM ebay_links WHERE synergy_id = ANY(%s)
        """, (wanted,))
        rows = {r["synergy_id"]: r for r in cur.fetchall()}
    # include empties so UI can map deterministically
    return {sid: rows.get(sid) for sid in wanted}
    
@app.post("/imports/ai-commit-jobs")
async def start_ai_commit_job(
    po_number: str = Form(...),
    file: UploadFile = File(...),
    vendor_id: str | None = Form(None),
    vendor_name: str | None = Form(None),
    category_id: str | None = Form(None),
    expand_units: bool = Form(False),
    allow_append: bool = Form(False),
):
    raw = await file.read()
    filename = file.filename or "upload"
    ext = (filename.rsplit(".", 1)[-1] or "").lower()
    raw_b64 = base64.b64encode(raw).decode("ascii")

    job_id = str(uuid.uuid4())
    q = rq.Queue("ai", connection=_redis_conn())
    q.enqueue(
        "jobs.ai_commit_job",
        job_id,
        po_number,
        vendor_id,
        vendor_name,
        category_id,
        expand_units,
        allow_append,
        raw_b64,       # <-- base64 string
        filename,
        ext,
        os.getenv("DATABASE_URL"),
        os.getenv("REDIS_URL"),
        job_id=job_id,
        description=f"AI commit for PO {po_number}",
    )
    return {"ok": True, "job_id": job_id}

def _decode_bytes(x):
    if isinstance(x, (bytes, bytearray)):
        return x.decode("utf-8", "ignore")
    return x if isinstance(x, str) else str(x)

def _stream_pubsub_sse(channel: str) -> StreamingResponse:
    """
    Subscribes to Redis Pub/Sub `channel` and streams JSON events as SSE.
    - Sends an initial "Queued" progress event so the UI renders immediately.
    - Emits a lightweight heartbeat every ~10s to keep proxies from buffering.
    - Stops on terminal event types: "complete" or "error".
    """
    r = get_redis()
    pubsub = r.pubsub(ignore_subscribe_messages=True)
    pubsub.subscribe(channel)

    async def gen():
        # Draw UI immediately
        yield _sse({"type": "progress", "pct": 1, "label": "Queued"})
        last_beat = time.monotonic()
        try:
            while True:
                msg = pubsub.get_message(timeout=1.0)
                if msg and msg.get("type") == "message":
                    raw = msg.get("data")
                    try:
                        event = json.loads(_decode_bytes(raw))
                    except Exception:
                        event = {"type": "error", "message": "invalid_event_payload"}
                    # Ship the event
                    yield _sse(event)

                    # Free the server coroutine on terminal events
                    if event.get("type") in ("complete", "error"):
                        break

                # Heartbeat every ~10s so proxies don't buffer the stream
                now = time.monotonic()
                if now - last_beat > 10:
                    yield ": keep-alive\n\n"
                    last_beat = now

                await asyncio.sleep(0.1)
        finally:
            try:
                pubsub.unsubscribe(channel)
            except Exception:
                pass
            try:
                pubsub.close()
            except Exception:
                pass

    headers = {
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no",  # nginx
    }
    return StreamingResponse(gen(), media_type="text/event-stream", headers=headers)

# Preview job events (single source of truth)
@app.get("/imports/ai-preview-jobs/{job_id}/events")
async def stream_ai_preview_events(job_id: str):
    return _stream_pubsub_sse(f"ai:preview:{job_id}")

# Commit job events (single source of truth)
@app.get("/imports/ai-commit-jobs/{job_id}/events")
async def stream_ai_commit_events(job_id: str):
    return _stream_pubsub_sse(f"ai:commit:{job_id}")


@app.get("/admin/ebay/token")
def admin_ebay_token():
    tok = get_ebay_token(force=True)
    if not tok:
        raise HTTPException(status_code=500, detail="no_token")
    return {
        "ok": True,
        "prefix": tok[:12],
        "expires_in": max(0, int(EBAY_TOKEN_CACHE["exp"] - time.time())),
        "marketplace": EBAY_MARKETPLACE_ID,
    }
# -----------------------------------------------------------------------------
# Gemini AI (safe wiring + diagnostics)
# -----------------------------------------------------------------------------
_genai_loaded = False
try:
    import google.generativeai as genai
    _genai_loaded = True
except Exception as e:
    print("[AI] google-generativeai not available:", e)

GEMINI_API_KEY = (
    os.getenv("GEMINI_API_KEY", "") or
    os.getenv("VITE_GEMINI_API_KEY", "")
)
if not GEMINI_API_KEY:
    print("[AI] No Gemini API key found in GEMINI_API_KEY or VITE_GEMINI_API_KEY")
elif not _genai_loaded:
    print("[AI] Package missing: pip install google-generativeai")
else:
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        print("[AI] Gemini configured âœ”")
    except Exception as e:
        print("[AI] configure failed:", e)

AI_RESPONSE_SCHEMA = {
    "type": "object",
    "properties": {
        "detected_headers": {"type": "array", "items": {"type": "string"}},
        "lines": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "product_name_raw": {"type": "string"},
                    "qty": {"type": "integer", "minimum": 1},
                    "unit_cost": {"type": ["number", "null"]},
                    "msrp": {"type": ["number", "null"]},
                    "upc": {"type": ["string", "null"]},
                    "asin": {"type": ["string", "null"]},
                    "category_guess": {"type": ["string", "null"]},
                },
                "required": ["product_name_raw", "qty"],
            },
        },
        "notes": {"type": "string"},
    },
    "required": ["lines"],
    "additionalProperties": False,
}

AI_PARSER_PROMPT = """You are a purchasing-sheet parser. Input is a spreadsheet or PDF
containing a purchase order (PO) or inventory manifest. Output is a SINGLE
JSON object matching the schema belowâ€”no prose, no code fences.

You must respect these MODE FLAGS (set by the caller at the end of this prompt):
- EXPAND_UNITS = true|false
  â€¢ true  â†’ If a row indicates Quantity = N (>1), expand that row into N separate line objects.
            Each expanded line must have qty = 1 and the same per-unit unit_cost and other fields.
            If unit_cost is not given but a total/extended amount exists, compute unit_cost = round(total/N, 2).
  â€¢ false â†’ Do not expand. Keep each vendor row as one output line with qty = N.

Core mapping rules:
- product_name_raw: The best available item description/title.
- qty: integer â‰¥ 1 (see EXPAND_UNITS).
- unit_cost:
    * Prefer the per-unit price column if present.
    * If the sheet has a column named â€œSTOCK IMAGEâ€ (any case), that value is the ACTUAL price paid per unit â†’ map it to unit_cost.
    * If only totals exist, compute unit_cost = total/qty (2 decimals).
    * Never use totals/extended amounts directly as unit_cost without normalizing to per-unit.
- msrp: Manufacturer suggested retail/original/list price. Synonyms include â€œOrig. Retailâ€, â€œOriginal Retailâ€, â€œListâ€, â€œMSRPâ€.
- upc / asin: Extract if present; also scan description text for clear UPC/ASIN tokens.
- category_guess:
    * If a â€œCategoryâ€ column is present, use it (trimmed).
    * Otherwise infer from the product name/specs. Keep this short and human-readable
      (e.g., â€œLaptopâ€, â€œDesktopâ€, â€œMonitorâ€, â€œGPUâ€, â€œCPUâ€, â€œMotherboardâ€, â€œMemoryâ€, â€œStorageâ€,
       â€œKeyboardâ€, â€œMouseâ€, â€œHeadsetâ€, â€œConsoleâ€, â€œControllerâ€, â€œNetworkâ€, â€œPowerâ€, â€œAccessoryâ€, â€œOtherâ€).
    * Choose the most specific one youâ€™re confident about; else null.

Header synonyms:
- product_name_raw: â€œProductâ€, â€œItemâ€, â€œListing Titleâ€, â€œTitleâ€, â€œNameâ€, â€œDescriptionâ€.
- unit_cost: â€œUnit Costâ€, â€œPrice Paidâ€, â€œCostâ€, â€œOur Costâ€, â€œSTOCK IMAGEâ€.
- msrp: â€œOrig. Retailâ€, â€œOriginal Retailâ€, â€œListâ€, â€œMSRPâ€.
- qty: â€œQtyâ€, â€œQuantityâ€, â€œQTYâ€.

Currency & hygiene:
- Parse currency like â€œ$ 1,499.99â€ â†’ 1499.99 (float). Missing/blank â†’ null.
- Ignore non-item rows: headers, subtotals, grand totals, shipping, tax, footers, ads.
- If multiple header rows, detect the best header row.
- Keep JSON compact. Do NOT include markdown, comments, or extra keys.

SCHEMA (return EXACTLY this shape):
{
  "detected_headers": [string],     // the literal column headers you used
  "lines": [
    {
      "product_name_raw": string,
      "qty": integer,               // = 1 if EXPAND_UNITS=true; else may be >1
      "unit_cost": number|null,
      "msrp": number|null,
      "upc": string|null,
      "asin": string|null,
      "category_guess": string|null
    }
  ],
  "notes": string                   // brief, e.g., "Expanded 3 rows into 70 units; used STOCK IMAGE for unit_cost."
}

MODE FLAGS:
EXPAND_UNITS = {expand_units}
"""


def make_ai_parser_prompt(expand_units: bool) -> str:
    """
    Compose the parsing prompt with mode flags. Keeps the base instructions in AI_PARSER_PROMPT
    and toggles expansion behavior via {expand_units}.
    """
    flag = "true" if expand_units else "false"
    return AI_PARSER_PROMPT.replace("{expand_units}", flag)


def _postprocess_lines(lines: list[dict], expand_units: bool) -> list[dict]:
    """
    Enforce numeric types and optionally expand rows into qty=1 units on the server side,
    regardless of what the model did. This guarantees UI consistency.
    """
    out: list[dict] = []
    for ln in lines or []:
        name = (ln.get("product_name_raw") or "").strip()
        if not name:
            continue
        try:
            qty = int(ln.get("qty") or 1)
        except Exception:
            qty = 1
        qty = max(1, qty)
        unit_cost = to_num(ln.get("unit_cost"))
        msrp = to_num(ln.get("msrp"))
        upc = (str(ln.get("upc")).strip() if ln.get("upc") else None)
        asin = (str(ln.get("asin")).strip() if ln.get("asin") else None)
        category_guess = (str(ln.get("category_guess")).strip() if ln.get("category_guess") else None)

        base = {
            "product_name_raw": name,
            "unit_cost": unit_cost,
            "msrp": msrp,
            "upc": upc,
            "asin": asin,
            "category_guess": category_guess,
        }

        if expand_units:
            for _ in range(qty):
                out.append({**base, "qty": 1})
        else:
            out.append({**base, "qty": qty})
    return out



def _to_float(x):
    if x is None: return None
    if isinstance(x, dict):
        v = x.get('value') or x.get('Value') or x.get('_value') or x.get('__value__')
        return float(v) if v is not None else None
    try:
        return float(x)
    except Exception:
        return None
def _to_float2(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    try:
        s = re.sub(r"[^\d\.\-]", "", str(v))
        return float(s) if s else None
    except Exception:
        return None

@app.get("/ai/health")
def ai_health():
    """Never crash: report what the API sees at runtime."""
    return {
        "genai_imported": bool(_genai_loaded),
        "has_key": bool(GEMINI_API_KEY),
        "ai_first": bool(AI_FIRST),
        "configured": bool(_genai_loaded and GEMINI_API_KEY),
    }

# -----------------------------------------------------------------------------
# eBay proxy (passes Authorization & marketplace header)
# -----------------------------------------------------------------------------
@app.api_route("/ebay/{subpath:path}", methods=["GET", "POST", "PUT", "DELETE"])
async def ebay_proxy(subpath: str, request: Request):
    token = get_ebay_token()
    if not token:
        raise HTTPException(401, "no_token")

    url = f"https://api.ebay.com/{subpath}"
    headers = {
        "Authorization": f"Bearer {token}",
        "X-EBAY-C-MARKETPLACE-ID": EBAY_MARKETPLACE_ID,
        "Accept": "application/json",
    }

    method = request.method.upper()
    params = dict(request.query_params)

    data = None
    raw = None
    ctype = request.headers.get("content-type", "")

    if method in ("POST", "PUT", "DELETE"):
        if ctype.startswith("application/json"):
            try:
                data = await request.json()
            except Exception:
                data = None
            headers["Content-Type"] = "application/json"
        else:
            raw = await request.body()
            headers["Content-Type"] = ctype or "application/octet-stream"

    resp = session.request(method, url, headers=headers, params=params, json=data, data=raw)
    return Response(
        content=resp.content,
        status_code=resp.status_code,
        media_type=resp.headers.get("Content-Type", "application/json"),
    )

def get_price_from_getitem(item: dict) -> tuple[float|None, str|None]:
    listing_type = item.get('ListingType')
    currency = None

    # variations first (if present)
    vars = (item.get('Variations') or {}).get('Variation')
    if vars:
        if isinstance(vars, dict):
            vars = [vars]
        prices = []
        for v in vars:
            sp = v.get('StartPrice')
            price = _to_float(sp)
            if price is not None:
                prices.append((price, (sp or {}).get('currencyID') or (sp or {}).get('_currencyID')))
        if prices:
            price, currency = min(prices, key=lambda t: t[0])  # or match specifics to your row.specs
            return price, currency

    # fixed price vs auction
    if listing_type == 'FixedPriceItem':
        sp = item.get('StartPrice')
        price = _to_float(sp)
        currency = (sp or {}).get('currencyID') or (sp or {}).get('_currencyID')
        return price, currency

    # auction (or fallback)
    cp = ((item.get('SellingStatus') or {}).get('CurrentPrice'))
    price = _to_float(cp)
    currency = (cp or {}).get('currencyID') or (cp or {}).get('_currencyID')
    return price, currency



# -----------------------------------------------------------------------------
# Scraping helpers (sold_avg + scrape)
# -----------------------------------------------------------------------------
def soup_from_html(html: str) -> BeautifulSoup:
    try:
        return BeautifulSoup(html or "", "lxml")
    except FeatureNotFound:
        return BeautifulSoup(html or "", "html.parser")

SPEC_KEYS = ("brand", "model", "processor", "cpu", "ram", "storage", "screen", "size")

def build_query_from_row(row: dict) -> str:
    toks = []
    name = (row.get("productName") or row.get("product_name") or "").strip()
    if name:
        toks.append(name)

    specs = row.get("specs") or {}
    flat  = {k: row.get(k) for k in SPEC_KEYS if row.get(k)}
    merged = {**specs, **{k: v for k, v in flat.items() if v}}

    cpu = merged.get("processor") or merged.get("cpu")
    if cpu: toks.append(str(cpu))
    ram = merged.get("ram")
    if ram: toks.append(str(ram).replace(" GB", "GB"))
    storage = merged.get("storage")
    if storage: toks.append(str(storage).replace(" GB", "GB"))
    size = merged.get("screen") or merged.get("size")
    if size:
        toks.append(str(size).replace(' "', '').replace('"', "").replace(" inch", "").strip() + '"')

    brand = merged.get("brand")
    model = merged.get("model")
    if brand and (brand not in name): toks.insert(0, str(brand))
    if model and (model not in name): toks.append(str(model))
    return " ".join(t for t in toks if t).strip()

def parse_price(text: str) -> Tuple[str, Optional[float], str]:
    if not text:
        return ("", None, "")
    t = re.sub(r"\s+", " ", text).strip()
    m = re.search(r"(?:(US|C|CA)\s*)?([$Â£â‚¬])\s?(\d[\d,]*\.?\d*)", t, re.I)
    if not m:
        return ("", None, t)
    region = (m.group(1) or "").upper()
    sym, val = m.group(2), m.group(3)
    try:
        num = float(val.replace(",", ""))
    except ValueError:
        num = None
    if sym == "$":
        cur = "CA" if region in ("C", "CA") else "US"
    elif sym == "Â£":
        cur = "GB"
    elif sym == "â‚¬":
        cur = "EU"
    else:
        cur = ""
    return (cur, num, (region + " " if region else "") + sym + val)

def shape_rows_from_items(items, hits: int) -> List[Dict[str, Any]]:
    rows = []
    for li in items[:hits]:
        a = li.select_one("a.s-item__link")
        title_el = li.select_one(".s-item__title") or li.select_one("h3")
        price_el = li.select_one(".s-item__price")
        if not a or not a.get("href") or not title_el:
            continue
        title = re.sub(r"\s+", " ", title_el.get_text(strip=True))
        price_text = re.sub(r"\s+", " ", price_el.get_text(strip=True)) if price_el else ""
        cur, num, pretty = parse_price(price_text)
        rows.append({
            "url": a["href"],
            "title": title,
            "priceText": pretty or price_text,
            "currency": cur,
            "priceNum": num,
        })
    return rows

def _http_session() -> requests.Session:
    s = requests.Session()
    retry = Retry(
        total=3, connect=3, read=3, backoff_factor=0.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET", "HEAD"),
        raise_on_status=False,
    )
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.mount("http://", HTTPAdapter(max_retries=retry))
    return s

def is_uuid_like(s) -> bool:
    try:
        UUID(str(s))
        return True
    except Exception:
        return False

def to_ymd(v):
    if v in (None, ""): return None
    if isinstance(v, datetime): v = v.date()
    if isinstance(v, date):     return v.isoformat()
    try:
        return datetime.fromisoformat(str(v).replace("Z","+00:00")).date().isoformat()
    except Exception:
        s = str(v)[:10]
        return s if len(s) == 10 else None

def summarize_rows(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    usd = [r for r in rows if r["currency"] == "US" and isinstance(r.get("priceNum"), (int, float))]
    valid = len(usd)
    avg = round(sum(r["priceNum"] for r in usd) / valid, 2) if valid else 0.0
    return {"sampled": len(rows), "valid": valid, "avg": avg}

def scrape_via_playwright(url: str, hits: int = 20) -> Dict[str, Any]:
    """Optional fallback; safe no-op if Playwright isn't installed."""
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except Exception as e:
        return {"ok": False, "error": f"playwright_not_available: {e}"}
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            ctx = browser.new_context(locale="en-US", viewport={"width": 1400, "height": 900})
            page = ctx.new_page()
            page.goto(url, wait_until="domcontentloaded", timeout=20000)
            try:
                page.wait_for_load_state("networkidle", timeout=12000)
            except Exception:
                pass
            for _ in range(6):
                page.mouse.wheel(0, 1200)
                page.wait_for_timeout(400)
            selectors = [
                "ul.srp-results li.s-item",
                "li.s-item",
                "[data-testid='item-cell']",
            ]
            cards = []
            for sel in selectors:
                try:
                    page.wait_for_selector(sel, timeout=4000)
                    cards = page.query_selector_all(sel)
                    if cards:
                        break
                except PWTimeout:
                    continue

            rows: List[Dict[str, Any]] = []
            def extract(li):
                a = li.query_selector("a.s-item__link") or li.query_selector("a[href*='/itm/']")
                title_el = li.query_selector(".s-item__title") or li.query_selector("h3, .s-item__subtitle")
                price_el = li.query_selector(".s-item__price, .s-item__detail--primary span")
                if not a or not title_el:
                    return None
                title = (title_el.inner_text() or "").strip()
                href = a.get_attribute("href") or ""
                price_text = (price_el.inner_text() or "").strip() if price_el else ""
                m = re.search(r"([$Â£â‚¬])\s?(\d[\d,]*\.?\d*)", price_text or "")
                price_num = float(m.group(2).replace(",", "")) if m else None
                return {"title": title, "url": href, "priceText": price_text, "priceNum": price_num,
                        "currency": "US" if price_num is not None else ""}

            for li in cards[:hits]:
                row = extract(li)
                if row:
                    rows.append(row)

            browser.close()

        if not rows:
            return {"ok": False, "error": "challenge_or_no_results_after_browser"}
        summary = summarize_rows(rows)
        return {"ok": True, **summary, "rows": rows}
    except Exception as e:
        return {"ok": False, "error": f"playwright_error: {e}"}

# -----------------------------------------------------------------------------
# Posted-row lookup (mock)
# -----------------------------------------------------------------------------
class PostedSearch(BaseModel):
    productName: Optional[str] = None
    grade: Optional[str] = None
def _is_uuid_like(x: str) -> bool:
    try:
        UUID(str(x)); return True
    except Exception:
        return False

@app.post("/rows/search_posted")
def search_posted_row(body: PostedSearch):
    name = (body.productName or "").strip().lower()
    if name and "laptop" in name:
        return {"ebayItemUrl": "https://www.ebay.com/itm/MOCK_COMP_LINK"}
    return {}

    # GET /rows/brief?ids=12345-0001,12345-0002
@app.get("/rows/brief")
def rows_brief(ids: str):
    arr: List[str] = [s.strip() for s in (ids or "").split(",") if s.strip()]
    if not arr:
        return []

    out: List[Dict] = []

    with db() as (_, cur):
        # ---- A) Aggregate by PO LINE codes (TEXT) ----
        # Works whether your pl.synergy_id holds '23452-00023' style codes or UUID-looking strings.
        cur.execute(
            """
            WITH agg AS (
              SELECT
                pl.synergy_id::text AS "synergyId",
                COALESCE(
                  BOOL_OR(
                    i.status = 'TESTED'
                    OR (i.grade IS NOT NULL AND i.tested_by IS NOT NULL AND i.tested_date IS NOT NULL)
                  ),
                  FALSE
                ) AS tested,
                MAX(i.grade) FILTER (WHERE i.grade IS NOT NULL)                 AS grade,
                MAX(i.tested_by::text) FILTER (WHERE i.tested_by IS NOT NULL)  AS "testedBy",
                MAX(i.tested_date)                                            AS tested_date,
                COALESCE(
                  BOOL_OR(
                    COALESCE(i.ebay_item_url,'') <> ''
                    OR i.posted_at IS NOT NULL
                    OR i.posted_by IS NOT NULL
                  ),
                  FALSE
                ) AS posted,
                MAX(i.posted_at)                                              AS "postedAt",
                MAX(i.posted_by::text)                                        AS "postedBy",
                MAX(i.ebay_price)                                             AS "ebayPrice",
                MAX(i.ebay_item_url)                                          AS "ebayItemUrl"
              FROM public.po_lines pl
              LEFT JOIN public.inventory_items i ON i.po_line_id = pl.id
              WHERE pl.synergy_id = ANY(%s)            -- TEXT compare (no uuid cast!)
              GROUP BY pl.synergy_id
            )
            SELECT
              a."synergyId",
              CASE WHEN a.tested THEN 'TESTED' ELSE 'UNTESTED' END            AS status,
              a.grade,
              a."testedBy",
              TO_CHAR(a.tested_date,'YYYY-MM-DD')                             AS "testedDate",
              a.posted,
              a."postedAt",
              a."postedBy",
              a."ebayPrice",
              a."ebayItemUrl"
            FROM agg a
            """,
            (arr,),
        )
        line_rows = [dict(r) for r in cur.fetchall()]
        out.extend(line_rows)

        # Build a set of IDs we already satisfied via line aggregation
        line_ids = {r["synergyId"] for r in line_rows}

        # ---- B) Back-compat: item codes (inventory_items.synergy_code) for any remaining ids ----
        remaining = [x for x in arr if x not in line_ids]
        if remaining:
            cur.execute(
                """
                SELECT
                  i.synergy_code                                                AS "synergyId",
                  i.status,
                  i.grade,
                  i.tested_by                                                   AS "testedBy",
                  TO_CHAR(i.tested_date,'YYYY-MM-DD')                           AS "testedDate",
                  (COALESCE(i.ebay_item_url,'') <> '' OR i.posted_at IS NOT NULL OR i.posted_by IS NOT NULL) AS posted,
                  i.posted_at                                                   AS "postedAt",
                  i.posted_by::text                                             AS "postedBy",
                  i.ebay_price                                                  AS "ebayPrice",
                  i.ebay_item_url                                               AS "ebayItemUrl"
                FROM public.inventory_items i
                WHERE i.synergy_code = ANY(%s)
                """,
                (remaining,),
            )
            out.extend([dict(r) for r in cur.fetchall()])

    return out

def _browse_get_item_by_legacy_id(legacy_id: str) -> tuple[dict | None, int]:
    """
    Returns (json, status_code) from eBay Buy Browse API get_item_by_legacy_id.
    """
    token = _ebay_access_token_from_refresh()
    r = session.get(
        "https://api.ebay.com/buy/browse/v1/item/get_item_by_legacy_id",
        params={"legacy_item_id": legacy_id},
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "X-EBAY-C-MARKETPLACE-ID": EBAY_MARKETPLACE_ID,
        },
        timeout=(6, 12),
    )
    try:
        j = r.json() if r.content else None
    except Exception:
        j = None
    return j, r.status_code

def _finding_completed_for_item_id(legacy_id: str) -> tuple[float | None, str | None, bool]:
    """
    API-only fallback using Finding API (findCompletedItems) to infer sold price.
    Requires EBAY_APP_ID in env. Returns (price, currency, sold).
    """
    if not EBAY_APP_ID:
        return None, None, False

    # Docs: https://developer.ebay.com/Devzone/finding/Concepts/FindingAPIGuide.html
    # Note: findCompletedItems doesn't accept itemId directly; we fake a tight keyword with item ID.
    # This is best-effort: some items wonâ€™t resolve. Still API-only.
    r = session.get(
        "https://svcs.ebay.com/services/search/FindingService/v1",
        params={
            "OPERATION-NAME": "findCompletedItems",
            "SERVICE-VERSION": "1.13.0",
            "SECURITY-APPNAME": EBAY_APP_ID,
            "RESPONSE-DATA-FORMAT": "JSON",
            "REST-PAYLOAD": "true",
            "keywords": legacy_id,
            "paginationInput.entriesPerPage": "1",
            "sortOrder": "EndTimeSoonest",
        },
        timeout=(6, 12),
    )
    try:
        j = r.json()
        items = (
            j.get("findCompletedItemsResponse", [{}])[0]
             .get("searchResult", [{}])[0]
             .get("item", [])
        )
        if not items:
            return None, None, False
        it = items[0]
        selling_state = (it.get("sellingStatus", [{}])[0].get("sellingState", [""])[0] or "").lower()
        sold = selling_state in ("endedwithsales", "sold")
        price_obj = it.get("sellingStatus", [{}])[0].get("currentPrice", [{}])[0]
        price = price_obj.get("__value__")
        currency = price_obj.get("@currencyId")
        return (float(price) if price is not None else None), currency, sold
    except Exception:
        return None, None, False


@app.post("/ebay/sync_by_url")
def ebay_sync_by_url(body: EbaySyncBody):
    url = (body.ebayUrl or "").strip()
    if not url:
        raise HTTPException(400, "ebayUrl required")

    legacy_id = _parse_legacy_item_id(url)
    if not legacy_id:
        raise HTTPException(400, "Could not extract itemId from ebayUrl")

    price = None
    currency = None
    sold = False
    title = None

    # 1) Browse API (primary, API-only)
    j, status = _browse_get_item_by_legacy_id(legacy_id)

    if status == 200 and isinstance(j, dict):
        title = j.get("title") or None
        p = (j.get("price") or {}).get("value")
        if p is not None:
            try:
                price = float(p)
            except Exception:
                price = None
        currency = (j.get("price") or {}).get("currency") or currency

        # basic sold/end signal using only API fields
        availability = ((j.get("availability") or {}).get("status") or "").upper()
        item_end = j.get("itemEndDate") or j.get("itemEndTime")
        sold = (availability in {"OUT_OF_STOCK", "UNAVAILABLE"}) or bool(item_end)

    elif status in (403, 404, 410):
        # Not found or ended / hidden via Browse â€” optional API fallback to Finding
        # Remove this block if you truly want Browse-only.
        p2, cur2, sold2 = _finding_completed_for_item_id(legacy_id)
        if p2 is not None:
            price = p2
            currency = cur2 or currency
            sold = sold or sold2

    else:
        # Other API status: keep link, return minimal data
        pass

    # Persist minimal fields back to your DB (no HTML-derived data)
    with db() as (con, cur):
        cur.execute(
            """
            UPDATE inventory_items
               SET ebay_item_url = %s,
                   ebay_price    = COALESCE(%s, ebay_price),
                   status        = CASE WHEN %s THEN 'SOLD' ELSE status END
             WHERE synergy_code  = %s
            """,
            (url, price, sold, body.synergyId),
        )
        con.commit()

    return {
        "ok": True,
        "synergyId": body.synergyId,
        "ebayItemUrl": url,
        "ebayPrice": price,     # None if API didnâ€™t provide
        "currency": currency or "USD",
        "sold": bool(sold),
        "title": title,
    }

@app.post("/api/auth/manager/login")
def manager_login(payload: dict = Body(...)):
    
    user_id = payload.get("userId")
    password = payload.get("password")
    if not user_id or not password:
        raise HTTPException(status_code=400, detail="Missing credentials")

    with db() as (con, cur):
        cur.execute(
            """
            SELECT id, name
            FROM public.app_users
            WHERE id = %s
              AND role = 'manager'
              AND active = TRUE
              AND password_hash = crypt(%s, password_hash)
            """,
            (user_id, password),
        )
        row = cur.fetchone()

    if not row:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Opaque token is fine since your UI only checks presence of a token.
    token = base64.urlsafe_b64encode(os.urandom(24)).decode("ascii").rstrip("=")
    return {"token": token}


@app.get("/rows/count")
def rows_count(status: str | None = None):
    # status can be "ready" or "incomplete"
    where = []
    if status == "ready":
        where.append("grade IS NOT NULL AND tested_by IS NOT NULL AND (tested_date IS NOT NULL OR tested_at IS NOT NULL)")
    elif status == "incomplete":
        where.append("NOT (grade IS NOT NULL AND tested_by IS NOT NULL AND (tested_date IS NOT NULL OR tested_at IS NOT NULL))")
    sql = "SELECT COUNT(*) AS total FROM public.inventory_items " + ("WHERE " + " AND ".join(where) if where else "")
    with db() as (_, cur):
        cur.execute(sql)
        return {"total": int(cur.fetchone()["total"])}

@app.get("/rows/counts")
def rows_counts(groupBy: str | None = None):
    # returns totals in one shot
    sql = """
      SELECT
        COUNT(*) AS total,
        COUNT(*) FILTER (WHERE grade IS NOT NULL AND tested_by IS NOT NULL AND (tested_date IS NOT NULL OR tested_at IS NOT NULL)) AS ready,
        COUNT(*) FILTER (WHERE NOT (grade IS NOT NULL AND tested_by IS NOT NULL AND (tested_date IS NOT NULL OR tested_at IS NOT NULL))) AS incomplete,
        COUNT(*) FILTER (WHERE tested_date = CURRENT_DATE) AS today_tested
      FROM public.inventory_items
    """
    with db() as (_, cur):
      cur.execute(sql)
      r = cur.fetchone()
      return {
        "total": int(r["total"]),
        "ready": int(r["ready"]),
        "incomplete": int(r["incomplete"]),
        "todayTested": int(r["today_tested"]),
      }
def _tokenize_for_like(s: str) -> List[str]:
    import re
    raw = re.split(r"[\s\-_/|,.;:()+]+", s or "")
    toks: List[str] = []
    for t in raw:
        t = t.strip()
        if len(t) < 2:
            continue
        toks.append(t[:64])
        if len(toks) >= 8:
            break
    return toks
@app.get("/search")
def global_search(
    q: str = Query(..., min_length=2, description="Free-text search"),
    limit: int = Query(50, ge=1, le=200),
):
    """
    Site-wide search over POs, PO lines, vendors, and categories.
    Tokens are ANDed (all tokens must appear in a row).
    """
    tokens = [t for t in re.findall(r"[A-Za-z0-9\-._]+", q) if t]
    if not tokens:
        return {"po": [], "line": [], "vendor": [], "category": []}

    def ilike_where(cols: List[str]) -> tuple[str, list]:
        # Build: AND (col1 ILIKE %tok% OR col2 ILIKE %tok% â€¦)
        parts = []
        params: list = []
        for tok in tokens:
            ors = " OR ".join([f"{c} ILIKE %s" for c in cols])
            parts.append(f"({ors})")
            params.extend([f"%{tok}%"] * len(cols))
        return " AND ".join(parts), params

    with db() as (con, cur):
        # POs
        po_where, po_params = ilike_where(["LOWER(TRIM(po.po_number))", "LOWER(TRIM(v.name))"])
        cur.execute(f"""
            SELECT po.id, po.po_number, COALESCE(v.name,'') AS vendor_name
            FROM purchase_orders po
            LEFT JOIN vendors v ON v.id = po.vendor_id
            WHERE {po_where}
            ORDER BY po.created_at DESC NULLS LAST
            LIMIT %s
        """, po_params + [limit])
        pos = [{"id": r["id"], "po_number": r["po_number"], "vendor_name": r["vendor_name"]} for r in cur.fetchall()]

        # PO lines (across ALL POs)
        line_where, line_params = ilike_where(["LOWER(pl.product_name_raw)", "pl.upc", "pl.asin", "pl.synergy_id::text"])
        cur.execute(f"""
            SELECT
              pl.id,
              pl.product_name_raw,
              pl.upc,
              pl.asin,
              COALESCE(pl.qty,1)      AS qty,
              COALESCE(pl.unit_cost,0) AS unit_cost,
              COALESCE(pl.msrp,0)      AS msrp,
              pl.category_guess        AS category_id,
              pl.synergy_id::text      AS synergy_id,
              pl.purchase_order_id     AS po_id
            FROM po_lines pl
            WHERE {line_where}
            ORDER BY pl.id DESC
            LIMIT %s
        """, line_params + [limit])
        lines = [dict(r) for r in cur.fetchall()]

        # Vendors
        vend_where, vend_params = ilike_where(["LOWER(v.name)"])
        cur.execute(f"""
            SELECT v.id, v.name
            FROM vendors v
            WHERE {vend_where}
            ORDER BY v.name ASC
            LIMIT %s
        """, vend_params + [limit])
        vendors = [{"id": r["id"], "name": r["name"]} for r in cur.fetchall()]

        # Categories
        cat_where, cat_params = ilike_where(["LOWER(c.label)", "LOWER(c.prefix)"])
        cur.execute(f"""
            SELECT c.id, c.label, COALESCE(c.prefix,'') AS prefix
            FROM categories c
            WHERE {cat_where}
            ORDER BY c.label ASC
            LIMIT %s
        """, cat_params + [limit])
        cats = [{"id": r["id"], "label": r["label"], "prefix": r["prefix"]} for r in cur.fetchall()]

    return {
        "po": pos,
        "line": lines,
        "vendor": vendors,
        "category": cats,
    }

@app.post("/rows/search_similar")
def rows_search_similar(payload: Dict[str, Any] = Body(...)):
    """
    Find similar items in our DB (heuristic, no eBay).
    Body:
      {
        "productName": str?,
        "specs": { processor?, ram?, storage?, screen?, batteryHealth?, color? }?,
        "grade": str?,
        "categoryId": str?,     # uuid or label/prefix text
        "excludeSynergyId": str?,
        "limit": int?           # default 5
      }
    """
    name = (payload.get("productName") or "").strip()
    specs = payload.get("specs") or {}
    grade = (payload.get("grade") or None) or None
    category_id = payload.get("categoryId") or None
    exclude_synergy = (payload.get("excludeSynergyId") or "").strip()
    limit = int(payload.get("limit") or 5)

    blob = " ".join([
        name,
        str(specs.get("processor") or ""),
        str(specs.get("ram") or ""),
        str(specs.get("storage") or ""),
        str(specs.get("screen") or ""),
        str(specs.get("batteryHealth") or ""),
        str(specs.get("color") or ""),
    ])
    tokens = _tokenize_for_like(blob)

    where = ["TRUE"]
    params: Dict[str, Any] = {"limit": limit}

    if exclude_synergy:
        where.append("i.synergy_code <> %(exclude)s")
        params["exclude"] = exclude_synergy

    if category_id:
        # allow uuid OR free-text match on category label/prefix
        if is_uuid_like(str(category_id)):  # uses your existing helper
            where.append("i.category_id = %(cat_uuid)s::uuid")
            params["cat_uuid"] = str(category_id)
        else:
            params["cat_like"] = "%" + str(category_id).replace("%", r"\%").replace("_", r"\_") + "%"
            where.append("(c.label ILIKE %(cat_like)s ESCAPE '\\' OR c.prefix ILIKE %(cat_like)s ESCAPE '\\')")

    if grade:
        where.append("i.grade = %(grade)s")
        params["grade"] = grade

    match_clauses: List[str] = []
    score_parts: List[str] = ["0"]
    for idx, tok in enumerate(tokens):
        key = f"tok{idx}"
        like = "%" + tok.replace("%", r"\%").replace("_", r"\_") + "%"
        params[key] = like
        match_clauses.append(
            f"(pl.product_name_raw ILIKE %({key})s ESCAPE '\\' OR CAST(i.specs AS text) ILIKE %({key})s ESCAPE '\\')"
        )
        # Name hit is weight 2, specs hit is weight 1
        score_parts.append(f"CASE WHEN pl.product_name_raw ILIKE %({key})s ESCAPE '\\' THEN 2 ELSE 0 END")
        score_parts.append(f"CASE WHEN CAST(i.specs AS text) ILIKE %({key})s ESCAPE '\\' THEN 1 ELSE 0 END")

    text_where = f"({' OR '.join(match_clauses)})" if match_clauses else "TRUE"

    sql = f"""
      SELECT
        i.synergy_code                    AS "synergyId",
        COALESCE(pl.product_name_raw, '') AS "productName",
        i.grade                           AS "grade",
        i.status                          AS "status",
        i.category_id                     AS "categoryId",
        c.label                           AS "categoryLabel",
        COALESCE(i.specs, '{{}}'::jsonb)  AS "specs",
        COALESCE(i.price, 0)              AS "price",
        COALESCE(i.ebay_price, 0)         AS "ebayPrice",
        ({' + '.join(score_parts)})       AS score
      FROM inventory_items i
      LEFT JOIN po_lines   pl ON pl.id = i.po_line_id
      LEFT JOIN categories c  ON c.id = i.category_id
      WHERE {' AND '.join(where)} AND {text_where}
      ORDER BY score DESC,
               COALESCE(i.posted_at, i.tested_date, i.tested_at) DESC NULLS LAST,
               i.synergy_code ASC
      LIMIT %(limit)s
    """

    with db() as (con, cur):
        cur.execute(sql, params)
        return {"items": [dict(r) for r in cur.fetchall()]}
# -----------------------------------------------------------------------------
# Local sheet parsing helpers (CSV/XLSX)  â€” used as AI fallback
# -----------------------------------------------------------------------------
# Optional XLSX support
try:
    from openpyxl import load_workbook  # for .xlsx
except Exception:
    load_workbook = None

def _parse_locally_from_file(tmp_path: str, suffix: str) -> list[dict]:
    """Returns rows with normalized keys using map_header()."""
    rows: list[dict] = []
    ext = suffix.lower().lstrip(".")
    if ext in ("csv", "tsv", "txt"):
        with open(tmp_path, "r", encoding="utf-8", errors="ignore") as fh:
            text = fh.read()
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
            reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        except Exception:
            reader = csv.DictReader(io.StringIO(text))
        for raw in reader:
            r = {}
            for k, v in (raw or {}).items():
                r[map_header(k)] = v
            rows.append(r)
    elif ext in ("xlsx", "xlsm", "xltx", "xltm"):
        if load_workbook is None:
            raise RuntimeError("openpyxl not installed")
        wb = load_workbook(filename=tmp_path, read_only=True, data_only=True)
        ws = wb.active
        hdr_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [map_header(str(c) if c is not None else "") for c in hdr_row]
        for rr in ws.iter_rows(min_row=2, values_only=True):
            obj = {}
            for i, val in enumerate(rr):
                key = headers[i] if i < len(headers) else f"col_{i}"
                obj[key] = val
            rows.append(obj)
    else:
        raise RuntimeError(f"no_local_parser_for_ext:{ext}")
    return rows

def _rows_to_csv_text(rows: list[dict], limit: int = 200) -> str:
    rows = rows[:max(1, limit)]
    headers, seen = [], set()
    for r in rows:
        for k in r.keys():
            m = str(k or "").strip()
            if m and m not in seen:
                seen.add(m); headers.append(m)
    sio = io.StringIO()
    w = csv.DictWriter(sio, fieldnames=headers)
    w.writeheader()
    for r in rows:
        w.writerow({k: r.get(k, "") for k in headers})
    return sio.getvalue()

def _rows_to_lines(rows: list[dict]) -> list[dict]:
    """Convert mapped rows to the canonical line schema the AI returns."""
    out: list[dict] = []
    for r in rows or []:
        if not any((r or {}).values()):
            continue

        name = (r.get("product_name_raw") or r.get("product") or r.get("name") or "").strip()
        if not name:
            continue

        try:
            qty = int(re.sub(r"[^\d\-]", "", str(r.get("qty") or 1)) or 1)
        except Exception:
            qty = 1

        unit_cost = to_num(r.get("unit_cost"))
        msrp = to_num(r.get("msrp"))
        upc = (str(r.get("upc")).strip() if r.get("upc") else None)
        asin = (str(r.get("asin")).strip() if r.get("asin") else None)
        category_guess = (str(r.get("category_guess")).strip() if r.get("category_guess") else None)

        out.append({
            "product_name_raw": name,
            "qty": max(1, qty),
            "unit_cost": unit_cost,
            "msrp": msrp,
            "upc": upc,
            "asin": asin,
            "category_guess": category_guess,
        })
    return out
def _norm_key(name: str | None, upc: str | None, asin: str | None) -> tuple:
    def _clean_id(x):
        if not x: return ""
        s = str(x).strip()
        # keep digits for UPC; drop spaces/dashes
        if re.fullmatch(r"[0-9E+.\-]+", s):  # handle scientific notation like 8.88E+11
            try:
                # try to turn into an int-like string if it's scientific
                from decimal import Decimal
                s = str(int(Decimal(s)))
            except Exception:
                s = re.sub(r"\D", "", s)
        return s

    def _clean_name(x):
        return re.sub(r"[\s\W]+", " ", (x or "").lower()).strip()

    return (_clean_id(upc), _clean_id(asin), _clean_name(name))


def _merge_ai_with_local(ai_lines: list[dict], local_lines: list[dict]) -> list[dict]:
    """Fill missing fields in AI lines from local deterministic parse."""
    # index local by (upc, asin, name)
    idx: dict[tuple, dict] = {}
    for ln in local_lines or []:
        k = _norm_key(ln.get("product_name_raw"), ln.get("upc"), ln.get("asin"))
        if k not in idx:
            idx[k] = ln

    merged: list[dict] = []
    for ai in ai_lines or []:
        k = _norm_key(ai.get("product_name_raw"), ai.get("upc"), ai.get("asin"))
        local = idx.get(k, {})

        merged.append({
            "product_name_raw": (ai.get("product_name_raw") or local.get("product_name_raw")),
            "qty": ai.get("qty") or local.get("qty") or 1,
            "unit_cost": (
                to_num(ai.get("unit_cost"))
                if ai.get("unit_cost") is not None
                else to_num(local.get("unit_cost"))
            ),
            "msrp": (
                to_num(ai.get("msrp"))
                if ai.get("msrp") is not None
                else to_num(local.get("msrp"))
            ),
            "upc": ai.get("upc") or local.get("upc"),
            "asin": ai.get("asin") or local.get("asin"),
            "category_guess": ai.get("category_guess") or local.get("category_guess"),
        })
    return merged


# -----------------------------------------------------------------------------
# AI-first upload (falls back to local parser)
# -----------------------------------------------------------------------------
@app.post("/imports/ai-preview/{vendor_id}")
async def ai_preview(
    vendor_id: str,
    po_file: UploadFile = File(...),
    expand_units: bool = Form(False),
):
    if po_file is None:
        raise HTTPException(400, "form field 'po_file' (file) is required")

    filename = po_file.filename or "upload"
    ext = (filename.rsplit(".", 1)[-1] or "").lower()
    content = await po_file.read()

    used_ai = False
    model_name = ""
    ai_notes = ""
    headers_seen: list[str] = []
    lines: list[dict] = []

    try:
        # Model (JSON output mode)
        model, model_name, _structured = make_gemini_model()

        # âœ… Normalize spreadsheets to CSV BEFORE Gemini to avoid XLSX MIME issues
        filename, ext, content = normalize_spreadsheet_upload(filename, ext, content)

        # ðŸ‘‰ Ask Gemini with safe inline parts (CSV -> text/csv, PDF -> inline_data)
        data = _gemini_parse_inline(
            file_bytes=content,
            filename=filename,
            ext=ext,
            expand_units=False,     # AI returns vendor rows; expand server-side
            model=model,
        )
        used_ai = True

        headers_seen = (data or {}).get("detected_headers") or []
        ai_notes = (data or {}).get("notes") or "Parsed directly by Gemini (inline file)."

        # Post-process & (optionally) expand units server-side for consistency
        raw_lines = (data or {}).get("lines") or []
        lines = _postprocess_lines(raw_lines, expand_units=expand_units)

        print(
            f"[AI Preview] Gemini ({model_name}) lines_in={len(raw_lines)} "
            f"lines_out={len(lines)} expand_units={expand_units}",
            flush=True,
        )

    except Exception as e:
        # Deterministic fallback
        print("[AI Preview] AI unavailable/failed; using local preview:", repr(e), flush=True)
        raw_lines = _local_preview_lines_from_bytes(content, ext)
        lines = _postprocess_lines(raw_lines, expand_units=expand_units)
        ai_notes = f"AI unavailable or failed; used local preview. ({type(e).__name__})"

    # Existing PO summary (unchanged)
    with db_conn() as con, con.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT p.id,
                   p.po_number,
                   COALESCE(p.created_at, NOW()) AS created_at,
                   COALESCE(SUM(pl.qty), 0)::int AS total_units,
                   COALESCE(SUM(COALESCE(pl.qty,0) * COALESCE(pl.unit_cost,0)), 0)::numeric(12,2) AS estimated_total_cost
            FROM purchase_orders p
            LEFT JOIN po_lines pl ON pl.purchase_order_id = p.id
            WHERE p.vendor_id = %s
            GROUP BY p.id, p.po_number, p.created_at
            ORDER BY COALESCE(p.created_at, NOW()) DESC, p.id DESC
            LIMIT 6;
        """, (vendor_id,))
        existing = []
        for r in cur.fetchall():
            d = dict(r)
            try:
                d["estimated_total_cost"] = float(d.get("estimated_total_cost") or 0)
            except Exception:
                pass
            existing.append(d)

    payload = {
        "ok": True,
        "vendor_id": vendor_id,
        "file_name": filename,
        "new_po_lines": lines,
        "existing_pos_summary": existing,
        "ai_notes": (ai_notes + (f" | merged_with_local | expand_units={expand_units}"
                   if ai_notes else f"merged_with_local | expand_units={expand_units}")).strip(),
        "headers_seen": headers_seen,
        "via": "gemini" if used_ai else ("xlsx" if ext.startswith("xl") else "csv"),
        "model": model_name,
    }
    return JSONResponse(
        content=jsonable_encoder(payload),
        headers={"X-AI-Used": "1" if used_ai else "0", "X-AI-Model": model_name or ""},
    )

def ai_upload_job(
    job_id: str,
    vendor_id: str,
    po_number: str,
    raw_b64: str,
    original_name: str,
    ext: str,
    expand_units: bool,
    category_id: Optional[str],
    allow_append: bool,
    vendor_name: Optional[str],
):
    # Setup for status updates and imports
    REDIS_URL = os.getenv("REDIS_URL", "redis://redis:6379/0")
    rds = redis.from_url(REDIS_URL, decode_responses=False)
    _pub = lambda evt: _publish(rds, f"ai:commit:{job_id}", evt)

@app.post("/imports/ai-hinted-preview", response_model=UploadPreviewResponse)
async def ai_hinted_preview(
    vendor_id: Optional[str] = Form(None),
    vendor_name: Optional[str] = Form(None),
    hints_json: str = Form(...),
    po_file: Optional[UploadFile] = File(None),
):
    # 1) parse hints
    try:
        hints = HintedPreviewHints.model_validate_json(hints_json)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"invalid hints: {e}")

    # 2) ensure vendor id (create if needed)
    vid = vendor_id
    if not vid:
        vname = (vendor_name or "").strip()
        if not vname:
            raise HTTPException(status_code=400, detail="vendor_id or vendor_name required")
        with db() as (con, cur):
            cur.execute("SELECT id FROM vendors WHERE name=%s LIMIT 1", (vname,))
            row = cur.fetchone()
            if row: vid = row["id"]
            else:
                cur.execute("INSERT INTO vendors(name) VALUES (%s) RETURNING id", (vname,))
                vid = cur.fetchone()["id"]
                con.commit()

    # 3) load rows from file (or later: accept raw rows JSON to avoid re-upload)
    if not po_file:
        raise HTTPException(status_code=400, detail="po_file required for now")
    raw_bytes = await po_file.read()

    # 4) try AI if configured, else deterministic fallback using the hints
    used_ai, model_name, ai_notes = False, None, None
    try:
        if HAVE_GENAI and GENAI_READY:  # whatever flag you already compute
            used_ai = True
            model_name = CURRENT_GENAI_MODEL
            # âœ³ï¸ System prompt: keep it deterministic & schema-bound
            system = """You are a purchase-order normalizer. Given a header row index,
column role mapping, and a 2D table, produce an array of JSON objects:
{product_name_raw, qty, unit_cost, msrp}. Use only provided columns/rows.
If a value is missing, set it null. Never invent items."""
            table = extract_table_from_xlsx(raw_bytes, hints.header_row, hints.selection_rows)  # implement using openpyxl / pandas
            sample = table[:400]  # cap tokens
            ai_input = {
                "column_roles": hints.column_roles,
                "rows": sample,
                "examples": hints.examples or [],
                "expand_units": hints.expand_units or False,
            }
            new_po_lines = call_genai_and_validate(system, ai_input)  # write: calls Gemini, then strict jsonschema validation
            ai_notes = f"AI normalized {len(new_po_lines)} lines from selection."
        else:
            raise RuntimeError("AI not configured")
    except Exception as e:
        # 5) fallback: use hints to parse deterministically
        used_ai = False
        model_name = None
        ai_notes = f"AI disabled or failed: {type(e).__name__}: {e}. Fallback parser used."
        table = extract_table_from_xlsx(raw_bytes, hints.header_row, hints.selection_rows)
        new_po_lines = deterministic_parse(table, hints.column_roles, hints.expand_units or False)

    # 6) respond (existing_pos_summary can be filled with your current helper)
    return {
        "ok": True,
        "vendor_id": vid,
        "file_name": po_file.filename,
        "new_po_lines": new_po_lines,
        "existing_pos_summary": get_existing_po_summaries(vid, limit=5),
        "ai_notes": ai_notes,
        "ai_model": model_name,
    }

@app.post("/imports/ai-upload")
async def ai_upload(
    # either provide the file (first-time upload)...
    po_file: UploadFile | None = File(None),
    file: UploadFile | None = File(None),
    # ...or skip the file and pass the preview result:
    parsed_lines: Optional[str] = Form(None),

    po_number: str = Form(...),
    vendor_id: Optional[str] = Form(None),
    vendor_name: Optional[str] = Form(None),
    expand_units: bool = Form(False),
    category_id: Optional[str] = Form(None),
    allow_append: bool = Form(False),
):
    if not DATABASE_URL:
        raise HTTPException(500, "DATABASE_URL not set")

    # ========== 0) Accept parsed lines from preview to skip AI ==========
    client_lines: list[dict] = []
    if parsed_lines:
        try:
            client_lines = json.loads(parsed_lines)
            if not isinstance(client_lines, list):
                raise ValueError("parsed_lines must be a JSON array")
        except Exception as e:
            raise HTTPException(400, f"parsed_lines invalid JSON: {e}")

    # if we don't have preview lines, we need a file to parse
    f = po_file or file
    if not client_lines and f is None:
        raise HTTPException(400, "Upload a file or supply parsed_lines JSON from preview")

    # ========== 1) Resolve vendor & PO (strict by vendor_id + po_number) ==========
    with db_conn() as con, con.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("BEGIN")

        if not vendor_id:
            vname = (vendor_name or "").strip()
            if not vname:
                con.rollback()
                raise HTTPException(400, "vendor_id or vendor_name is required")
            cur.execute("SELECT id FROM vendors WHERE name=%s LIMIT 1", (vname,))
            v = cur.fetchone()
            if v:
                vendor_id = v["id"]
            else:
                cur.execute("INSERT INTO vendors(name) VALUES (%s) RETURNING id", (vname,))
                vendor_id = cur.fetchone()["id"]
                con.commit()
                cur.execute("BEGIN")

        cur.execute(
            "SELECT id FROM purchase_orders WHERE vendor_id=%s AND po_number=%s LIMIT 1",
            (vendor_id, po_number),
        )
        p = cur.fetchone()

        if p and not allow_append:
            con.rollback()
            raise HTTPException(status_code=409, detail={"error": "DuplicatePO", "id": str(p["id"])})

        if p and allow_append:
            po_id = p["id"]
        else:
            cur.execute(
                "INSERT INTO purchase_orders (po_number, vendor_id, created_at) VALUES (%s,%s,NOW()) RETURNING id",
                (po_number, vendor_id),
            )
            po_id = cur.fetchone()["id"]
            con.commit()

    # ========== 2) Build final lines ==========
    used_ai = False
    ai_model = ""
    ai_notes = ""
    lines: list[dict] = []

    if client_lines:
        # âœ… Use previewâ€™s result; skip AI entirely
        lines = _postprocess_lines(client_lines, expand_units=expand_units)
        ai_notes = "Used preview lines; AI skipped."
    else:
        # Need to parse now (local + try AI once)
        filename = f.filename or "upload"
        ext_in = (filename.rsplit(".", 1)[-1] or "").lower()
        raw = await f.read()

        # normalize spreadsheets to CSV like preview does
        filename, ext, content = normalize_spreadsheet_upload(filename, ext_in, raw)

        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=("." + ext)) as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            local_rows = _parse_locally_from_file(tmp_path, "." + ext)
            local_lines = _rows_to_lines(local_rows)

            try:
                model, ai_model, _ = make_gemini_model()
                data = _gemini_parse_inline(
                    file_bytes=content,
                    filename=filename,
                    ext=ext,
                    expand_units=False,
                    model=model,
                )
                ai_lines = (data or {}).get("lines") or []
                merged = _merge_ai_with_local(ai_lines, local_lines)
                lines = _postprocess_lines(merged, expand_units=expand_units)
                ai_notes = (data or {}).get("notes") or "Parsed via Gemini (inline file)."
                used_ai = True
            except Exception as e:
                lines = _postprocess_lines(local_lines, expand_units=expand_units)
                ai_notes = f"AI unavailable; used local parser. ({type(e).__name__})"
        finally:
            if tmp_path:
                try: os.remove(tmp_path)
                except Exception: pass

    if not lines:
        raise HTTPException(400, "No item lines were parsed or provided.")

    # ========== 3) Insert lines ==========
    created = 0
    with db_conn() as con:
        cur = con.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("BEGIN")
        params = []
        for ln in lines:
            name = (ln.get("product_name_raw") or "").strip()
            if not name:
                continue
            qty = int(ln.get("qty") or 1)
            unit_cost = to_num(ln.get("unit_cost"))
            msrp = to_num(ln.get("msrp"))
            upc = ln.get("upc")
            asin = ln.get("asin")
            resolved_cat_id = category_id or resolve_category_id(cur, ln.get("category_id") or ln.get("category_guess"))
            params.append((po_id, name, upc, asin, qty, unit_cost, msrp, resolved_cat_id, json.dumps(ln, default=str)))

        if params:
            psycopg2.extras.execute_values(
                cur,
                """
                INSERT INTO po_lines
                  (purchase_order_id, product_name_raw, upc, asin, qty, unit_cost, msrp, category_guess, raw_json)
                VALUES %s
                """,
                params,
                page_size=len(params),
            )
        con.commit()
        created = len(params)

    # ========== 4) Return a compat job-shaped response ==========
    job_id = f"inline-{uuid.uuid4().hex}"
    final_notes = (ai_notes + f" | expand_units={expand_units}").strip()

    return {
        "job_id": job_id,           # ðŸ‘ˆ keeps your current UI happy
        "status": "completed",      # ðŸ‘ˆ tells the UI there's nothing to poll
        "ok": True,
        "po_id": str(po_id),
        "created_lines": created,
        "ai_model": ai_model,
        "ai_notes": final_notes,
        "used_ai": used_ai,
    }



def resolve_category_id(cur, guess: str | None) -> Optional[str]:
    """
    Try to map a human label/prefix guess to categories.id (UUID).
    Returns a UUID string if found, else None. Never raises.
    """
    if not guess:
        return None
    q = (str(guess) or "").strip()
    if not q:
        return None
    try:
        # exact match on label or prefix (case-insensitive)
        cur.execute(
            """
            SELECT id
            FROM categories
            WHERE lower(label) = lower(%s) OR lower(prefix) = lower(%s)
            LIMIT 1
            """,
            (q, q),
        )
        row = cur.fetchone()
        if row and row.get("id"):
            return row["id"]

        # starts-with fallback
        cur.execute(
            """
            SELECT id
            FROM categories
            WHERE lower(label) LIKE lower(%s) || '%%'
               OR lower(prefix) LIKE lower(%s) || '%%'
            LIMIT 1
            """,
            (q, q),
        )
        row = cur.fetchone()
        return row["id"] if row and row.get("id") else None
    except Exception:
        return None


# -----------------------------------------------------------------------------
# scrape + sold_avg
# -----------------------------------------------------------------------------


@app.get("/scrape")
def scrape(url: str, hits: int = 15):
    started = time.time()
    hits = max(1, min(200, hits))
    cookie = COOKIE_STORE.get("local", "")
    try:
        r = http_get(url, cookie=cookie)
        soup = soup_from_html(r.text)
        items = soup.select(".srp-results .s-item")
    except Exception as e:
        raise HTTPException(502, str(e))

    should_retry = (not items) or (r is not None and r.status_code >= 500)
    if should_retry:
        try:
            warm = http_get("https://www.ebay.com/")
            warm_jar = dict_from_cookiejar(session.cookies)
            if warm_jar:
                COOKIE_STORE["local"] = "; ".join(f"{k}={v}" for k, v in warm_jar.items())
        except Exception:
            pass

        time.sleep(0.8)
        try:
            r = http_get(url, cookie=COOKIE_STORE.get("local", ""))
            soup = soup_from_html(r.text)
            items = soup.select(".srp-results .s-item")
        except Exception:
            items = None

        if not items:
            pw = scrape_via_playwright(url, hits)
            if pw.get("ok"):
                return {
                    "ok": True,
                    "endpoint": url,
                    "status": r.status_code if isinstance(r, requests.Response) else None,
                    "durationMs": int((time.time() - started) * 1000),
                    "sampled": pw["sampled"],
                    "valid": pw["valid"],
                    "avg": pw["avg"],
                    "rows": pw["rows"],
                    "via": "playwright",
                }
            title_snippet = (soup.title.string if soup and soup.title else "")[:160] if isinstance(soup, BeautifulSoup) else ""
            raise HTTPException(502, f"no_results_from_index; snippet={title_snippet}")

    rows = shape_rows_from_items(items, hits)
    summary = summarize_rows(rows)
    return {
        "ok": True,
        "endpoint": url,
        "status": r.status_code,
        "durationMs": int((time.time() - started) * 1000),
        "sampled": summary["sampled"],
        "valid": summary["valid"],
        "avg": summary["avg"],
        "rows": rows,
        "via": "requests",
    }

@app.post("/purchase_orders", status_code=201)
def purchase_order_create(body: CreatePOBody):
    po_number = (body.po_number or "").strip()
    if not po_number:
        raise HTTPException(400, "po_number is required")
    if not body.vendor_id and not (body.vendor_name and body.vendor_name.strip()):
        raise HTTPException(400, "vendor_id or vendor_name is required")

    with db() as (con, cur):
        try:
            cur.execute("BEGIN")

            # Resolve vendor_id (verify given id OR upsert by name)
            if body.vendor_id:
                vendor_id = body.vendor_id
                cur.execute("SELECT 1 FROM vendors WHERE id=%s", (vendor_id,))
                if not cur.fetchone():
                    raise HTTPException(400, "vendor_id not found")
            else:
                vname = body.vendor_name.strip()
                cur.execute("SELECT id FROM vendors WHERE name=%s LIMIT 1", (vname,))
                row = cur.fetchone()
                if row:
                    vendor_id = row["id"]
                else:
                    cur.execute("INSERT INTO vendors(name) VALUES (%s) RETURNING id", (vname,))
                    vendor_id = cur.fetchone()["id"]

            # STRICT duplicate check by (vendor_id, po_number)
            cur.execute(
                "SELECT id FROM purchase_orders WHERE vendor_id=%s AND po_number=%s LIMIT 1",
                (vendor_id, po_number),
            )
            existing = cur.fetchone()
            if existing:
                con.rollback()
                # Same shape your frontend already expects for 409 handling
                raise HTTPException(status_code=409, detail={"reason": "exists", "id": str(existing["id"])})

            # Create new PO
            cur.execute(
                "INSERT INTO purchase_orders (po_number, vendor_id, created_at) "
                "VALUES (%s, %s, NOW()) RETURNING id",
                (po_number, vendor_id),
            )
            po_id = cur.fetchone()["id"]

            con.commit()
            return {"ok": True, "id": str(po_id), "vendor_id": str(vendor_id)}

        except psycopg2.errors.UniqueViolation:
            # If you later add a DB UNIQUE constraint, handle it gracefully
            con.rollback()
            cur.execute(
                "SELECT id FROM purchase_orders WHERE vendor_id=%s AND po_number=%s LIMIT 1",
                (vendor_id, po_number),
            )
            row = cur.fetchone()
            raise HTTPException(status_code=409, detail={"reason": "exists", "id": str(row["id"]) if row else None})
        except HTTPException:
            con.rollback()
            raise
        except Exception as e:
            con.rollback()
            raise HTTPException(500, f"Database error creating PO: {e}")

@app.get("/sold_avg")
def sold_avg(q: Optional[str] = None, row: Optional[str] = None, limit: int = 60):
    if not q and row:
        try:
            obj = json.loads(row)
            q = build_query_from_row(obj)
        except Exception:
            pass
    if not q:
        raise HTTPException(400, "q or row required")

    hits = max(1, min(200, limit))
    site = "https://www.ebay.com/sch/i.html"
    url = f"{site}?_nkw={quote(q)}&LH_Complete=1&LH_Sold=1&_sop=13"

    started = time.time()
    cookie = COOKIE_STORE.get("local", "")

    try:
        r = http_get(url, cookie=cookie)
        soup = soup_from_html(r.text)
        items = (
            soup.select(".srp-results .s-item")
            or soup.select("ul.srp-results li.s-item")
            or soup.select("li.s-item")
            or soup.select("[data-testid='item-cell']")
        )
    except Exception:
        items = None
        r = None

    if not items:
        pw = scrape_via_playwright(url, hits)
        if pw.get("ok"):
            return {
                "ok": True,
                "endpoint": url,
                "status": r.status_code if isinstance(r, requests.Response) else None,
                "durationMs": int((time.time() - started) * 1000),
                "sampled": pw["sampled"],
                "valid": pw["valid"],
                "avg": pw["avg"],
                "rows": pw["rows"],
                "via": "playwright",
            }
        raise HTTPException(502, "no_results_from_index")

    rows = shape_rows_from_items(items, hits)
    summary = summarize_rows(rows)
    return {
        "ok": True,
        "endpoint": url,
        "status": (r.status_code if isinstance(r, requests.Response) else None),
        "durationMs": int((time.time() - started) * 1000),
        "sampled": summary["sampled"],
        "valid": summary["valid"],
        "avg": summary["avg"],
        "rows": rows[:10],
        "via": "requests",
    }

    EBAY_ITEM_RE = re.compile(r"(?:/itm/|item=)(\d{9,14})")

def _parse_ebay_legacy_id(url: str) -> Optional[str]:
    if not url: 
        return None
    m = EBAY_ITEM_RE.search(url)
    return m.group(1) if m else None

def _http_ok(url: str) -> bool:
    try:
        r = requests.get(url, timeout=6, allow_redirects=True)
        return r.status_code < 400
    except Exception:
        return False



@app.post("/listings/link-ebay")
def link_ebay_listing(body: LinkEbayBody):
    url = (body.ebayUrl or "").strip()
    if not url:
        raise HTTPException(400, "ebayUrl required")

    legacy_id = _parse_legacy_item_id(url)  # may be None; not fatal for storing link
    link_ok = True  # donâ€™t ping HTML; trust link
    status = "ACTIVE"  # optimistic until Browse says ended

    with db() as (con, cur):
        inv_id = None
        if body.inventoryItemId:
            cur.execute("SELECT id FROM public.inventory_items WHERE id = %s", (body.inventoryItemId,))
            r = cur.fetchone()
            inv_id = (r or [None])[0]
        elif body.synergyCode:
            cur.execute("""
                SELECT id FROM public.inventory_items
                WHERE synergy_code = %s
                ORDER BY posted_at NULLS FIRST, created_at NULLS LAST
                LIMIT 1
            """, (body.synergyCode,))
            r = cur.fetchone()
            inv_id = (r or [None])[0]
        elif body.poLineSynergyId:
            cur.execute("""
                SELECT i.id FROM public.inventory_items i
                JOIN public.po_lines pl ON pl.id = i.po_line_id
                WHERE pl.synergy_id::text = %s
                ORDER BY (COALESCE(i.ebay_item_url,'') <> '' OR i.posted_at IS NOT NULL) ASC, i.created_at
                LIMIT 1
            """, (body.poLineSynergyId,))
            r = cur.fetchone()
            inv_id = (r or [None])[0]

        cur.execute("""
            INSERT INTO public.external_listings
                (platform, sku, synergy_code, po_line_id, inventory_item_id, url, ebay_legacy_id,
                 price_num, currency, status)
            SELECT
                'ebay',
                %s,
                %s,
                (SELECT pl.id FROM public.po_lines pl 
                 WHERE %s IS NOT NULL AND pl.synergy_id::text = %s LIMIT 1),
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            RETURNING id
        """, (
            body.sku,
            body.synergyCode,
            body.poLineSynergyId, body.poLineSynergyId,
            inv_id,
            url,
            legacy_id,
            body.price, body.currency, status
        ))
        _ = cur.fetchone()
        if inv_id:
            cur.execute("""
                UPDATE public.inventory_items
                   SET ebay_item_url = COALESCE(%s, ebay_item_url),
                       ebay_price    = COALESCE(%s, ebay_price),
                       posted_at     = COALESCE(posted_at, now()),
                       posted_by     = COALESCE(posted_by, %s::uuid)
                 WHERE id = %s
            """, (url, body.price, body.postedBy, inv_id))
        con.commit()

    # Chain into API-only sync
    return ebay_sync_by_url(EbaySyncBody(synergyId=body.synergyId, ebayUrl=body.ebayUrl))


# (Optional) naive status checker by scraping the public page
@app.get("/listings/{legacy_id}/refresh")
def refresh_ebay_listing_status(legacy_id: str):
    if not legacy_id or not re.fullmatch(r"\d{9,14}", legacy_id):
        raise HTTPException(400, "invalid legacy item id")

    url = f"https://www.ebay.com/itm/{legacy_id}"
    try:
        r = requests.get(url, timeout=8)
        txt = r.text.lower()
        if r.status_code == 404:
            state = "ENDED"
        elif "this listing has ended" in txt or "bids ended" in txt:
            state = "ENDED"
        elif "sold" in txt and "this item is out of stock" in txt:
            state = "SOLD"
        else:
            state = "ACTIVE"
    except Exception:
        state = "UNKNOWN"

    with db() as (con, cur):
        cur.execute("""
            UPDATE public.external_listings
            SET status = %s,
                ended_at = CASE WHEN %s IN ('ENDED','SOLD') THEN COALESCE(ended_at, now()) ELSE ended_at END
            WHERE platform='ebay' AND ebay_legacy_id = %s
        """, (state, state, legacy_id))
        con.commit()

    return {"ok": True, "legacyId": legacy_id, "status": state, "url": url}

# -----------------------------------------------------------------------------
# Browse API average (+ optional Directus cache)
# -----------------------------------------------------------------------------
def _avg_key(q, fixed_price, condition, currency, limit, min_price, max_price, category_ids):
    payload = {
        "q": (q or "").strip().lower(),
        "fixed": bool(fixed_price),
        "condition": (condition or "").upper(),
        "currency": (currency or "").upper(),
        "limit": int(limit or 100),
        "min_price": min_price,
        "max_price": max_price,
        "category_ids": category_ids or "",
    }
    s = json.dumps(payload, sort_keys=True)
    return hashlib.sha1(s.encode()).hexdigest()

def _directus_get(key: str):
    if not (DIRECTUS_URL and DIRECTUS_TOKEN):
        return None
    r = session.get(
        f"{DIRECTUS_URL}/items/market_avg",
        headers={"Authorization": f"Bearer {DIRECTUS_TOKEN}"},
        params={"filter[key][_eq]": key, "limit": 1},
        timeout=(CONNECT_TIMEOUT, READ_TIMEOUT),
    )
    if r.status_code != 200:
        return None
    data = r.json().get("data") or []
    return data[0] if data else None

def _directus_save(record: dict):
    if not (DIRECTUS_URL and DIRECTUS_TOKEN):
        return
    session.post(
        f"{DIRECTUS_URL}/items/market_avg",
        headers={
            "Authorization": f"Bearer {DIRECTUS_TOKEN}",
            "Content-Type": "application/json",
        },
        json=record,
        timeout=(CONNECT_TIMEOUT, READ_TIMEOUT),
    )

@app.get("/browse_avg")
def browse_avg(
    q: str,
    limit: int = 100,
    filter_raw: Optional[str] = Query(None, alias="filter"),
    priceCurrency: str = "USD",
):
    if not q:
        raise HTTPException(400, "q required")

    currency = (priceCurrency or "USD").upper()
    fixed_price = "buyingOptions:{FIXED_PRICE}" in (filter_raw or "") if filter_raw else True
    condition = "USED"
    if filter_raw and "conditions:{" in filter_raw:
        try:
            condition = filter_raw.split("conditions:{", 1)[1].split("}", 1)[0].split("|")[0].strip().upper()
        except Exception:
            pass

    key = _avg_key(q, fixed_price, condition, currency, limit, None, None, None)
    now_ms = int(time.time() * 1000)
    cached = _directus_get(key)
    if cached:
        age = now_ms - int(cached.get("created_at_ms", 0))
        ttl = int(cached.get("ttl_ms", AVG_CACHE_TTL_MS))
        if 0 <= age < ttl:
            return {
                "ok": True,
                "source": "cache",
                "currency": cached.get("currency"),
                "sampled": cached.get("sampled", 0),
                "valid": cached.get("valid", 0),
                "avg": cached.get("avg", 0.0),
                "rows": cached.get("rows") or [],
            }

    if not (EBAY_CLIENT_ID and EBAY_CLIENT_SECRET):
        raise HTTPException(500, "EBAY_CLIENT_ID/SECRET required")

    auth = base64.b64encode(f"{EBAY_CLIENT_ID}:{EBAY_CLIENT_SECRET}".encode()).decode()
    tok = session.post(
        "https://api.ebay.com/identity/v1/oauth2/token",
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Authorization": f"Basic {auth}",
        },
        data="grant_type=client_credentials&scope=https://api.ebay.com/oauth/api_scope",
        timeout=(CONNECT_TIMEOUT, READ_TIMEOUT),
    )
    if tok.status_code != 200:
        raise HTTPException(502, f"oauth_failed: {tok.status_code} {tok.text[:200]}")
    access = tok.json().get("access_token")

    params = {"q": q, "limit": max(1, min(200, int(limit))), "priceCurrency": currency}
    params["filter"] = filter_raw or "buyingOptions:{FIXED_PRICE},conditions:{USED}"

    r = session.get(
        "https://api.ebay.com/buy/browse/v1/item_summary/search",
        params=params,
        headers={"Authorization": f"Bearer {access}"},
        timeout=(CONNECT_TIMEOUT, READ_TIMEOUT),
    )
    if r.status_code != 200:
        raise HTTPException(502, f"browse_failed: {r.status_code} {r.text[:200]}")

    data = r.json()
    rows: List[Dict[str, Any]] = []
    for it in data.get("itemSummaries", []):
        p = it.get("price") or {}
        price_num = None
        if "value" in p:
            try:
                price_num = float(p["value"])
            except Exception:
                price_num = None
        rows.append({
            "title": it.get("title"),
            "url": it.get("itemWebUrl"),
            "priceText": f"{p.get('currency','')} {p.get('value')}",
            "priceNum": price_num,
            "currency": p.get("currency", ""),
        })

    from collections import Counter
    currs = [r["currency"] for r in rows if r.get("currency")]
    dominant = "USD" if "USD" in currs else (Counter(currs).most_common(1)[0][0] if currs else "")
    prices = [r["priceNum"] for r in rows if isinstance(r.get("priceNum"), (int, float)) and r["currency"] == dominant]
    avg = round(sum(prices) / len(prices), 2) if prices else 0.0

    payload = {
        "ok": True,
        "source": "browse_api",
        "currency": dominant,
        "sampled": len(rows),
        "valid": len(prices),
        "avg": avg,
        "rows": rows[:10],
    }

    try:
        _directus_save({
            "key": key,
            "q": q,
            "filters": params.get("filter", ""),
            "currency": dominant,
            "sampled": len(rows),
            "valid": len(prices),
            "avg": avg,
            "rows": rows[:10],
            "created_at_ms": now_ms,
            "ttl_ms": AVG_CACHE_TTL_MS,
        })
    except Exception:
        pass

    return payload

# Aliases with /api prefix (optional)
@app.get("/api/browse_avg")
def browse_avg_alias(
    q: str,
    limit: int = 100,
    filter_raw: Optional[str] = Query(None, alias="filter"),
    priceCurrency: str = "USD",
):
    return browse_avg(q=q, limit=limit, filter_raw=filter_raw, priceCurrency=priceCurrency)

@app.get("/api/sold_avg")
def sold_avg_alias(q: Optional[str] = None, row: Optional[str] = None, limit: int = 60):
    return sold_avg(q=q, row=row, limit=limit)

@app.get("/pricing-config")
def get_pricing_config():
    with db() as (_, cur):
        cur.execute("SELECT data FROM public.pricing_config WHERE id = 1;")
        row = cur.fetchone()
        return {"data": (row["data"] if row and "data" in row else None)}

@app.put("/pricing-config")
def put_pricing_config(body: dict = Body(...)):
    data = body.get("data")
    if data is None or not isinstance(data, dict):
        return JSONResponse({"ok": False, "error": "Body must be { data: { ... } }"}, status_code=400)
    with db() as (con, cur):
        cur.execute(
            """
            INSERT INTO public.pricing_config (id, data, updated_at)
            VALUES (1, %s::jsonb, now())
            ON CONFLICT (id) DO UPDATE
            SET data = EXCLUDED.data, updated_at = now()
            """,
            (_json.dumps(data),),
        )
        con.commit()
    return {"ok": True}

# --- MSRP lookup (fallback = 999) --------------------------------------------
@app.get("/msrp")
def get_msrp(model_key: str = Query(...), year: int = Query(...)):
    with db() as (_, cur):
        cur.execute(
            "SELECT msrp_usd FROM public.msrp_reference WHERE model_key = %s AND year = %s;",
            (model_key, year),
        )
        row = cur.fetchone()
    msrp = float(row["msrp_usd"]) if row and row.get("msrp_usd") is not None else 999.00
    return {"model_key": model_key, "year": year, "msrp": round(msrp, 2)}

# -----------------------------------------------------------------------------
# Categories (UI fallback when Directus off)
# -----------------------------------------------------------------------------
class CategoryBody(BaseModel):
    label: str
    prefix: str
    notes: Optional[str] = None

@app.get("/categories")
def categories_list():
    with db() as (con, cur):
        cur.execute("SELECT id, label, prefix, notes FROM categories ORDER BY label;")
        return [dict(r) for r in cur.fetchall()]

@app.post("/categories")
def categories_create(body: CategoryBody):
    with db() as (con, cur):
        cur.execute(
            "INSERT INTO categories(label, prefix, notes) VALUES (%s,%s,%s) RETURNING id,label,prefix,notes",
            (body.label, body.prefix, body.notes),
        )
        con.commit()
        return dict(cur.fetchone())

@app.patch("/categories/{cat_id}")
def categories_patch(cat_id: str, body: dict):
    fields, vals = [], []
    for k in ("label","prefix","notes"):
        if k in body:
            fields.append(f"{k}=%s"); vals.append(body[k])
    if not fields:
        raise HTTPException(400, "no fields")
    vals.append(cat_id)
    with db() as (con, cur):
        cur.execute(
            f"UPDATE categories SET {', '.join(fields)} WHERE id=%s RETURNING id,label,prefix,notes",
            vals,
        )
        if cur.rowcount == 0: raise HTTPException(404, "not found")
        con.commit()
        return dict(cur.fetchone())

@app.delete("/categories/{cat_id}")
def categories_delete(cat_id: str):
    with db() as (con, cur):
        cur.execute("DELETE FROM categories WHERE id=%s", (cat_id,))
        con.commit()
        return {"ok": cur.rowcount > 0}

# -----------------------------------------------------------------------------
# Vendors CRUD + counts + vendor â†’ POs
# -----------------------------------------------------------------------------
class VendorIn(BaseModel):
    name: str

class VendorUpdate(BaseModel):
    name: str

@app.get("/vendors")
def vendors_list(q: Optional[str] = Query(None)):
    with db() as (con, cur):
        if q:
            cur.execute("""
                SELECT v.id, v.name, COALESCE(COUNT(p.id),0) AS po_count
                FROM vendors v
                LEFT JOIN purchase_orders p ON p.vendor_id = v.id
                WHERE v.name ILIKE %s
                GROUP BY v.id, v.name
                ORDER BY v.name ASC
            """, (f"%{q}%",))
        else:
            cur.execute("""
                SELECT v.id, v.name, COALESCE(COUNT(p.id),0) AS po_count
                FROM vendors v
                LEFT JOIN purchase_orders p ON p.vendor_id = v.id
                GROUP BY v.id, v.name
                ORDER BY v.name ASC
            """)
        return [dict(r) for r in cur.fetchall()]

@app.post("/vendors")
def create_vendor(body: VendorIn):
    with db() as (con, cur):
        cur.execute("""
            INSERT INTO vendors (name) VALUES (%s)
            ON CONFLICT (name) DO UPDATE SET name = EXCLUDED.name
            RETURNING id, name;
        """, (body.name.strip(),))
        con.commit()
        return dict(cur.fetchone())

@app.patch("/vendors/{vendor_id}")
def update_vendor(vendor_id: str, body: VendorUpdate):
    with db() as (con, cur):
        cur.execute("""
            UPDATE vendors SET name = %s
            WHERE id = %s
            RETURNING id, name;
        """, (body.name.strip(), vendor_id))
        con.commit()
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Vendor not found")
        return dict(row)

@app.delete("/vendors/{vendor_id}")
def delete_vendor(vendor_id: str):
    with db() as (con, cur):
        try:
            cur.execute("DELETE FROM vendors WHERE id = %s RETURNING id;", (vendor_id,))
            con.commit()
        except Exception as e:
            raise HTTPException(409, f"Cannot delete vendor with existing POs: {e}")
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Vendor not found")
        return {"ok": True, "id": row["id"]}

@app.get("/vendors/{vendor_id}/pos")
def vendor_pos(vendor_id: str = Path(...)):
    with db() as (con, cur):
        cur.execute("""
            SELECT p.id, p.po_number,
                   COALESCE(SUM(pl.qty),0) AS total_lines_qty,
                   COALESCE(SUM(pl.qty * COALESCE(pl.unit_cost,0)),0)::numeric(12,2) AS est_cost
            FROM purchase_orders p
            LEFT JOIN po_lines pl ON pl.purchase_order_id = p.id
            WHERE p.vendor_id = %s
            GROUP BY p.id, p.po_number
            ORDER BY p.po_number DESC NULLS LAST, p.id DESC
        """, (vendor_id,))
        return [dict(r) for r in cur.fetchall()]

        

# -----------------------------------------------------------------------------
# PO summaries, summary, lines, profit, groups
# -----------------------------------------------------------------------------
@app.get("/pos/summaries")
def po_summaries():
    with db() as (con, cur):
        cur.execute("""
            SELECT
              p.id,
              p.po_number,
              v.name AS vendor_name,
              p.created_at
            FROM purchase_orders p
            LEFT JOIN vendors v ON v.id = p.vendor_id
            ORDER BY p.created_at DESC NULLS LAST,
                     p.po_number DESC NULLS LAST,
                     p.id DESC
            LIMIT 200
        """)
        return [dict(r) for r in cur.fetchall()]

@app.get("/pos/{po_id}/summary")
def po_summary(po_id: str):
    with db() as (con, cur):
        # 1) Header
        cur.execute(
            """
            SELECT
              p.id,
              p.po_number,
              p.vendor_id,
              v.name AS vendor_name,
              p.created_at
            FROM purchase_orders p
            LEFT JOIN vendors v ON v.id = p.vendor_id
            WHERE p.id = %s
            """,
            (po_id,),
        )
        header = cur.fetchone()
        if not header:
            raise HTTPException(404, "PO not found")

        summary = dict(header)

        # 2) Line totals (from po_lines)
        cur.execute(
            """
            SELECT
              COUNT(*)                         AS line_count,
              COALESCE(SUM(qty), 0)::int       AS qty_total,
              COALESCE(SUM(qty * unit_cost),0) AS cost_total,
              COALESCE(SUM(qty * msrp), 0)     AS msrp_total
            FROM po_lines
            WHERE purchase_order_id = %s
            """,
            (po_id,),
        )
        summary.update(dict(cur.fetchone()))

        # 3) Inventory status counts (from inventory_items)
        cur.execute(
            """
            SELECT
              COUNT(*)::int                                               AS minted,
              COUNT(*) FILTER (WHERE status = 'INTAKE')::int              AS intake,
              COUNT(*) FILTER (WHERE status = 'TESTING')::int             AS testing,
              COUNT(*) FILTER (WHERE status = 'TESTED')::int              AS tested,
              COUNT(*) FILTER (WHERE status = 'PRICED')::int              AS priced,
              COUNT(*) FILTER (WHERE status = 'RESERVED')::int            AS reserved,
              COUNT(*) FILTER (WHERE status = 'SOLD')::int                AS sold,
              COUNT(*) FILTER (WHERE status = 'DONATED')::int             AS donated,
              COUNT(*) FILTER (WHERE status = 'SCRAPPED')::int            AS scrapped,
              COUNT(*) FILTER (WHERE status = 'LOST')::int                AS lost
            FROM inventory_items
            WHERE purchase_order_id = %s
            """,
            (po_id,),
        )
        summary.update(dict(cur.fetchone()))

        # 4) Optional testing window (use tested_at; if your column is tested_date, swap the name below)
        cur.execute(
            """
            SELECT
              MIN(tested_at) AS first_tested_at,
              MAX(tested_at) AS last_tested_at
            FROM inventory_items
            WHERE purchase_order_id = %s AND tested_at IS NOT NULL
            """,
            (po_id,),
        )
        test_window = cur.fetchone()
        if test_window:
            summary.update({k: v for k, v in dict(test_window).items() if v is not None})

        return summary



class LineUpdate(BaseModel):
    product_name_raw: Optional[str] = None
    qty: Optional[int] = None
    unit_cost: Optional[float] = None
    category_id: Optional[str] = None  # stored in po_lines.category_guess

@app.get("/pos/{po_id}/lines")
def get_po_lines(po_id: str):
    with db() as (con, cur):
        cur.execute("""
            SELECT pl.id,
                   pl.product_name_raw,
                   pl.upc,
                   pl.asin,
                   pl.qty,
                   pl.unit_cost,
                   pl.msrp,
                   pl.category_guess AS category_id,
                   pl.synergy_id::text AS synergy_id
            FROM po_lines pl
            WHERE pl.purchase_order_id = %s
            ORDER BY pl.id ASC;
        """, (po_id,))
        return [dict(r) for r in cur.fetchall()]

@app.patch("/pos/lines/{line_id}")
def patch_po_line(line_id: str, body: LineUpdate):
    sets = []
    vals: List[object] = []

    if body.product_name_raw is not None:
        sets.append("product_name_raw = %s")
        vals.append(body.product_name_raw)
    if body.qty is not None:
        sets.append("qty = %s")
        vals.append(body.qty)
    if body.unit_cost is not None:
        sets.append("unit_cost = %s")
        vals.append(body.unit_cost)
    if body.category_id is not None:
        sets.append("category_guess = %s")
        vals.append(body.category_id)

    with db() as (con, cur):
        if not sets:
            cur.execute("""
                SELECT id, product_name_raw, upc, asin, qty, unit_cost, msrp,
                       category_guess AS category_id
                FROM po_lines WHERE id = %s;
            """, (line_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(404, "PO line not found")
            return dict(row)

        q = f"""
            UPDATE po_lines SET {", ".join(sets)}
            WHERE id = %s
            RETURNING id, product_name_raw, upc, asin, qty, unit_cost, msrp,
                      category_guess AS category_id;
        """
        vals.append(line_id)
        cur.execute(q, tuple(vals))
        con.commit()
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "PO line not found")
        return dict(row)

def _resolve_po_id(cur, po_ref: str) -> str:
    """
    Accepts either a UUID id or a human PO number.
    Returns the internal purchase_orders.id or raises HTTPException(404).
    """
    # Try UUID path first
    try:
        uuid_obj = uuid.UUID(po_ref)
        cur.execute("SELECT id FROM purchase_orders WHERE id = %s LIMIT 1", (str(uuid_obj),))
        row = cur.fetchone()
        if row:
            return row["id"]
    except Exception:
        pass  # not a uuid or not found

    # Fallback: treat as PO number (case/space-insensitive)
    cur.execute(
        """
        SELECT id
        FROM purchase_orders
        WHERE LOWER(TRIM(po_number)) = LOWER(TRIM(%s))
        LIMIT 1
        """,
        (po_ref,),
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail=f"Purchase order '{po_ref}' not found")
    return row["id"]


@app.get("/pos/{po_ref}/profit")
def unified_po_profit(po_ref: str):
    """
    Universal profit endpoint:
      - /pos/<UUID>/profit
      - /pos/<PO_NUMBER>/profit   (e.g., /pos/MAUNAL/profit)

    Returns numeric fields even if some sources are missing.
    """
    with db() as (con, cur):
        # Resolve id from either UUID or PO number
        try:
            po_id = _resolve_po_id(cur, po_ref)
        except HTTPException as e:
            # return JSON error
            raise
        except Exception as e:
            return JSONResponse(status_code=500, content={"error": "resolve_failed", "detail": str(e)})

        payload: dict[str, object] = {}

        # 1) Optional: start with your view (if present)
        try:
            cur.execute(
                "SELECT * FROM vw_po_profit WHERE purchase_order_id=%s LIMIT 1",
                (po_id,),
            )
            r = cur.fetchone()
            if r:
                payload.update(dict(r))
        except Exception:
            pass  # view may not exist; ignore

        # 2) Baseline totals from po_lines (always available for this PO)
        try:
            cur.execute(
                """
                SELECT
                  COALESCE(SUM(COALESCE(qty,1)), 0)                           AS total_units,
                  COALESCE(SUM(COALESCE(unit_cost,0) * COALESCE(qty,1)), 0.0) AS total_inventory_cost
                FROM po_lines
                WHERE purchase_order_id = %s
                """,
                (po_id,),
            )
            base = cur.fetchone() or {}
        except Exception as e:
            base = {}
            payload["_base_error"] = f"po_lines aggregate failed: {e.__class__.__name__}"

        payload["total_units"] = int(payload.get("total_units") or base.get("total_units") or 0)
        payload["total_inventory_cost"] = float(
            payload.get("total_inventory_cost") or base.get("total_inventory_cost") or 0.0
        )

        # Ensure legacy keys exist for UI
        if payload.get("units_sold") is None: payload["units_sold"] = 0
        if payload.get("sales_net_revenue") is None: payload["sales_net_revenue"] = 0.0
        if payload.get("gross_profit") is None: payload["gross_profit"] = 0.0
        if payload.get("cost_in_unsold_inventory") is None: payload["cost_in_unsold_inventory"] = 0.0

        # 3) Posted/unposted (NULL-safe + qty-weighted). Treat no matching item as NOT posted.
        try:
            cur.execute(
             """
             WITH inv AS (
               SELECT
                 ii.po_line_id,
                 CASE
                   WHEN (ii.ebay_item_url IS NOT NULL AND ii.ebay_item_url <> '')
                     OR ii.posted_at IS NOT NULL
                   THEN TRUE ELSE FALSE
                 END AS posted,
                 COALESCE(ii.ebay_price, 0.0) AS ebay_price
               FROM public.inventory_items ii
             ),
             join_items AS (
               SELECT
                 pl.id AS po_line_id,
                 pl.unit_cost,
                 COALESCE(inv.posted, FALSE) AS posted,
                 inv.ebay_price
               FROM public.po_lines pl
               LEFT JOIN inv ON inv.po_line_id = pl.id
               WHERE pl.purchase_order_id = %s
             )
             SELECT
               COUNT(*) FILTER (WHERE posted) AS units_posted,
               COUNT(*) FILTER (WHERE NOT posted) AS units_unposted,
               COALESCE(SUM(ebay_price) FILTER (WHERE posted), 0.0) AS posted_value,
               COALESCE(SUM(unit_cost) FILTER (WHERE NOT posted), 0.0) AS unposted_cost
             FROM join_items;
             """,
                (po_id,),
            )
            agg = cur.fetchone() or {}
        except Exception as e:
            agg = {}
            payload["_agg_error"] = f"posted/unposted aggregate failed: {e.__class__.__name__}"

        payload["units_posted"]   = int(agg.get("units_posted") or 0)
        payload["units_unposted"] = int(agg.get("units_unposted") or 0)
        payload["posted_value"]   = float(agg.get("posted_value") or 0.0)

        unposted_cost = float(agg.get("unposted_cost") or 0.0)
        payload["unposted_cost"] = unposted_cost
        if not payload.get("cost_in_unsold_inventory"):
            payload["cost_in_unsold_inventory"] = unposted_cost

        # Helpful echo so curl shows what we resolved
        payload["_resolved_po_id"] = po_id
        payload["_ref"] = po_ref

        return payload



@app.get("/pos/{po_id}/groups")
def po_groups(po_id: str):
    with db() as (con, cur):
        cur.execute("""
            SELECT
              COALESCE(pl.category_guess::text, 'UNASSIGNED') AS category_id,
              COUNT(*)                              AS line_count,
              SUM(COALESCE(pl.qty,1))               AS units,
              ROUND(SUM(COALESCE(pl.unit_cost,0) * COALESCE(pl.qty,1))::numeric, 2) AS total_cost,
              (SELECT label FROM categories c WHERE c.id = pl.category_guess) AS label,
              (SELECT prefix FROM categories c WHERE c.id = pl.category_guess) AS prefix
            FROM po_lines pl
            WHERE pl.purchase_order_id = %s
            GROUP BY category_id
            ORDER BY label NULLS LAST, category_id
        """, (po_id,))
        return [dict(r) for r in cur.fetchall()]

# -----------------------------------------------------------------------------
# Prefix counters
# -----------------------------------------------------------------------------
@app.get("/prefix/{prefix}/peek")
def prefix_peek(prefix: str):
    if not DATABASE_URL:
        raise HTTPException(500, "DATABASE_URL not set")
    with db() as (con, cur):
        cur.execute("SELECT next_seq FROM id_prefix_counters WHERE prefix=%s", (prefix,))
        row = cur.fetchone()
        next_seq = row["next_seq"] if row else 1
        return Response(content=fmt_synergy(prefix, next_seq), media_type="text/plain")

@app.post("/prefix/{prefix}/take")
def prefix_take(prefix: str):
    if not DATABASE_URL:
        raise HTTPException(500, "DATABASE_URL not set")
    with db() as (con, cur):
        cur.execute("BEGIN")
        cur.execute("SELECT next_seq FROM id_prefix_counters WHERE prefix=%s FOR UPDATE", (prefix,))
        row = cur.fetchone()
        if row:
            next_seq = row["next_seq"]
            cur.execute("UPDATE id_prefix_counters SET next_seq = next_seq + 1 WHERE prefix=%s", (prefix,))
        else:
            next_seq = 1
            cur.execute("INSERT INTO id_prefix_counters(prefix, next_seq) VALUES (%s,%s)", (prefix, 2))
        con.commit()
        return Response(content=fmt_synergy(prefix, next_seq), media_type="text/plain")

@app.post("/prefix/{prefix}/set")
def prefix_set(prefix: str, body: Dict[str, Any]):
    if not DATABASE_URL:
        raise HTTPException(500, "DATABASE_URL not set")
    nxt = int(body.get("next", 1))
    if nxt < 1:
        raise HTTPException(400, "next must be >=1")
    with db() as (con, cur):
        cur.execute(
            """
            INSERT INTO id_prefix_counters(prefix, next_seq) VALUES (%s,%s)
            ON CONFLICT (prefix) DO UPDATE SET next_seq = EXCLUDED.next_seq
            """,
            (prefix, nxt),
        )
        con.commit()
        return {"ok": True, "next": nxt}

# -----------------------------------------------------------------------------
# Imports: preview (vendor) â€” this is the one your UI calls with a file
# -----------------------------------------------------------------------------

@app.post("/imports/preview/{vendor_id}")
async def imports_preview(vendor_id: str, po_file: UploadFile = File(...)):
    # Defensive guards so we fail fast with a readable 4xx/5xx
    if po_file is None:
        raise HTTPException(400, "form field 'po_file' (file) is required")

    filename = po_file.filename or "upload"
    ext = (filename.rsplit(".", 1)[-1] or "").lower()

    try:
        content = await po_file.read()
    except Exception as e:
        print("[/imports/preview] read() failed:", repr(e), file=sys.stderr)
        traceback.print_exc()
        raise HTTPException(400, f"could_not_read_upload: {e}")

    rows = []
    try:
        if ext in ("csv", "tsv", "txt"):
            text = content.decode("utf-8", errors="ignore")
            try:
                dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
                reader = csv.DictReader(io.StringIO(text), dialect=dialect)
            except Exception:
                reader = csv.DictReader(io.StringIO(text))
            for raw in reader:
                r = {}
                for k, v in raw.items():
                    r[map_header(k)] = v
                rows.append(r)

        elif ext in ("xlsx", "xlsm", "xltx", "xltm"):
            if load_workbook is None:
                # <- This is the openpyxl case
                raise HTTPException(400, "openpyxl not installed (needed for .xlsx preview)")
            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            hdr = [map_header(str(c) if c is not None else "")
                   for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            for rr in ws.iter_rows(min_row=2, values_only=True):
                obj = {}
                for idx, val in enumerate(rr):
                    key = hdr[idx] if idx < len(hdr) else f"col_{idx}"
                    obj[key] = val
                rows.append(obj)

        else:
            raise HTTPException(400, "Unsupported file type. Use CSV or XLSX.")

    except HTTPException:
        raise
    except Exception as e:
        print("[/imports/preview] parse failed:", repr(e), file=sys.stderr)
        traceback.print_exc()
        raise HTTPException(400, f"parse failed: {e}")

    preview_lines = []
    for r in rows:
        if not any(filter(None, r.values())):
            continue
        preview_lines.append({
            "product_name_raw": (r.get("product_name_raw") or r.get("product") or r.get("name") or ""),
            "upc": r.get("upc"),
            "asin": r.get("asin"),
            "qty": int(re.sub(r"[^0-9]", "", str(r.get("qty") or 1)) or 1),
            "unit_cost": to_num(r.get("unit_cost")),
            "msrp": to_num(r.get("msrp")),
        })

    with db_conn() as con, con.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT p.id,
                   p.po_number,
                   COALESCE(p.created_at, NOW()) AS created_at,
                   COALESCE(SUM(pl.qty), 0)::int AS total_units,
                   COALESCE(SUM(COALESCE(pl.qty,0) * COALESCE(pl.unit_cost,0)), 0)::numeric(12,2) AS estimated_total_cost
            FROM purchase_orders p
            LEFT JOIN po_lines pl ON pl.purchase_order_id = p.id
            WHERE p.vendor_id = %s
            GROUP BY p.id, p.po_number, p.created_at
            ORDER BY COALESCE(p.created_at, NOW()) DESC, p.id DESC
            LIMIT 6;
        """, (vendor_id,))
        existing = [dict(r) for r in cur.fetchall()]

    return {
        "ok": True,
        "vendor_id": vendor_id,
        "file_name": filename,
        "new_po_lines": preview_lines,
        "existing_pos_summary": existing,
        # Optional: tell UI what parser path it took
        "via": "xlsx" if ext.startswith("xl") else "csv",
    }



# (Rename of the placeholder to avoid route collision; keeping functionality)
@app.post("/imports/preview/by-po/{po_id}")
def import_preview_by_po(po_id: str):
    return {"ok": True, "po_id": po_id, "message": "Preview logic to be implemented"}

# -----------------------------------------------------------------------------
# Imports: upload (non-AI)
# -----------------------------------------------------------------------------
@app.post("/imports/upload")
async def import_upload(
    file: UploadFile = File(...),
    po_number: str = Form(...),
    vendor_name: Optional[str] = Form(None),
    category_id: Optional[str] = Form(None),
):
    if not DATABASE_URL:
        raise HTTPException(500, "DATABASE_URL not set")
    if not file or not po_number:
        raise HTTPException(400, "Missing 'file' or 'po_number'")

    filename = file.filename or "upload"
    ext = (filename.split(".")[-1] or "").lower()
    content = await file.read()

    rows: List[Dict[str, Any]] = []
    try:
        if ext in ("csv", "tsv", "txt"):
            text = content.decode("utf-8", errors="ignore")
            try:
                dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
                reader = csv.DictReader(io.StringIO(text), dialect=dialect)
            except Exception:
                reader = csv.DictReader(io.StringIO(text))
            for raw in reader:
                r = {}
                for k, v in raw.items():
                    r[map_header(k)] = v
                rows.append(r)
        elif ext in ("xlsx", "xlsm", "xltx", "xltm"):
            if load_workbook is None:
                raise HTTPException(500, "openpyxl not installed")
            wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            hdr = [
                map_header(str(c) if c is not None else "")
                for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            ]
            for rr in ws.iter_rows(min_row=2, values_only=True):
                obj = {}
                for idx, val in enumerate(rr):
                    key = hdr[idx] if idx < len(hdr) else f"col_{idx}"
                    obj[key] = val
                rows.append(obj)
        else:
            raise HTTPException(400, "Unsupported file type. Use CSV or XLSX.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"parse failed: {e}")

    with db() as (con, cur):
        try:
            cur.execute("BEGIN")

            # vendor
            vendor_id = None
            if vendor_name:
                cur.execute("SELECT id FROM vendors WHERE name=%s LIMIT 1", (vendor_name,))
                v = cur.fetchone()
                if v:
                    vendor_id = v["id"]
                else:
                    cur.execute("INSERT INTO vendors(name) VALUES (%s) RETURNING id", (vendor_name,))
                    vendor_id = cur.fetchone()["id"]

            # purchase order (upsert by po_number)
            cur.execute("SELECT id FROM purchase_orders WHERE po_number=%s LIMIT 1", (po_number,))
            p = cur.fetchone()
            if p:
                po_id = p["id"]
                cur.execute(
                    "UPDATE purchase_orders SET created_at = COALESCE(created_at, NOW()) WHERE id=%s",
                    (po_id,),
                )
            else:
                cur.execute(
                    "INSERT INTO purchase_orders(po_number, vendor_id, created_at) VALUES (%s,%s,NOW()) RETURNING id",
                    (po_number, vendor_id),
                )
                po_id = cur.fetchone()["id"]

            # optional category validation
            if category_id:
                cur.execute("SELECT 1 FROM categories WHERE id=%s", (category_id,))
                if not cur.fetchone():
                    raise HTTPException(400, "category_id not found")

            created = 0
            for r in rows:
                if not any(filter(None, r.values())):
                    continue
                line = {
                    "purchase_order_id": po_id,
                    "product_name_raw": r.get("product_name_raw"),
                    "upc": r.get("upc"),
                    "asin": r.get("asin"),
                    "msrp": to_num(r.get("msrp")),
                    "unit_cost": to_num(r.get("unit_cost")),
                    "qty": int(re.sub(r"[^0-9]", "", str(r.get("qty") or 1)) or 1),
                    "category_guess": category_id,
                    "raw_json": json.dumps(r, default=str),
                }
                cols = list(line.keys())
                cur.execute(
                    f"INSERT INTO po_lines ({','.join(cols)}) VALUES ({','.join(['%s']*len(cols))})",
                    [line[c] for c in cols],
                )
                created += 1

            con.commit()
        except HTTPException:
            con.rollback()
            raise
        except Exception as e:
            con.rollback()
            raise HTTPException(500, str(e))

    return {
        "ok": True,
        "po_id": str(po_id),
        "vendor_id": (str(vendor_id) if vendor_id else None),
        "category_id": category_id,
        "created_lines": created,
    }

# -----------------------------------------------------------------------------
# Explode (by-line + group)
# -----------------------------------------------------------------------------
class ExplodeBody(BaseModel):
    useLineCategory: bool = False
    categoryId: Optional[str] = None
    prefix: Optional[str] = None

class ExplodeGroupBody(BaseModel):
    categoryId: Optional[str] = None
    prefix: Optional[str] = None

def _take_synergy(cur, prefix: str) -> str:
    cur.execute("SELECT next_seq FROM id_prefix_counters WHERE prefix=%s FOR UPDATE", (prefix,))
    r = cur.fetchone()
    if r:
        next_seq = r["next_seq"]
        cur.execute("UPDATE id_prefix_counters SET next_seq = next_seq + 1 WHERE prefix=%s", (prefix,))
    else:
        next_seq = 1
        cur.execute("INSERT INTO id_prefix_counters(prefix, next_seq) VALUES (%s,%s)", (prefix, 2))
    return fmt_synergy(prefix, next_seq)

def explode_po_endpoint(po_id: str, body: ExplodeBody):
    if not DATABASE_URL:
        raise HTTPException(500, "DATABASE_URL not set")
    with db() as (con, cur):
        cat_id = body.categoryId or None
        default_prefix = body.prefix or "GEN"
        if cat_id:
            cur.execute("SELECT prefix FROM categories WHERE id=%s", (cat_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(400, "categoryId not found")
            if row["prefix"]:
                default_prefix = row["prefix"]

        cur.execute(
            """
            SELECT pl.id, pl.purchase_order_id, pl.msrp, pl.unit_cost, pl.qty,
                   CASE WHEN %s THEN (SELECT c.prefix FROM categories c WHERE c.id = pl.category_guess)
                        ELSE NULL END AS line_prefix,
                   CASE WHEN %s THEN pl.category_guess ELSE %s::uuid END AS category_id
            FROM po_lines pl
            WHERE pl.purchase_order_id = %s
            """,
            (body.useLineCategory, body.useLineCategory, cat_id, po_id),
        )
        lines = cur.fetchall()
        if not lines:
            return {"ok": True, "created": 0}

        created = 0
        try:
            cur.execute("BEGIN")
            for ln in lines:
                prefix = ln["line_prefix"] or default_prefix
                if not prefix:
                    con.rollback()
                    raise HTTPException(
                        400,
                        "No prefix for one or more lines; pass categoryId or set per-line category",
                    )
                qty = int(ln["qty"] or 1)
                for _ in range(qty):
                    code = _take_synergy(cur, prefix)
                    cur.execute(
                        """
                        INSERT INTO public.inventory_items
                          (synergy_code, purchase_order_id, po_line_id, category_id, cost_unit, msrp, status)
                        VALUES (%s,%s,%s,%s,%s,%s,'INTAKE')
                        """,
                        (
                            code,
                            ln["purchase_order_id"],
                            ln["id"],
                            ln["category_id"],
                            ln["unit_cost"] or 0,
                            ln["msrp"] or 0,
                        ),
                    )
                    created += 1
            con.commit()
        except HTTPException:
            con.rollback()
            raise
        except Exception as e:
            con.rollback()
            raise HTTPException(500, str(e))

        return {"ok": True, "created": created}

@app.get("/pos/{po_id}/mint-stats")
def mint_stats(po_id: str):
    with db() as (con, cur):
        cur.execute("""
            SELECT COALESCE(SUM(COALESCE(qty,1)),0) AS total_qty
            FROM public.po_lines
            WHERE purchase_order_id = %s
        """, (po_id,))
        total_qty = int((cur.fetchone() or {}).get("total_qty") or 0)

        cur.execute("""
            SELECT COUNT(*) AS minted
            FROM public.inventory_items
            WHERE purchase_order_id = %s
        """, (po_id,))
        minted = int((cur.fetchone() or {}).get("minted") or 0)

        pending = max(0, total_qty - minted)
        return {"total_qty": total_qty, "minted": minted, "pending": pending}


# Idempotent explode: only create the delta that isn't already in inventory_items
@app.post("/imports/{po_id}/explode-by-line")
def explode_by_line(po_id: str):
    """
    Create intake rows (inventory_items) for each PO line up to its qty.
    - If a line already has N rows, and qty == N, we do nothing.
    - If qty increased or a new line was added, we create the missing rows only.
    Safe to call multiple times (idempotent).
    """
    with db() as (con, cur):
        # Existing counts per line for this PO
        cur.execute(
            """
            SELECT po_line_id, COUNT(*) AS n
            FROM public.inventory_items
            WHERE purchase_order_id = %s
            GROUP BY po_line_id
            """,
            (po_id,),
        )
        existing_counts = {str(r["po_line_id"]): int(r["n"]) for r in cur.fetchall()}

        cur.execute(
         """
         SELECT
           pl.id,
           COALESCE(pl.qty, 1)              AS qty,
           pl.unit_cost,
           pl.msrp,
           pl.category_guess                AS category_id, 
           COALESCE(c.prefix, '')           AS prefix
         FROM public.po_lines pl
         LEFT JOIN public.categories c
           ON c.id = pl.category_guess     
         WHERE pl.purchase_order_id = %s
         ORDER BY pl.id
         """,
      (po_id,),
  )
        rows = cur.fetchall()

        created = 0
        skipped = 0

        try:
            cur.execute("BEGIN")
            for ln in rows:
                line_id   = str(ln["id"])
                qty       = int(ln["qty"] or 1)
                prefix    = (ln["prefix"] or "").strip()
                cat_id    = ln["category_id"]

                # Need a category + prefix to mint
                if not cat_id or not prefix:
                    skipped += qty
                    continue

                have = int(existing_counts.get(line_id, 0))
                need = max(0, qty - have)
                if need <= 0:
                    skipped += qty
                    continue

                # Create only the delta
                for _ in range(need):
                    code = _take_synergy(cur, prefix)
                    cur.execute(
                        """
                        INSERT INTO public.inventory_items
                          (synergy_code, purchase_order_id, po_line_id, category_id, cost_unit, msrp, status)
                        VALUES (%s, %s, %s, %s, %s, %s, 'INTAKE')
                        """,
                        (
                            code,
                            po_id,
                            line_id,
                            cat_id,
                            ln["unit_cost"] or 0,
                            ln["msrp"] or 0,
                        ),
                    )
                    created += 1

            con.commit()
        except Exception:
            con.rollback()
            raise

        # Simple state flag for the UI
        state = "done" if created > 0 and skipped == 0 else ("partial" if created > 0 else "already")
        return {"ok": True, "created": created, "skipped": skipped, "state": state}



@app.post("/imports/ai-preview-stream/{vendor_id}")
async def ai_preview_stream(
    vendor_id: str,
    po_file: UploadFile = File(...),
    require_ai: bool = Form(False),
    expand_units: bool = Form(False),
    limit_rows: int = Form(120),
):
    # Read upload BEFORE streaming
    try:
        raw = await po_file.read()
    except Exception as e:
        return StreamingResponse(
            iter([_sse({"type": "error", "error_type": type(e).__name__, "message": str(e)})]),
            media_type="text/event-stream",
        )

    filename_in = po_file.filename or "upload"
    ext_in = (filename_in.rsplit(".", 1)[-1] or "").lower()

    async def run():
        tmp_path = None
        try:
            # 1) announce + heartbeat
            yield _sse({"type": "progress", "pct": 2, "label": "Receiving file"})
            yield _heartbeat()

            # 2) normalize for local parsing & Gemini (XLSXâ†’CSV)
            yield _sse({"type": "progress", "pct": 8, "label": "Normalizing spreadsheet"})
            filename, ext, content = normalize_spreadsheet_upload(filename_in, ext_in, raw)

            # 3) local baseline (sample) for merge/fallback
            yield _sse({"type": "progress", "pct": 18, "label": "Local baseline parse"})
            with tempfile.NamedTemporaryFile(delete=False, suffix=("." + ext)) as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            local_rows = _parse_locally_from_file(tmp_path, "." + ext)
            sample_rows = local_rows[:limit_rows] if isinstance(local_rows, list) else []
            local_lines = _rows_to_lines(sample_rows)
            csv_text = _rows_to_csv_text(sample_rows, limit=limit_rows)

            # 4) Ask Gemini with CSV BYTES via inline_data (no 'mime_type' on a text part)
            headers_seen: list[str] = []
            ai_notes = ""
            ai_model = ""
            ai_lines: list[dict] = []

            try:
                yield _sse({"type": "progress", "pct": 62, "label": "Asking Gemini (CSV inline bytes)"})
                model, ai_model, _ = make_gemini_model()
                prompt = make_ai_parser_prompt(expand_units=False)  # vendor rows; expand on server

                resp = model.generate_content([
                    {"text": prompt + "\n\nA CSV sample is attached. Return JSON only."},
                    {
                        "inline_data": {
                            "mime_type": "text/csv",
                            "data": csv_text.encode("utf-8"),  # <-- BYTES, not {"mime_type","text":...}
                        }
                    },
                ])

                data = parse_gemini_json(resp)
                ai_lines = (data or {}).get("lines") or []
                headers_seen = (data or {}).get("detected_headers") or []
                ai_notes = (data or {}).get("notes") or "Parsed via CSV inline."
            except Exception as e:
                if require_ai:
                    yield _sse({"type": "error", "error_type": type(e).__name__, "message": f"AI required: {e}"})
                    return
                ai_notes = f"AI unavailable or failed; used local preview. ({type(e).__name__})"

            # 5) merge + postprocess
            yield _sse({"type": "progress", "pct": 78, "label": "Merging & post-processing"})
            merged = _merge_ai_with_local(ai_lines, local_lines) if ai_lines else local_lines
            lines = _postprocess_lines(merged, expand_units=bool(expand_units))

            # 6) sidebar: recent POs
            with db_conn() as con, con.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute("""
                    SELECT p.id, p.po_number, COALESCE(p.created_at, NOW()) AS created_at,
                           COALESCE(SUM(pl.qty),0)::int AS total_units,
                           COALESCE(SUM(COALESCE(pl.qty,0)*COALESCE(pl.unit_cost,0)),0)::numeric(12,2) AS estimated_total_cost
                    FROM purchase_orders p
                    LEFT JOIN po_lines pl ON pl.purchase_order_id = p.id
                    WHERE p.vendor_id = %s
                    GROUP BY p.id, p.po_number, p.created_at
                    ORDER BY COALESCE(p.created_at, NOW()) DESC, p.id DESC
                    LIMIT 6
                """, (vendor_id,))
                existing = [dict(r) for r in cur.fetchall()]

            yield _sse({"type": "progress", "pct": 98, "label": "Preparing preview"})

            # 7) complete
            payload = {
                "new_po_lines": lines,
                "existing_pos_summary": existing,
                "headers_seen": headers_seen,
                "ai_notes": (ai_notes + f" | expand_units={bool(expand_units)}").strip(),
                "ai_model": ai_model,
            }
            yield _sse({"type": "complete", "payload": payload})

        except Exception as e:
            yield _sse({"type": "error", "error_type": type(e).__name__, "message": str(e)})
        finally:
            if tmp_path:
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

    headers = {
        "Cache-Control": "no-cache, no-transform",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no",
    }
    return StreamingResponse(run(), media_type="text/event-stream", headers=headers)

@app.post("/imports/{po_id}/explode_group")
def explode_group(po_id: str, body: ExplodeGroupBody):
    if not DATABASE_URL:
        raise HTTPException(500, "DATABASE_URL not set")
    with db() as (con, cur):
        default_prefix = body.prefix or "GEN"
        if body.categoryId:
            cur.execute("SELECT prefix FROM categories WHERE id=%s", (body.categoryId,))
            row = cur.fetchone()
            if row and row["prefix"]:
                default_prefix = row["prefix"]

        if body.categoryId:
            cur.execute("""
                SELECT id, purchase_order_id, msrp, unit_cost, qty, category_guess AS category_id
                FROM po_lines
                WHERE purchase_order_id=%s AND category_guess=%s
            """, (po_id, body.categoryId))
        else:
            cur.execute("""
                SELECT id, purchase_order_id, msrp, unit_cost, qty, category_guess AS category_id
                FROM po_lines
                WHERE purchase_order_id=%s AND category_guess IS NULL
            """, (po_id,))
        lines = cur.fetchall()
        if not lines:
            return {"ok": True, "created": 0}

        created = 0
        try:
            cur.execute("BEGIN")
            for ln in lines:
                prefix = default_prefix
                qty = int(ln["qty"] or 1)
                for _ in range(qty):
                    code = _take_synergy(cur, prefix)
                    cur.execute("""
                        INSERT INTO public.inventory_items
                          (synergy_code, purchase_order_id, po_line_id, category_id, cost_unit, msrp, status)
                        VALUES (%s,%s,%s,%s,%s,%s,'INTAKE')
                    """, (
                        code, ln["purchase_order_id"], ln["id"],
                        ln["category_id"], ln["unit_cost"] or 0, ln["msrp"] or 0
                    ))
                    created += 1
            con.commit()
        except Exception as e:
            con.rollback()
            raise HTTPException(500, str(e)) 
        return {"ok": True, "created": created}

# -----------------------------------------------------------------------------
# Misc
# -----------------------------------------------------------------------------
class PurchaseOrderUpdate(BaseModel):
    vendor_id: str

@app.patch("/purchase_orders/{po_id}")
def purchase_order_patch(po_id: str, body: PurchaseOrderUpdate):
    if not po_id:
        raise HTTPException(400, "Missing Purchase Order ID.")
    updated_count = 0
    with db() as (con, cur):
        try:
            cur.execute(
                "UPDATE purchase_orders SET vendor_id = %s, updated_at = NOW() WHERE id = %s",
                (body.vendor_id, po_id,)
            )
            updated_count = cur.rowcount
            if updated_count == 0:
                 con.rollback()
                 raise HTTPException(404, f"Purchase Order with ID {po_id} not found.")
            con.commit()
        except Exception as e:
            con.rollback()
            raise HTTPException(500, f"Database error during PO update: {e}")
    return {"ok": True, "id": po_id, "updated_count": updated_count}

@app.delete("/purchase_orders/{po_id}")
def purchase_order_delete(po_id: str):
    if not po_id:
        raise HTTPException(400, "Missing Purchase Order ID.")
    deleted_po_count = 0
    with db() as (con, cur):
        try:
            cur.execute("DELETE FROM inventory_items WHERE purchase_order_id = %s", (po_id,))
            deleted_inventory_items = cur.rowcount
            cur.execute("DELETE FROM po_lines WHERE purchase_order_id = %s", (po_id,))
            deleted_po_lines = cur.rowcount
            cur.execute("DELETE FROM purchase_orders WHERE id = %s", (po_id,))
            deleted_po_count = cur.rowcount
            if deleted_po_count == 0:
                 con.rollback()
                 raise HTTPException(404, f"Purchase Order with ID {po_id} not found.")
            con.commit()
        except Exception as e:
            con.rollback()
            raise HTTPException(500, f"Database error during PO deletion: {e}")
    return {
        "ok": True, 
        "id": po_id, 
        "deleted_po_count": deleted_po_count,
        "deleted_po_lines": deleted_po_lines,
        "deleted_inventory_items": deleted_inventory_items
    }
class RowPatch(BaseModel):
    grade: Optional[str] = None
    testedBy: Optional[str] = None
    testedDate: Optional[str] = None
    testerComment: Optional[str] = None
    specs: Optional[dict] = None
    price: Optional[float] = None
    ebayPrice: Optional[float] = None
    categoryId: Optional[str] = None
    postedAt: Optional[str] = None
    postedBy: Optional[str] = None
    ebayItemUrl: Optional[str] = None
    status: Optional[str] = None  

import json
from uuid import UUID
from fastapi import HTTPException

@app.patch("/rows/{synergy_id}")
def rows_patch(synergy_id: str, body: RowPatch):
    if not synergy_id:
        raise HTTPException(400, "Missing synergyId")

    mapping = {
        "grade":         "grade",
        "testedBy":      "tested_by",
        "testedDate":    "tested_date",
        "testerComment": "tester_comment",
        "specs":         "specs",
        "price":         "price",
        "ebayPrice":     "ebay_price",
        "categoryId":    "category_id",
        "postedAt":      "posted_at",
        "postedBy":      "posted_by",
        "ebayItemUrl":   "ebay_item_url",
        "status":        "status",
    }
    data = body.model_dump(exclude_unset=True)

    def to_num(v):
        if v in (None, ""): return None
        try: return float(re.sub(r"[^0-9.\-]", "", str(v)))
        except Exception: return None

    def to_uuid(v):
        if not v: return None
        try: return str(UUID(str(v)))
        except Exception: return None

    fields, vals = [], []
    for k, v in data.items():
        col = mapping.get(k)
        if not col: continue
        if col == "tested_date":
            fields.append(f"{col} = %s::date");        vals.append(to_ymd(v))
        elif col == "posted_at":
            fields.append(f"{col} = %s::timestamptz"); vals.append(v or None)
        elif col in ("price","ebay_price"):
            fields.append(f"{col} = %s");              vals.append(to_num(v))
        elif col == "category_id":
            fields.append(f"{col} = %s::uuid");        vals.append(to_uuid(v))
        elif col == "specs":
            fields.append(f"{col} = %s::jsonb");       vals.append(json.dumps(v or {}))
        else:
            fields.append(f"{col} = %s");              vals.append(v if v != "" else None)

    if not fields:
        return {"ok": True, "updated": 0}

    # âœ… Pick the correct key to match on
    if is_uuid_like(synergy_id):
        where_sql = "WHERE synergy_id = %s::uuid"
        key_vals  = [synergy_id]
    else:
        where_sql = "WHERE synergy_code = %s"
        key_vals  = [synergy_id]

    update_sql = f"UPDATE public.inventory_items SET {', '.join(fields)} {where_sql}"

    with db() as (con, cur):
        try:
            cur.execute(update_sql, vals + key_vals)
            if cur.rowcount == 0:
                con.rollback()
                raise HTTPException(404, f"No row found for '{synergy_id}'")
            con.commit()

            # refetch in the shape the UI expects
            select_sql = f"""
                SELECT
                  i.synergy_code                     AS "synergyId",
                  COALESCE(pl.product_name_raw, '')  AS "productName",
                  i.category_id                      AS "categoryId",
                  i.status                           AS "status",
                  '{{}}'::jsonb                      AS "attrs",
                  COALESCE(i.cost_unit, 0)           AS "purchaseCost",
                  COALESCE(pl.qty, 1)                AS "qty",
                  COALESCE(i.msrp, pl.msrp, 0)       AS "msrp",
                  i.purchase_order_id                AS "poId",
                  i.po_line_id                       AS "lineId",
                  i.grade                            AS "grade",
                  i.tested_by                        AS "testedBy",
                  i.tested_date                      AS "testedDate",
                  i.tester_comment                   AS "testerComment",
                  COALESCE(i.specs, '{{}}'::jsonb)   AS "specs",
                  COALESCE(i.price, 0)               AS "price",
                  COALESCE(i.ebay_price, 0)          AS "ebayPrice",
                  i.posted_at                        AS "postedAt",         
                  i.posted_by::text                  AS "postedBy",        
                  i.ebay_item_url                    AS "ebayItemUrl"      
                FROM public.inventory_items i
                LEFT JOIN po_lines pl ON pl.id = i.po_line_id
                {where_sql}
            """
            cur.execute(select_sql, key_vals)
            row = cur.fetchone()
            if row:
                _broadcast("row.upserted", dict(row))
            return {"ok": True, "updated": 1}
        except HTTPException:
            raise
        except Exception as e:
            con.rollback()
            raise HTTPException(500, f"Database error updating row: {e}")

@app.get("/rows/similar")
def rows_similar(
    productName: str = Query(..., min_length=2),
    limit: int = Query(5, ge=1, le=20),
):
    # Split the incoming name into search tokens (min length 3 to reduce noise)
    terms = [t for t in re.findall(r"[A-Za-z0-9]+", productName) if len(t) >= 3]
    if not terms:
        return []

    # Build a parameterized ILIKE WHERE (no pg_trgm dependency)
    where = " AND ".join([f"pl.product_name_raw ILIKE %s" for _ in terms])
    params = [f"%{t}%" for t in terms]

    sql = f"""
        SELECT DISTINCT ON (pl.product_name_raw)
            pl.product_name_raw              AS "productName",
            COALESCE(i.specs, '{{}}'::jsonb) AS "specs"
        FROM po_lines pl
        LEFT JOIN public.inventory_items i ON i.po_line_id = pl.id
        WHERE pl.product_name_raw <> %s
          AND ({where})
        ORDER BY pl.product_name_raw
        LIMIT %s
    """
    with db() as (con, cur):
        cur.execute(sql, [productName] + params + [limit])
        rows = cur.fetchall() or []
        return rows
@app.get("/rows")
def rows_list(
    q: Optional[str] = Query(None, description="Free text across synergy_code, product_name_raw, specs"),
    grade: Optional[str] = Query(None, description="A|B|C|D|P"),
    category: Optional[str] = Query(None, description="UUID of category OR a category label/prefix"),
    categoryId: Optional[str] = Query(None),  
    status: Optional[str] = Query(None, description="ready|incomplete or raw inventory_items.status"),
    limit: int = Query(200, ge=1, le=1000),
    offset: int = Query(0, ge=0),
):
    if categoryId and not category:
        category = categoryId
    """
    List inventory rows with optional filters. `category` may be:
      - a UUID (matches i.category_id), OR
      - a human-friendly label/prefix (matches categories.label/prefix via ILIKE).
    """
    with db() as (con, cur):
        base_sql = """
            SELECT
              i.synergy_code                     AS "synergyId",
              COALESCE(pl.product_name_raw, '')  AS "productName",
              i.category_id                      AS "categoryId",
              i.status                           AS "status",
              '{}'::jsonb                        AS "attrs",
              COALESCE(i.cost_unit, 0)           AS "purchaseCost",
              COALESCE(pl.qty, 1)                AS "qty",
              COALESCE(i.msrp, pl.msrp, 0)       AS "msrp",
              i.purchase_order_id                AS "poId",
              i.po_line_id                       AS "lineId",
              i.grade                            AS "grade",
              i.tested_by                        AS "testedBy",
              i.tested_date                      AS "testedDate",
              i.tester_comment                   AS "testerComment",
              i.posted_at                        AS "postedAt",
              i.posted_by::text                  AS "postedBy",
              i.ebay_item_url                    AS "ebayItemUrl",
              COALESCE(i.specs, '{}'::jsonb)     AS "specs",
              COALESCE(i.price, 0)               AS "price",
              COALESCE(i.ebay_price, 0)          AS "ebayPrice",
              c.label  AS "categoryLabel",
              c.prefix AS "categoryPrefix"
            FROM public.inventory_items i
            LEFT JOIN po_lines     pl ON pl.id = i.po_line_id
            LEFT JOIN categories    c ON c.id  = i.category_id
        """

        where: list[str] = []
        params: dict[str, object] = {"limit": limit, "offset": offset}

        # status filter (supports special "ready"/"incomplete")
        # Note: keep both tested_date/tested_at to match prior behavior.
        ready_sql = "(i.grade IS NOT NULL AND i.tested_by IS NOT NULL AND (i.tested_date IS NOT NULL OR i.tested_at IS NOT NULL))"
        if status == "ready":
            where.append(ready_sql)
        elif status == "incomplete":
            where.append(f"NOT {ready_sql}")
        elif status:
            where.append("i.status = %(status)s")
            params["status"] = status
        else:
            where.append("i.status IN ('INTAKE','TESTING','READY','HOLD','TESTED')")

        # exact filters
        if grade:
            where.append("i.grade = %(grade)s")
            params["grade"] = grade

        # category filter: accept UUID or human label/prefix
        if category:
            if is_uuid_like(category):
                where.append("i.category_id = %(category_uuid)s::uuid")
                params["category_uuid"] = category
            else:
                # ILIKE over label/prefix; escape wildcards
                cat_like = "%" + str(category).replace("%", r"\%").replace("_", r"\_") + "%"
                where.append("(c.label ILIKE %(cat_like)s ESCAPE '\\' OR c.prefix ILIKE %(cat_like)s ESCAPE '\\')")
                params["cat_like"] = cat_like

        # free-text search across code, product name, and specs
        if q:
            needle = "%" + q.replace("%", r"\%").replace("_", r"\_") + "%"
            params["needle"] = needle
            where.append(
                """
                (
                  i.synergy_code ILIKE %(needle)s ESCAPE '\\'
                  OR pl.product_name_raw ILIKE %(needle)s ESCAPE '\\'
                  OR CAST(i.specs AS text) ILIKE %(needle)s ESCAPE '\\'
                )
                """
            )

        sql = f"""{base_sql}
            WHERE {" AND ".join(where) if where else "TRUE"}
            ORDER BY i.synergy_code
            LIMIT %(limit)s OFFSET %(offset)s
        """

        cur.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]



@app.get("/events")
async def events_list():
    q: asyncio.Queue = asyncio.Queue()
    _subscribers.append(q)

    async def gen():
        try:
            # Initial comment to keep connections alive under proxies
            yield b": ok\n\n"
            while True:
                evt = await q.get()
                yield f"event: {evt['type']}\n".encode("utf-8")
                yield f"data: {json.dumps(evt['data'], default=str)}\n\n"

        finally:
            try:
                _subscribers.remove(q)
            except ValueError:
                pass

    return StreamingResponse(gen(), media_type="text/event-stream")

# -----------------------------------------------------------------------------
# main
# -----------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=PORT, reload=True)