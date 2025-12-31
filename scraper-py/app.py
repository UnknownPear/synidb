#!/usr/bin/env python3
from __future__ import annotations

# ============================
# Stdlib
# ============================
import os
import re
import io
import csv
import json
import time
import base64
import hashlib
import signal
import sys
import socket
import subprocess
from typing import Tuple, Optional, List, Dict, Any

# ============================
# Third-party
# ============================
import requests
from flask import Flask, request, jsonify
from flask_cors import CORS
from bs4 import BeautifulSoup, FeatureNotFound
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from requests.utils import dict_from_cookiejar
from werkzeug.utils import secure_filename

# Optional: load .env files
try:
    from dotenv import load_dotenv
    load_dotenv()                       # ./.env
    load_dotenv(".env.development", override=False)
    load_dotenv(".env.local",        override=False)
except Exception:
    pass

# Optional XLSX support
try:
    from openpyxl import load_workbook  # for .xlsx
except Exception:
    load_workbook = None

# Postgres (Directus DB)
import psycopg2
import psycopg2.extras


# ============================
# Config / ENV
# ============================
DEFAULT_PORT      = int(os.environ.get("PORT", "3333"))  # default to 3333 for dev
CONNECT_TIMEOUT   = float(os.environ.get("CONNECT_TIMEOUT", "5"))
READ_TIMEOUT      = float(os.environ.get("READ_TIMEOUT", "45"))
MAX_RETRIES       = int(os.environ.get("MAX_RETRIES", "3"))

# Directus
DIRECTUS_URL   = os.getenv("DIRECTUS_URL", "http://localhost:8055").rstrip("/")
DIRECTUS_TOKEN = os.getenv("DIRECTUS_TOKEN", "")

# Postgres (same DB your Directus uses)
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    print("[server] WARNING: DATABASE_URL not set; /prefix and /imports will fail", file=sys.stderr)

# eBay OAuth (user refresh + client creds)
EBAY_CLIENT_ID      = os.getenv("EBAY_CLIENT_ID", "")
EBAY_CLIENT_SECRET  = os.getenv("EBAY_CLIENT_SECRET", "")
EBAY_REFRESH_TOKEN  = os.getenv("EBAY_REFRESH_TOKEN", "")
EBAY_MARKETPLACE_ID = os.getenv("VITE_EBAY_MARKETPLACE_ID", "EBAY_US")

# Market avg cache
AVG_CACHE_TTL_MS = int(os.getenv("AVG_CACHE_TTL_MS", str(6 * 60 * 60 * 1000)))  # 6h default

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome Safari"
)


# ============================
# Flask
# ============================
app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})
app.url_map.strict_slashes = False

# Very simple in-memory cookie store (dev only)
COOKIE_STORE: Dict[str, str] = {"local": ""}


# ============================
# HTTP session with retries
# ============================
session = requests.Session()
session.headers.update({
    "User-Agent":      UA,
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Cache-Control":   "no-cache",
    "Pragma":          "no-cache",
})
retry = Retry(
    total=MAX_RETRIES,
    backoff_factor=0.7,
    status_forcelist=(429, 500, 502, 503, 504),
    allowed_methods=frozenset({"GET", "POST"}),
    raise_on_status=False,
)
session.mount("https://", HTTPAdapter(max_retries=retry))
session.mount("http://",  HTTPAdapter(max_retries=retry))

def http_get(
    url: str,
    cookie: str = "",
    timeout: Tuple[float, float] = (CONNECT_TIMEOUT, READ_TIMEOUT),
) -> requests.Response:
    headers = {}
    if cookie:
        headers["Cookie"] = cookie
    return session.get(url, headers=headers, timeout=timeout, allow_redirects=True)


# ============================
# DB helpers
# ============================
def db_conn():
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL not set")
    return psycopg2.connect(DATABASE_URL, sslmode=os.getenv("PGSSLMODE", "prefer"))

def to_num(v):
    if v is None:
        return None
    try:
        return float(re.sub(r"[^0-9.\-]", "", str(v)))
    except Exception:
        return None

def map_header(h: str) -> str:
    s = (h or "").strip().lower()
    if "orig" in s or "msrp" in s or "retail" in s: return "msrp"
    if "stock image" in s or "unit cost" in s or "price paid" in s or s == "cost": return "unit_cost"
    if s == "qty" or "quantity" in s: return "qty"
    if "upc" in s: return "upc"
    if "asin" in s: return "asin"
    if "serial" in s: return "serial"
    if "product" in s or "desc" in s or "name" in s: return "product_name_raw"
    return h

def fmt_synergy(prefix: str, seq: int) -> str:
    return f"{prefix}-{str(int(seq)).zfill(4)}"


# ============================
# eBay token (USER refresh)
# ============================
_access_token: Optional[str] = None
_expires_at_ms: int = 0

def get_ebay_token(force: bool = False) -> Optional[str]:
    global _access_token, _expires_at_ms
    now = int(time.time() * 1000)
    if not force and _access_token and now < _expires_at_ms - 5 * 60_000:
        return _access_token

    client_id     = EBAY_CLIENT_ID
    client_secret = EBAY_CLIENT_SECRET
    refresh_token = EBAY_REFRESH_TOKEN
    if not (client_id and client_secret and refresh_token):
        print("[eBay OAuth] Missing credentials", file=sys.stderr)
        return None

    auth  = base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()
    scope = (
        "https://api.ebay.com/oauth/api_scope/sell.inventory "
        "https://api.ebay.com/oauth/api_scope/sell.account "
        "https://api.ebay.com/oauth/api_scope/buy.marketplace.insights"
    )

    r = requests.post(
        "https://api.ebay.com/identity/v1/oauth2/token",
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Authorization": f"Basic {auth}",
        },
        data={"grant_type": "refresh_token", "refresh_token": refresh_token, "scope": scope},
        timeout=(CONNECT_TIMEOUT, READ_TIMEOUT),
    )
    if r.status_code == 200:
        data = r.json()
        _access_token  = data["access_token"]
        _expires_at_ms = now + int(data.get("expires_in", 3600)) * 1000
        print(f"[eBay OAuth] Refreshed user token (~{int((_expires_at_ms - now)/60000)} min)")
        return _access_token

    print(f"[eBay OAuth] Failed: {r.status_code} {r.text}", file=sys.stderr)
    return None


# ============================
# HTML helpers / parsing
# ============================
SPEC_KEYS = ("brand", "model", "processor", "cpu", "ram", "storage", "screen", "size")

def build_query_from_row(row: dict) -> str:
    toks = []
    name = (row.get("productName") or row.get("product_name") or "").strip()
    if name:
        toks.append(name)

    specs = row.get("specs") or {}
    flat  = {k: row.get(k) for k in SPEC_KEYS if row.get(k)}
    merged = {**specs, **{k: v for k, v in flat.items() if v}}

    cpu = merged.get("processor") or merged.get("cpu")
    if cpu: toks.append(str(cpu))
    ram = merged.get("ram")
    if ram: toks.append(str(ram).replace(" GB", "GB"))
    storage = merged.get("storage")
    if storage: toks.append(str(storage).replace(" GB", "GB"))
    size = merged.get("screen") or merged.get("size")
    if size:
        toks.append(str(size).replace(' "', '').replace('"', "").replace(" inch", "").strip() + '"')

    brand = merged.get("brand")
    model = merged.get("model")
    if brand and (brand not in name): toks.insert(0, str(brand))
    if model and (model not in name): toks.append(str(model))

    return " ".join(t for t in toks if t).strip()

def parse_price(text: str) -> Tuple[str, Optional[float], str]:
    if not text:
        return ("", None, "")
    t = re.sub(r"\s+", " ", text).strip()
    m = re.search(r"(?:(US|C|CA)\s*)?([$£€])\s?(\d[\d,]*\.?\d*)", t, re.I)
    if not m:
        return ("", None, t)
    region = (m.group(1) or "").upper()
    sym, val = m.group(2), m.group(3)
    try:
        num = float(val.replace(",", ""))
    except ValueError:
        num = None
    if sym == "$":
        cur = "CA" if region in ("C", "CA") else "US"
    elif sym == "£":
        cur = "GB"
    elif sym == "€":
        cur = "EU"
    else:
        cur = ""
    return (cur, num, (region + " " if region else "") + sym + val)

def soup_from_html(html: str) -> BeautifulSoup:
    try:
        return BeautifulSoup(html or "", "lxml")
    except FeatureNotFound:
        return BeautifulSoup(html or "", "html.parser")

def shape_rows_from_items(items, hits: int) -> List[Dict[str, Any]]:
    rows = []
    for li in items[:hits]:
        a = li.select_one("a.s-item__link")
        title_el = li.select_one(".s-item__title") or li.select_one("h3")
        price_el = li.select_one(".s-item__price")
        if not a or not a.get("href") or not title_el:
            continue
        title = re.sub(r"\s+", " ", title_el.get_text(strip=True))
        price_text = re.sub(r"\s+", " ", price_el.get_text(strip=True)) if price_el else ""
        cur, num, pretty = parse_price(price_text)
        rows.append({
            "url": a["href"],
            "title": title,
            "priceText": pretty or price_text,
            "currency": cur,
            "priceNum": num,
        })
    return rows

def summarize_rows(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    usd = [r for r in rows if r["currency"] == "US" and isinstance(r["priceNum"], (int, float))]
    valid = len(usd)
    avg = round(sum(r["priceNum"] for r in usd) / valid, 2) if valid else 0.0
    return {"sampled": len(rows), "valid": valid, "avg": avg}


# ============================
# Playwright fallback (optional)
# ============================
def scrape_via_playwright(url: str, hits: int = 20) -> Dict[str, Any]:
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except Exception as e:
        return {"ok": False, "error": f"playwright_not_available: {e}"}

    def fail(msg: str, page=None):
        try:
            if page:
                page.screenshot(path="/tmp/ebay_last.png", full_page=True)
                with open("/tmp/ebay_last.html", "w", encoding="utf-8") as f:
                    f.write(page.content())
        except Exception:
            pass
        return {"ok": False, "error": msg}

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--disable-features=IsolateOrigins,site-per-process",
                ],
            )
            ua = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36")
            storage_state_path = "storage_state.json"
            ctx = browser.new_context(
                storage_state=storage_state_path if os.path.exists(storage_state_path) else None,
                user_agent=ua,
                locale="en-US",
                timezone_id="America/Chicago",
                viewport={"width": 1400, "height": 900},
            )
            page = ctx.new_page()
            page.goto(url, wait_until="domcontentloaded")

            if "splashui/challenge" in page.url:
                page.wait_for_timeout(4000)
                page.goto(url, wait_until="domcontentloaded")

            for sel in [
                'button:has-text("Accept all")',
                'button:has-text("Accept")',
                'button:has-text("Got it")',
                'button:has-text("OK")',
                '[aria-label="Close"]',
            ]:
                try:
                    page.locator(sel).first.click(timeout=1200)
                except Exception:
                    pass

            try:
                page.wait_for_load_state("networkidle", timeout=12000)
            except PWTimeout:
                pass
            for _ in range(6):
                page.mouse.wheel(0, 1200)
                page.wait_for_timeout(500)

            selectors = [
                "ul.srp-results li.s-item",
                "li.s-item",
                "[data-testid='item-cell']",
                "[data-view*='srp'] li.s-item",
            ]
            cards = []
            for sel in selectors:
                try:
                    page.wait_for_selector(sel, timeout=5000)
                    cards = page.query_selector_all(sel)
                    if cards:
                        break
                except PWTimeout:
                    continue

            rows: List[Dict[str, Any]] = []
            def extract(li):
                a = li.query_selector("a.s-item__link") or li.query_selector("a[href*='/itm/']")
                title_el = li.query_selector(".s-item__title") or li.query_selector("h3, .s-item__subtitle")
                price_el = li.query_selector(".s-item__price, .s-item__detail--primary span")
                if not a or not title_el:
                    return None
                title = (title_el.inner_text() or "").strip()
                href = a.get_attribute("href") or ""
                price_text = (price_el.inner_text() or "").strip() if price_el else ""
                m = re.search(r"([$£€])\s?(\d[\d,]*\.?\d*)", price_text or "")
                price_num = float(m.group(2).replace(",", "")) if m else None
                return {"title": title, "url": href, "priceText": price_text, "priceNum": price_num,
                        "currency": "US" if price_num is not None else ""}

            for li in cards[:hits]:
                row = extract(li)
                if row:
                    rows.append(row)

            if not rows:
                html = page.content()
                soup = soup_from_html(html)
                items = soup.select("ul.srp-results li.s-item") or soup.select("li.s-item") or soup.select("[data-testid='item-cell']")
                if items:
                    rows = shape_rows_from_items(items, hits)

            browser.close()

        if not rows:
            return fail("challenge_or_no_results_after_browser (see /tmp/ebay_last.html & /tmp/ebay_last.png)")
        summary = summarize_rows(rows)
        return {"ok": True, **summary, "rows": rows}
    except Exception as e:
        return {"ok": False, "error": f"playwright_error: {e}"}


# ============================
# Directus cache helpers
# ============================
def _avg_key(q, fixed_price, condition, currency, limit, min_price, max_price, category_ids):
    payload = {
        "q": (q or "").strip().lower(),
        "fixed": bool(fixed_price),
        "condition": (condition or "").upper(),
        "currency": (currency or "").upper(),
        "limit": int(limit or 100),
        "min_price": min_price,
        "max_price": max_price,
        "category_ids": category_ids or "",
    }
    s = json.dumps(payload, sort_keys=True)
    return hashlib.sha1(s.encode()).hexdigest()

def _directus_get(key: str):
    if not DIRECTUS_TOKEN:
        return None
    r = session.get(
        f"{DIRECTUS_URL}/items/market_avg",
        headers={"Authorization": f"Bearer {DIRECTUS_TOKEN}"},
        params={"filter[key][_eq]": key, "limit": 1},
        timeout=(CONNECT_TIMEOUT, READ_TIMEOUT),
    )
    if r.status_code != 200:
        return None
    data = r.json().get("data") or []
    return data[0] if data else None

def _directus_save(record: dict):
    if not DIRECTUS_TOKEN:
        return
    session.post(
        f"{DIRECTUS_URL}/items/market_avg",
        headers={
            "Authorization": f"Bearer {DIRECTUS_TOKEN}",
            "Content-Type": "application/json",
        },
        json=record,
        timeout=(CONNECT_TIMEOUT, READ_TIMEOUT),
    )


# ============================
# Routes
# ============================
@app.get("/health")
def health():
    return jsonify(ok=True, ts=time.time())

@app.post("/cookie")
def save_cookie():
    data = request.get_json(silent=True) or {}
    cookie = str(data.get("cookie", "")).strip()
    if not cookie:
        return jsonify(ok=False, error="cookie required"), 400
    COOKIE_STORE["local"] = cookie
    return jsonify(ok=True)

@app.get("/warmup")
def warmup():
    url = request.args.get("url", "")
    if not url:
        return jsonify(ok=False, error="url required"), 400
    try:
        r = http_get(url)
        jar = dict_from_cookiejar(session.cookies)
        if jar:
            COOKIE_STORE["local"] = "; ".join(f"{k}={v}" for k, v in jar.items())
        return jsonify(ok=True, status=r.status_code, cookies=len(jar))
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 502

# eBay proxy (passes Authorization & marketplace header)
@app.route("/ebay/<path:subpath>", methods=["GET", "POST", "PUT", "DELETE"])
def ebay_proxy(subpath):
    token = get_ebay_token()
    if not token:
        return jsonify(ok=False, error="no_token"), 401

    url = f"https://api.ebay.com/{subpath}"
    headers = {
        "Authorization": f"Bearer {token}",
        "X-EBAY-C-MARKETPLACE-ID": EBAY_MARKETPLACE_ID,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    method = request.method.upper()
    if method == "GET":
        resp = requests.get(url, headers=headers, params=request.args)
    elif method == "POST":
        resp = requests.post(url, headers=headers, json=request.get_json(silent=True))
    elif method == "PUT":
        resp = requests.put(url, headers=headers, json=request.get_json(silent=True))
    else:
        resp = requests.delete(url, headers=headers, json=request.get_json(silent=True))
    return resp.content, resp.status_code, {"Content-Type": resp.headers.get("Content-Type", "application/json")}

# Basic HTML scrape (active)
@app.get("/scrape")
def scrape():
    started = time.time()
    url = request.args.get("url", "")
    hits = max(1, min(200, int(request.args.get("hits", "15") or 15)))
    if not url:
        return jsonify(ok=False, error="url required"), 400

    cookie = (request.headers.get("x-ebay-cookie") or "").strip() or COOKIE_STORE.get("local", "")

    try:
        r = http_get(url, cookie=cookie)
        soup = soup_from_html(r.text)
        items = soup.select(".srp-results .s-item")
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 502

    should_retry = (not items) or (r is not None and r.status_code >= 500)

    if should_retry:
        try:
            warm = http_get("https://www.ebay.com/")
            warm_jar = dict_from_cookiejar(session.cookies)
            if warm_jar:
                COOKIE_STORE["local"] = "; ".join(f"{k}={v}" for k, v in warm_jar.items())
                cookie = COOKIE_STORE["local"]
        except Exception:
            pass

        time.sleep(0.8)
        try:
            r = http_get(url, cookie=cookie)
            soup = soup_from_html(r.text)
            items = soup.select(".srp-results .s-item")
        except Exception:
            items = None

        if not items:
            pw = scrape_via_playwright(url, hits)
            if pw.get("ok"):
                payload = {
                    "ok": True,
                    "endpoint": request.url,
                    "status": r.status_code if isinstance(r, requests.Response) else None,
                    "durationMs": int((time.time() - started) * 1000),
                    "sampled": pw["sampled"],
                    "valid": pw["valid"],
                    "avg": pw["avg"],
                    "rows": pw["rows"],
                    "via": "playwright",
                }
                return jsonify(payload), 200
            else:
                title_snippet = (soup.title.string if soup and soup.title else "")[:160] if isinstance(soup, BeautifulSoup) else ""
                return jsonify(ok=False, error=pw.get("error", "no_results_from_index"), snippet=title_snippet, status=(r.status_code if isinstance(r, requests.Response) else None)), 502

    rows = shape_rows_from_items(items, hits)
    summary = summarize_rows(rows)
    return jsonify(
        ok=True,
        endpoint=request.url,
        status=r.status_code,
        durationMs=int((time.time() - started) * 1000),
        sampled=summary["sampled"],
        valid=summary["valid"],
        avg=summary["avg"],
        rows=rows,
        via="requests",
    )

# Find an already-posted row (mock for now)
@app.route("/rows/search_posted", methods=["POST"])
def search_posted_row():
    try:
        criteria = request.get_json(silent=True) or {}
        product_name = (criteria.get("productName") or "").strip().lower()
        grade = (criteria.get("grade") or "").strip().upper()

        # MOCK for now (replace with real DB search)
        if product_name and "laptop" in product_name:
            return jsonify({"ebayItemUrl": "https://www.ebay.com/itm/MOCK_COMP_LINK"})
        return jsonify({}), 200
    except Exception as e:
        print(f"Error in /rows/search_posted: {e}", file=sys.stderr)
        return jsonify({"error": "Failed to search for existing listing"}), 500

# Completed+Sold scrape average (HTML + PW fallback)
@app.get("/sold_avg")
def sold_avg():
    q = request.args.get("q", "").strip()
    row_raw = request.args.get("row")
    if not q and row_raw:
        try:
            q = build_query_from_row(json.loads(row_raw))
        except Exception:
            pass
    if not q:
        return jsonify(ok=False, error="q or row required"), 400

    limit = max(1, min(200, int(request.args.get("limit", "60"))))
    hits = limit

    site = "https://www.ebay.com/sch/i.html"
    url = f"{site}?_nkw={requests.utils.quote(q)}&LH_Complete=1&LH_Sold=1&_sop=13"

    started = time.time()
    cookie = COOKIE_STORE.get("local", "")

    try:
        r = http_get(url, cookie=cookie)
        soup = soup_from_html(r.text)
        items = soup.select(".srp-results .s-item") or soup.select("ul.srp-results li.s-item") or soup.select("li.s-item") or soup.select("[data-testid='item-cell']")
    except Exception:
        items = None
        r = None

    should_retry = (not items) or (r is not None and r.status_code >= 500)

    if should_retry:
        try:
            warm = http_get("https://www.ebay.com/")
            warm_jar = dict_from_cookiejar(session.cookies)
            if warm_jar:
                COOKIE_STORE["local"] = "; ".join(f"{k}={v}" for k, v in warm_jar.items())
                cookie = COOKIE_STORE["local"]
        except Exception:
            pass

        time.sleep(0.8)
        try:
            r = http_get(url, cookie=cookie)
            soup = soup_from_html(r.text)
            items = soup.select(".srp-results .s-item") or soup.select("ul.srp-results li.s-item") or soup.select("li.s-item") or soup.select("[data-testid='item-cell']")
        except Exception:
            items = None

        if not items:
            pw = scrape_via_playwright(url, hits)
            if pw.get("ok"):
                payload = {
                    "ok": True,
                    "endpoint": request.url,
                    "status": r.status_code if isinstance(r, requests.Response) else None,
                    "durationMs": int((time.time() - started) * 1000),
                    "sampled": pw["sampled"],
                    "valid": pw["valid"],
                    "avg": pw["avg"],
                    "rows": pw["rows"],
                    "via": "playwright",
                }
                return jsonify(payload), 200
            else:
                title_snippet = (soup.title.string if soup and hasattr(soup, "title") and soup.title else "")[:160]
                return jsonify(ok=False, error=pw.get("error", "no_results_from_index"), snippet=title_snippet, status=(r.status_code if isinstance(r, requests.Response) else None)), 502

    rows = shape_rows_from_items(items, hits)
    summary = summarize_rows(rows)
    return jsonify(
        ok=True,
        endpoint=request.url,
        status=r.status_code if isinstance(r, requests.Response) else None,
        durationMs=int((time.time() - started) * 1000),
        sampled=summary["sampled"],
        valid=summary["valid"],
        avg=summary["avg"],
        rows=rows[:10],
        via="requests",
    )

# Active Browse API average (with Directus cache)
@app.get("/browse_avg")
def browse_avg():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify(ok=False, error="q required"), 400

    limit       = int(request.args.get("limit", "100") or 100)
    filters_raw = request.args.get("filter", "") or ""
    currency    = (request.args.get("priceCurrency", "USD") or "USD").upper()

    fixed_price = "buyingOptions:{FIXED_PRICE}" in filters_raw if filters_raw else True
    condition   = "USED"
    if "conditions:{" in filters_raw:
        try:
            condition = filters_raw.split("conditions:{", 1)[1].split("}", 1)[0].split("|")[0].strip().upper()
        except Exception:
            pass

    key = _avg_key(q, fixed_price, condition, currency, limit, None, None, None)

    now_ms = int(time.time() * 1000)
    cached = _directus_get(key)
    if cached:
        age = now_ms - int(cached.get("created_at_ms", 0))
        if 0 <= age < int(cached.get("ttl_ms", AVG_CACHE_TTL_MS)):
            return jsonify({
                "ok": True, "source": "cache",
                "currency": cached.get("currency"),
                "sampled": cached.get("sampled", 0),
                "valid": cached.get("valid", 0),
                "avg": cached.get("avg", 0.0),
                "rows": cached.get("rows") or [],
            })

    if not EBAY_CLIENT_ID or not EBAY_CLIENT_SECRET:
        return jsonify(ok=False, error="EBAY_CLIENT_ID/SECRET env vars required"), 500

    auth = base64.b64encode(f"{EBAY_CLIENT_ID}:{EBAY_CLIENT_SECRET}".encode()).decode()
    tok = session.post(
        "https://api.ebay.com/identity/v1/oauth2/token",
        headers={"Content-Type": "application/x-www-form-urlencoded",
                 "Authorization": f"Basic {auth}"},
        data="grant_type=client_credentials&scope=https://api.ebay.com/oauth/api_scope",
        timeout=(CONNECT_TIMEOUT, READ_TIMEOUT)
    )
    if tok.status_code != 200:
        return jsonify(ok=False, error="oauth_failed", status=tok.status_code, body=tok.text), 502
    access = tok.json().get("access_token")

    params = {"q": q, "limit": min(200, max(1, limit)), "priceCurrency": currency}
    if filters_raw:
        params["filter"] = filters_raw
    else:
        params["filter"] = "buyingOptions:{FIXED_PRICE},conditions:{USED}"

    r = session.get(
        "https://api.ebay.com/buy/browse/v1/item_summary/search",
        params=params,
        headers={"Authorization": f"Bearer {access}"},
        timeout=(CONNECT_TIMEOUT, READ_TIMEOUT),
    )
    if r.status_code != 200:
        return jsonify(ok=False, error="browse_failed", status=r.status_code, body=r.text), 502

    data = r.json()
    rows: List[Dict[str, Any]] = []
    for it in data.get("itemSummaries", []):
        p = it.get("price") or {}
        title = it.get("title")
        url   = it.get("itemWebUrl")
        price_num = None
        cur = p.get("currency", "")
        if "value" in p:
            try: price_num = float(p["value"])
            except Exception: price_num = None
        rows.append({
            "title": title,
            "url": url,
            "priceText": f"{cur} {p.get('value')}",
            "priceNum": price_num,
            "currency": cur
        })

    from collections import Counter
    currs    = [r["currency"] for r in rows if r.get("currency")]
    dominant = "USD" if "USD" in currs else (Counter(currs).most_common(1)[0][0] if currs else "")
    prices   = [r["priceNum"] for r in rows if isinstance(r.get("priceNum"), (int, float)) and r["currency"] == dominant]
    avg      = round(sum(prices) / len(prices), 2) if prices else 0.0

    payload = {
        "ok": True, "source": "browse_api", "currency": dominant,
        "sampled": len(rows), "valid": len(prices), "avg": avg, "rows": rows[:10],
    }

    rec = {
        "key": key, "q": q, "filters": params.get("filter", ""), "currency": dominant,
        "sampled": len(rows), "valid": len(prices), "avg": avg, "rows": rows[:10],
        "created_at_ms": now_ms, "ttl_ms": AVG_CACHE_TTL_MS,
    }
    try:
        _directus_save(rec)
    except Exception:
        pass

    return jsonify(payload), 200

# Aliases with /api prefix (optional)
@app.get("/api/browse_avg")
def browse_avg_alias():
    return browse_avg()

@app.get("/api/sold_avg")
def sold_avg_alias():
    return sold_avg()


# ============================
# Synergy ID prefix counters
# ============================
@app.get("/prefix/<prefix>/peek")
def prefix_peek(prefix):
    if not DATABASE_URL:
        return jsonify(error="DATABASE_URL not set"), 500
    with db_conn() as con, con.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        cur.execute("SELECT next_seq FROM id_prefix_counters WHERE prefix=%s", (prefix,))
        row = cur.fetchone()
        next_seq = row["next_seq"] if row else 1
        return (fmt_synergy(prefix, next_seq), 200, {"Content-Type": "text/plain"})

@app.post("/prefix/<prefix>/take")
def prefix_take(prefix):
    if not DATABASE_URL:
        return jsonify(error="DATABASE_URL not set"), 500
    with db_conn() as con:
        with con.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("BEGIN")
            cur.execute("SELECT next_seq FROM id_prefix_counters WHERE prefix=%s FOR UPDATE", (prefix,))
            row = cur.fetchone()
            if row:
                next_seq = row["next_seq"]
                cur.execute("UPDATE id_prefix_counters SET next_seq = next_seq + 1 WHERE prefix=%s", (prefix,))
            else:
                next_seq = 1
                cur.execute("INSERT INTO id_prefix_counters(prefix, next_seq) VALUES (%s, %s)", (prefix, 2))
            con.commit()
            return (fmt_synergy(prefix, next_seq), 200, {"Content-Type": "text/plain"})

@app.post("/prefix/<prefix>/set")
def prefix_set(prefix):
    if not DATABASE_URL:
        return jsonify(error="DATABASE_URL not set"), 500
    data = request.get_json(silent=True) or {}
    nxt = int(data.get("next", 1))
    if nxt < 1:
        return jsonify(error="next must be >=1"), 400
    with db_conn() as con, con.cursor() as cur:
        cur.execute("""
            INSERT INTO id_prefix_counters(prefix, next_seq) VALUES (%s,%s)
            ON CONFLICT (prefix) DO UPDATE SET next_seq = EXCLUDED.next_seq
        """, (prefix, nxt))
        con.commit()
        return jsonify(ok=True, next=nxt)


# ============================
# Manager upload (CSV/XLSX → po_lines)
# ============================
@app.post("/imports/upload")
def import_upload():
    if not DATABASE_URL:
        return jsonify(error="DATABASE_URL not set"), 500

    f           = request.files.get("file")
    po_number   = (request.form.get("po_number")  or "").strip()
    vendor_name = (request.form.get("vendor_name") or "").strip()
    category_id = (request.form.get("category_id") or "").strip() or None

    if not f or not po_number:
        return jsonify(error="Missing 'file' or 'po_number'"), 400

    filename = secure_filename(f.filename or "upload")
    ext = (filename.split(".")[-1] or "").lower()

    rows: List[Dict[str, Any]] = []
    try:
        if ext in ("csv", "tsv", "txt"):
            stream = io.StringIO(f.stream.read().decode("utf-8", errors="ignore"))
            try:
                dialect = csv.Sniffer().sniff(stream.read(4096), delimiters=",\t;|")
                stream.seek(0)
                reader = csv.DictReader(stream, dialect=dialect)
            except Exception:
                stream.seek(0)
                reader = csv.DictReader(stream)
            for raw in reader:
                r = {}
                for k, v in raw.items():
                    r[map_header(k)] = v
                rows.append(r)

        elif ext in ("xlsx", "xlsm", "xltx", "xltm"):
            if load_workbook is None:
                return jsonify(error="openpyxl not installed"), 500
            data = io.BytesIO(f.stream.read())
            wb = load_workbook(filename=data, read_only=True, data_only=True)
            ws = wb.active
            hdr = [map_header(str(c.value) if c else "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for rr in ws.iter_rows(min_row=2):
                obj = {}
                for idx, c in enumerate(rr):
                    key = hdr[idx] if idx < len(hdr) else f"col_{idx}"
                    obj[key] = c.value
                rows.append(obj)
        else:
            return jsonify(error="Unsupported file type. Use CSV or XLSX."), 400
    except Exception as e:
        return jsonify(error=f"parse failed: {e}"), 400

    with db_conn() as con:
        cur = con.cursor(cursor_factory=psycopg2.extras.DictCursor)
        try:
            cur.execute("BEGIN")

            # vendor
            vendor_id = None
            if vendor_name:
                cur.execute("SELECT id FROM vendors WHERE name=%s LIMIT 1", (vendor_name,))
                v = cur.fetchone()
                if v:
                    vendor_id = v["id"]
                else:
                    cur.execute("INSERT INTO vendors(name) VALUES (%s) RETURNING id", (vendor_name,))
                    vendor_id = cur.fetchone()["id"]

            # purchase order
            cur.execute("SELECT id FROM purchase_orders WHERE po_number=%s LIMIT 1", (po_number,))
            p = cur.fetchone()
            if p:
                po_id = p["id"]
            else:
                cur.execute(
                    "INSERT INTO purchase_orders(po_number, vendor_id, status) VALUES (%s,%s,'Here') RETURNING id",
                    (po_number, vendor_id)
                )
                po_id = cur.fetchone()["id"]

            # optional category validation
            if category_id:
                cur.execute("SELECT 1 FROM categories WHERE id=%s", (category_id,))
                if not cur.fetchone():
                    return jsonify(error="category_id not found"), 400

            created = 0
            for r in rows:
                if not any(list(filter(None, r.values()))):
                    continue

                line = {
                    "purchase_order_id": po_id,
                    "product_name_raw": r.get("product_name_raw"),
                    "upc": r.get("upc"),
                    "asin": r.get("asin"),
                    "msrp": to_num(r.get("msrp")),
                    "unit_cost": to_num(r.get("unit_cost")),
                    "qty": int(re.sub(r"[^0-9]", "", str(r.get("qty") or 1)) or 1),
                    "category_guess": category_id,
                    "raw_json": json.dumps(r, default=str)
                }

                cols = list(line.keys())
                cur.execute(
                    f"INSERT INTO po_lines ({','.join(cols)}) VALUES ({','.join(['%s']*len(cols))})",
                    [line[c] for c in cols]
                )
                created += 1

            con.commit()
        except Exception as e:
            con.rollback()
            return jsonify(error=str(e)), 500

    return jsonify(ok=True, po_id=str(po_id), vendor_id=str(vendor_id) if vendor_id else None,
                   category_id=category_id, created_lines=created)

# ============================
# Explode PO → inventory_items
# ============================
@app.post("/imports/<po_id>/explode")
def explode_po(po_id):
    """
    Body: { "categoryId": "<uuid>", "useLineCategory": true/false }
    For each po_lines row under the PO, create `qty` items with synergy_code.
    """
    if not DATABASE_URL:
        return jsonify(error="DATABASE_URL not set"), 500

    body = request.get_json(silent=True) or {}
    category_id = body.get("categoryId")
    use_line    = bool(body.get("useLineCategory"))

    with db_conn() as con:
        cur = con.cursor(cursor_factory=psycopg2.extras.DictCursor)

        default_prefix = None
        if category_id:
            cur.execute("SELECT prefix FROM categories WHERE id=%s", (category_id,))
            row = cur.fetchone()
            if not row:
                return jsonify(error="categoryId not found"), 400
            default_prefix = row["prefix"]

        cur.execute("""
            SELECT pl.id, pl.purchase_order_id, pl.msrp, pl.unit_cost, pl.qty,
                   CASE WHEN %s THEN (SELECT c.prefix FROM categories c WHERE c.id = pl.category_guess) ELSE NULL END AS line_prefix,
                   CASE WHEN %s THEN pl.category_guess ELSE %s::uuid END AS category_id
            FROM po_lines pl
            WHERE pl.purchase_order_id = %s
        """, (use_line, use_line, category_id, po_id))
        lines = cur.fetchall()

        if not lines:
            return jsonify(ok=True, created=0)

        def take(prefix: str) -> str:
            cur.execute("SELECT next_seq FROM id_prefix_counters WHERE prefix=%s FOR UPDATE", (prefix,))
            row = cur.fetchone()
            if row:
                next_seq = row["next_seq"]
                cur.execute("UPDATE id_prefix_counters SET next_seq = next_seq + 1 WHERE prefix=%s", (prefix,))
            else:
                next_seq = 1
                cur.execute("INSERT INTO id_prefix_counters(prefix, next_seq) VALUES (%s,%s)", (prefix, 2))
            return fmt_synergy(prefix, next_seq)

        created = 0
        try:
            cur.execute("BEGIN")
            for ln in lines:
                prefix = ln["line_prefix"] or default_prefix
                if not prefix:
                    con.rollback()
                    return jsonify(error="No prefix for one or more lines; pass categoryId or set per-line category"), 400
                qty = int(ln["qty"] or 1)
                for _ in range(qty):
                    code = take(prefix)
                    cur.execute("""
                        INSERT INTO inventory_items
                          (synergy_code, purchase_order_id, po_line_id, category_id, cost_unit, msrp, status)
                        VALUES (%s,%s,%s,%s,%s,%s,'INTAKE')
                    """, (
                        code, ln["purchase_order_id"], ln["id"], ln["category_id"],
                        ln["unit_cost"] or 0, ln["msrp"] or 0
                    ))
                    created += 1
            con.commit()
        except Exception as e:
            con.rollback()
            return jsonify(error=str(e)), 500

        return jsonify(ok=True, created=created)


# ============================
# Port auto-kill (best effort)
# ============================
def _pids_listening_on_port(port: int) -> List[int]:
    if os.name == "nt":
        try:
            out = subprocess.check_output(["netstat", "-ano"], stderr=subprocess.DEVNULL, text=True)
            pids = set()
            needle = f":{port} "
            for line in out.splitlines():
                if "LISTENING" in line and needle in line:
                    parts = line.split()
                    if parts:
                        pids.add(int(parts[-1]))
            return list(pids)
        except Exception:
            return []
    # POSIX: prefer lsof; fast & reliable with -nP
    try:
        out = subprocess.check_output(
            ["lsof", "-nP", f"-iTCP:{port}", "-sTCP:LISTEN", "-t"],
            stderr=subprocess.DEVNULL, text=True
        )
        return [int(x) for x in out.split()]
    except Exception:
        pass
    try:
        out = subprocess.check_output(["fuser", f"{port}/tcp"], stderr=subprocess.DEVNULL, text=True)
        return [int(tok) for tok in out.replace("\n", " ").split() if tok.isdigit()]
    except Exception:
        return []

def free_port(port: int, aggressive: bool = False) -> None:
    my_pid = os.getpid()
    pids = _pids_listening_on_port(port)
    if not pids:
        return

    same_user_pids: List[int] = []
    for pid in pids:
        if pid == my_pid:
            continue
        if hasattr(os, "getuid"):
            try:
                st = os.stat(f"/proc/{pid}")  # Linux; macOS won't have /proc, skip uid check
                if st.st_uid != os.getuid():
                    continue
            except Exception:
                pass
        same_user_pids.append(pid)

    if not same_user_pids:
        return

    print(f"[startup] Port {port} is busy; attempting to stop PIDs: {same_user_pids}", file=sys.stderr)
    for pid in same_user_pids:
        try:
            if os.name == "nt":
                subprocess.call(["taskkill", "/PID", str(pid), "/T", "/F"],
                                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:
                os.kill(pid, signal.SIGTERM)
        except Exception:
            pass

    time.sleep(0.5)
    remaining = [pid for pid in _pids_listening_on_port(port) if pid in same_user_pids]
    if remaining and aggressive and os.name != "nt":
        for pid in remaining:
            try:
                os.kill(pid, signal.SIGKILL)
            except Exception:
                pass
        time.sleep(0.2)

    still = [pid for pid in _pids_listening_on_port(port) if pid in same_user_pids]
    if still:
        print(f"[startup] Warning: port {port} still in use by {still}", file=sys.stderr)

def _port_is_free(port: int) -> bool:
    s = socket.socket()
    try:
        s.settimeout(0.2)
        s.bind(("0.0.0.0", port))
        return True
    except OSError:
        return False
    finally:
        try:
            s.close()
        except Exception:
            pass


# ============================
# Main
# ============================
if __name__ == "__main__":
    # optional: prefetch eBay token
    try:
        get_ebay_token(True)
    except Exception:
        pass

    port = DEFAULT_PORT
    # be aggressive and then wait a moment for the OS to release the socket
    free_port(port, aggressive=True)

    # wait up to ~2s for the port to become free
    start = time.time()
    while not _port_is_free(port) and (time.time() - start) < 2.0:
        time.sleep(0.1)

    app.run(host="0.0.0.0", port=port, use_reloader=False)
