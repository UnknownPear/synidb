import os
import io
import re
import csv
import json
import time
import html
import requests
from typing import Optional, List, Dict, Any, Tuple
from fastapi.responses import JSONResponse
from fastapi import HTTPException
from pydantic import BaseModel
from requests.utils import dict_from_cookiejar
# Import configs
from config import guess_mime
from db_utils import to_num, map_header

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


AI_RESPONSE_SCHEMA = {
    "type": "object",
    "properties": {
        "detected_headers": {"type": "array", "items": {"type": "string"}},
        "lines": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "product_name_raw": {"type": "string"},
                    "qty": {"type": "integer", "minimum": 1},
                    "unit_cost": {"type": ["number", "null"]},
                    "msrp": {"type": ["number", "null"]},
                    "upc": {"type": ["string", "null"]},
                    "asin": {"type": ["string", "null"]},
                    "category_guess": {"type": ["string", "null"]},
                    # --- ADDED: Specs and Notes fields ---
                    "specs": {
                        "type": "object",
                        "properties": {
                            "processor": {"type": ["string", "null"]},
                            "ram": {"type": ["string", "null"]},
                            "storage": {"type": ["string", "null"]},
                            "screen": {"type": ["string", "null"]},
                            "color": {"type": ["string", "null"]},
                            "batteryHealth": {"type": ["string", "null"]},
                        }
                    },
                    "item_notes": {"type": ["string", "null"]}
                    # -------------------------------------
                },
                "required": ["product_name_raw", "qty"],
            },
        },
        "notes": {"type": "string"},
    },
    "required": ["lines"],
    "additionalProperties": False,
}

AI_PARSER_PROMPT = """You are a purchasing-sheet parser. Input is a spreadsheet or PDF
containing a purchase order (PO) or inventory manifest. Output is a SINGLE
JSON object matching the schema below—no prose, no code fences.

You must respect these MODE FLAGS (set by the caller at the end of this prompt):
- EXPAND_UNITS = true|false
  • true  → If a row indicates Quantity = N (>1), expand that row into N separate line objects.
            Each expanded line must have qty = 1 and the same per-unit unit_cost and other fields.
            If unit_cost is not given but a total/extended amount exists, compute unit_cost = round(total/N, 2).
  • false → Do not expand. Keep each vendor row as one output line with qty = N.

Core mapping rules:
- product_name_raw: The best available item description/title.
- qty: integer ≥ 1 (see EXPAND_UNITS).
- unit_cost:
    * Prefer the per-unit price column if present.
    * If the sheet has a column named “STOCK IMAGE” (any case), that value is the ACTUAL price paid per unit → map it to unit_cost.
    * If only totals exist, compute unit_cost = total/qty (2 decimals).
    * Never use totals/extended amounts directly as unit_cost without normalizing to per-unit.
- msrp: Manufacturer suggested retail/original/list price. Synonyms include “Orig. Retail”, “Original Retail”, “List”, “MSRP”.
- upc / asin: Extract if present; also scan description text for clear UPC/ASIN tokens.
- category_guess:
    * If a “Category” column is present, use it (trimmed).
    * Otherwise infer from the product name/specs. Keep this short and human-readable
      (e.g., “Laptop”, “Desktop”, “Monitor”, “GPU”, “CPU”, “Motherboard”, “Memory”, “Storage”,
       “Keyboard”, “Mouse”, “Headset”, “Console”, “Controller”, “Network”, “Power”, “Accessory”, “Other”).
    * Choose the most specific one you’re confident about; else null.

Header synonyms:
- product_name_raw: “Product”, “Item”, “Listing Title”, “Title”, “Name”, “Description”.
- unit_cost: “Unit Cost”, “Price Paid”, “Cost”, “Our Cost”, “STOCK IMAGE”.
- msrp: “Orig. Retail”, “Original Retail”, “List”, “MSRP”.
- qty: “Qty”, “Quantity”, “QTY”.

Currency & hygiene:
- Parse currency like “$ 1,499.99” → 1499.99 (float). Missing/blank → null.
- Ignore non-item rows: headers, subtotals, grand totals, shipping, tax, footers, ads.
- If multiple header rows, detect the best header row.
- Keep JSON compact. Do NOT include markdown, comments, or extra keys.

SCHEMA (return EXACTLY this shape):
{
  "detected_headers": [string],     // the literal column headers you used
  "lines": [
    {
      "product_name_raw": string,
      "qty": integer,               // = 1 if EXPAND_UNITS=true; else may be >1
      "unit_cost": number|null,
      "msrp": number|null,
      "upc": string|null,
      "asin": string|null,
      "category_guess": string|null
    }
  ],
  "notes": string                   // brief, e.g., "Expanded 3 rows into 70 units; used STOCK IMAGE for unit_cost."
}

MODE FLAGS:
EXPAND_UNITS = {expand_units}
"""

# Kept for compatibility, though we use Pollinations now
def make_gemini_model():
    return None, "pollinations-openai", False

def make_ai_parser_prompt(
    expand_units: bool,
    category_descriptions: str | None = None,
) -> str:
    flag = "true" if expand_units else "false"
    
    # We append specific instructions for specs/notes to the existing prompt
    base = AI_PARSER_PROMPT.replace("{expand_units}", flag) + """

CRITICAL MAPPING RULES FOR SPECS:
You MUST check every column for technical specifications.
- If you see a column named "CPU", "Processor", or "Chip", map it to specs.processor.
- If you see a column named "RAM" or "Memory", map it to specs.ram.
- If you see a column named "SSD", "HDD", "Storage", or "Hard Drive", map it to specs.storage.
- If you see a column named "Screen", "Display", or "Size", map it to specs.screen.
- If you see a column named "Battery" or "Batt", map it to specs.batteryHealth.
- If you see "Color", map it to specs.color.

Example Spec Output:
"specs": { "processor": "i7-8650U", "ram": "16GB", "storage": "512 GB", "color": "Black" }

ITEM NOTES:
- Any column named "Condition", "Notes", "Comments", or "Defects" that contains text (not just model numbers) should be mapped to 'item_notes'.
"""

    if category_descriptions:
        base += f"""

ALLOWED INVENTORY CATEGORIES (for `category_guess`):
{category_descriptions}

Rules for `category_guess`:
- If the sheet contains a 'Category' column, use that value (trimmed).
- Otherwise infer from the product name/specs and choose EXACTLY ONE of the
  ALLOWED INVENTORY CATEGORIES above.
- Copy the category label text EXACTLY as written in the list.
- Do NOT invent new categories.
- If nothing fits, set category_guess = null.
"""
    return base


def _strip_code_fences(s: str) -> str:
    s = s.strip()
    if s.startswith("```"):
        s = re.sub(r"^```[a-zA-Z0-9_-]*\s*", "", s, flags=re.DOTALL)
        s = re.sub(r"\s*```$", "", s, flags=re.DOTALL)
    return s.strip()

def _find_balanced_json(s: str) -> str | None:
    starts = [i for i in (s.find("{"), s.find("[")) if i != -1]
    if not starts: return None
    start = min(starts)
    stack = []
    open_to_close = {"{": "}", "[": "]"}
    for i, ch in enumerate(s[start:], start):
        if ch in "{[": stack.append(open_to_close[ch])
        elif stack and ch == stack[-1]:
            stack.pop()
            if not stack: return s[start:i + 1]
    return None

def parse_gemini_json(resp) -> dict:
    # Kept for compatibility if older code calls it
    # We mainly use parse_ai_json now
    if hasattr(resp, "text"):
        return parse_ai_json(resp.text)
    return parse_ai_json(str(resp))

def parse_ai_json(text_response: str) -> dict:
    # 1. Try direct parse
    try: return json.loads(text_response)
    except: pass
    
    # 2. Strip Fences
    clean = _strip_code_fences(text_response)
    try: return json.loads(clean)
    except: pass
    
    # 3. Find substring
    extracted = _find_balanced_json(clean)
    if extracted:
        try: return json.loads(extracted)
        except: pass
        
    raise ValueError("Could not parse JSON from AI response")

# In ai_utils.py

def _call_pollinations(prompt: str, data_context: str = "", model: str = "openai") -> dict:
    """
    Sends the request to Pollinations.ai.
    Models: 'openai' (GPT-4o-mini equivalent), 'qwen' (Qwen 2.5), 'searchgpt'.
    """
    url = "https://text.pollinations.ai/"
    
    # Construct a strong system message
    full_prompt = f"{prompt}\n\nDATA:\n{data_context}" if data_context else prompt
    
    # Map friendly names to actual model tags if needed, or pass through
    # 'openai' is usually the most stable for JSON. 'qwen' is good for logic.
    selected_model = model if model in ["openai", "qwen", "searchgpt"] else "openai"
    
    payload = {
        "messages": [
            {"role": "system", "content": "You are a strict JSON data processing engine. You only output valid JSON. No conversational text."},
            {"role": "user", "content": full_prompt}
        ],
        "model": selected_model,
        "jsonMode": True
    }
    
    # Increased timeout to 60s and retries
    for attempt in range(2):
        try:
            r = requests.post(url, json=payload, timeout=60)
            if r.status_code == 200:
                return parse_ai_json(r.text)
        except Exception as e:
            print(f"[Pollinations] Attempt {attempt+1} Error: {e}")
            time.sleep(2)
            
    raise RuntimeError("Pollinations API failed to respond")
def _gemini_parse_inline(
    file_bytes: bytes,
    filename: str,
    ext: str,
    expand_units: bool,
    model=None,
    category_descriptions: str | None = None,   # NEW
) -> dict:
    # Redirect to Pollinations
    e = (ext or "").lower().lstrip(".")
    if e in ("xlsx", "xls"):
        # Increase max rows to safe limit for large files
        file_bytes = _xlsx_to_csv_bytes(file_bytes, max_rows=100000)
        filename = re.sub(r"\.(xlsx|xls)$", ".csv", filename, flags=re.I)
        e = "csv"

    prompt = make_ai_parser_prompt(
        expand_units=expand_units,
        category_descriptions=category_descriptions,
    )
    
    # Decode bytes to string for Pollinations
    try:
        csv_text = file_bytes.decode('utf-8')
    except:
        csv_text = file_bytes.decode('latin-1', errors='ignore')

    return _call_pollinations(prompt, csv_text)

def _xlsx_to_csv_bytes(xlsx_bytes: bytes, max_rows: int | None = 100000) -> bytes:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for XLS/XLSX conversion")
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    out = io.StringIO()
    w = csv.writer(out, lineterminator="\n")
    rows = 0
    for row in ws.iter_rows(values_only=True):
        w.writerow(["" if v is None else str(v) for v in row])
        rows += 1
        if max_rows and rows >= max_rows:
            break
    return out.getvalue().encode("utf-8")

def normalize_spreadsheet_upload(filename: str, ext: str, content: bytes) -> tuple[str, str, bytes]:
    clean_ext = (ext or "").lower().lstrip(".")
    if clean_ext in ("xlsx", "xls"):
        try:
            # Use high limit for upload normalization to avoid truncating user's 60k line files
            csv_bytes = _xlsx_to_csv_bytes(content, max_rows=100000)
        except Exception as conv_err:
            raise
        new_filename = re.sub(r"\.(xlsx|xls)$", ".csv", filename, flags=re.I)
        return new_filename, "csv", csv_bytes
    return filename, ext, content

def _postprocess_lines(lines: list[dict], expand_units: bool) -> list[dict]:
    out: list[dict] = []
    for ln in lines or []:
        name = (ln.get("product_name_raw") or "").strip()
        if not name: continue
        try: qty = int(ln.get("qty") or 1)
        except Exception: qty = 1
        qty = max(1, qty)
        unit_cost = to_num(ln.get("unit_cost"))
        msrp = to_num(ln.get("msrp"))
        upc = (str(ln.get("upc")).strip() if ln.get("upc") else None)
        asin = (str(ln.get("asin")).strip() if ln.get("asin") else None)
        category_guess = (str(ln.get("category_guess")).strip() if ln.get("category_guess") else None)
        
        # --- Capture Specs & Notes ---
        specs = ln.get("specs") or {}
        item_notes = ln.get("item_notes")
        # -----------------------------

        base = {
            "product_name_raw": name, "unit_cost": unit_cost, "msrp": msrp,
            "upc": upc, "asin": asin, "category_guess": category_guess,
            "specs": specs, 
            "item_notes": item_notes
        }
        if expand_units:
            for _ in range(qty): out.append({**base, "qty": 1})
        else:
            out.append({**base, "qty": qty})
    return out

def _local_preview_lines_from_bytes(content: bytes, ext: str) -> list[dict]:
    """
    Local parser that preserves raw headers to ensure Specs (CPU/RAM/SSD) are caught.
    """
    rows: list[dict] = []

    # --- LOAD CSV / TXT / TSV ---
    if ext in ("csv", "tsv", "txt"):
        text = content.decode("utf-8", errors="ignore")
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
            reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        except Exception:
            reader = csv.DictReader(io.StringIO(text))
        
        for raw in reader:
            # We map keys for standard fields, but KEEP raw for specs
            mapped = {map_header(k): v for k, v in raw.items()}
            mapped['_raw'] = raw  # <--- CRITICAL: Store original columns
            rows.append(mapped)

    # --- LOAD XLSX ---
    elif ext in ("xlsx", "xlsm", "xltx", "xltm"):
        if load_workbook is None:
            raise HTTPException(400, "openpyxl not installed")
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(header_cells)]
        
        # Map headers
        mapped_headers = [map_header(h) for h in headers]

        for rr in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            raw_dict = {}
            for idx, val in enumerate(rr):
                if idx < len(headers):
                    # Standard mapped key
                    key = mapped_headers[idx]
                    row_dict[key] = val
                    # Raw key for specs
                    raw_key = headers[idx]
                    raw_dict[raw_key] = val
            
            row_dict['_raw'] = raw_dict # <--- CRITICAL
            rows.append(row_dict)

    else:
        raise HTTPException(400, "Unsupported file type. Use CSV or XLSX.")

    # ------------------------------
    # SHAPE INTO PREVIEW LINES
    # ------------------------------

    preview_lines = []

    for r in rows:
        # Use the raw data for specs lookup to avoid 'map_header' stripping "SSD" -> "ssd" etc incorrectly
        raw_data = r.get('_raw', {})
        
        # 1. Product Name Logic
        make = str(r.get("make") or "").strip()
        model = str(r.get("model") or "").strip()
        condition = str(r.get("condition") or "").strip()

        # Heuristic: If condition is a Model Number (e.g. 1769), add to name
        is_model_number = len(condition) < 6 and any(c.isdigit() for c in condition)
        
        parts = [p for p in (make, model) if p]
        if is_model_number:
            parts.append(condition)
            
        structured_name = " ".join(parts).strip()

        fallback_name = (
            r.get("product_name_raw") or r.get("product") or r.get("name") or 
            r.get("description") or r.get("title") or ""
        )
        name = structured_name if structured_name else str(fallback_name).strip()
        
        if not name:
            continue

        # 2. Qty Logic
        try:
            qty = int(re.sub(r"[^0-9]", "", str(r.get("qty") or 1)) or 1)
        except Exception:
            qty = 1

        # 3. Spec Extraction (Iterate RAW headers)
        specs = {}
        for k, v in raw_data.items():
            if not v: continue
            
            # Strict Lowercase Check on RAW header
            key = str(k).strip().lower()
            val = str(v).strip()
            
            if key in ("cpu", "processor", "chip"):
                specs["processor"] = val
            elif key in ("ram", "memory"):
                specs["ram"] = val
            elif key in ("ssd", "hdd", "storage", "hard drive", "disk"):
                specs["storage"] = val
            elif key in ("screen", "display", "size", "monitor"):
                specs["screen"] = val
            elif key in ("color", "colour"):
                specs["color"] = val
            elif "battery" in key:
                specs["batteryHealth"] = val

        # 4. Notes Extraction
        # If 'Condition' was NOT a model number, treat it as a note
        note_parts = []
        if condition and not is_model_number:
            note_parts.append(f"Condition: {condition}")
            
        # Check raw headers for explicit note columns
        for k, v in raw_data.items():
            if not v: continue
            key = str(k).strip().lower()
            if key in ("notes", "comments", "item_notes", "issues", "defects"):
                note_parts.append(str(v).strip())

        preview_lines.append({
            "product_name_raw": name,
            "upc": r.get("upc"),
            "asin": r.get("asin"),
            "qty": qty,
            "unit_cost": to_num(r.get("unit_cost")),
            "msrp": to_num(r.get("msrp")),
            "specs": specs,
            "item_notes": " | ".join(note_parts) if note_parts else None
        })

    return preview_lines


# --- NEW: UPC FINDER UTILS (USING SERPER) ---

def search_web_for_upc_text(query: str) -> str:
    """
    Executes a multi-strategy search to find technical identifiers using Serper.dev.
    Requires SERPER_API_KEY in .env
    """
    api_key = os.getenv("SERPER_API_KEY")
    if not api_key:
        print("Error: SERPER_API_KEY not found in environment variables.")
        return "Search configuration missing."

    url = "https://google.serper.dev/search"
    headers = {
        'X-API-KEY': api_key,
        'Content-Type': 'application/json'
    }

    results = []
    seen_urls = set()

    # Strategy: Two targeted queries to maximize coverage
    queries = [
        f"{query} UPC barcode EAN",          # Broad search
        f"site:ebay.com {query} 'UPC'"       # Specific eBay listings
    ]

    try:
        for q in queries:
            payload = json.dumps({"q": q, "num": 5}) # Top 5 results per query
            response = requests.post(url, headers=headers, data=payload)
            
            if response.status_code == 200:
                data = response.json()
                organic = data.get("organic", [])
                results.extend(organic)
            else:
                print(f"Serper error {response.status_code}: {response.text}")

        # Format results for the AI context
        context_parts = []
        for r in results:
            link = r.get('link', '')
            if link in seen_urls: continue
            seen_urls.add(link)
            
            title = r.get('title', '')
            snippet = r.get('snippet', '')
            
            # Serper often extracts attributes (like Specs/UPC) automatically
            attributes = r.get('attributes', {})
            attr_str = " | ".join([f"{k}: {v}" for k, v in attributes.items()])

            context_parts.append(f"SOURCE: {title}\nTEXT: {snippet}\nATTRIBUTES: {attr_str}\n")

        return "\n".join(context_parts)

    except Exception as e:
        print(f"Search failed: {e}")
        return ""

def ai_find_upc(raw_product_name: str) -> dict:
    """
    Agentic workflow:
    1. AI cleans the title to core model keywords.
    2. Python performs Serper Google search.
    3. AI extracts the code.
    """
    
    # Step 1: Clean the title
    clean_prompt = f"""
    Raw Title: "{raw_product_name}"
    
    Task: Create a search string to find the UPC/Barcode.
    1. Remove "Grade A", "Used", "Good", "Unlocked", internal SKUs.
    2. KEEP specific model numbers (e.g. A1708, MPXQ2LL/A).
    3. KEEP screen size and year.
    
    Return JSON: {{ "search_term": "..." }}
    """
    
    try:
        clean_resp = _call_pollinations(clean_prompt, model="openai")
        search_term = clean_resp.get("search_term", raw_product_name)
    except:
        search_term = raw_product_name

    # Step 2: Perform targeted web search
    search_context = search_web_for_upc_text(search_term)
    
    if not search_context or "Search configuration missing" in search_context:
        err = search_context if search_context else "No search results found"
        return {"upc": None, "confidence": "low", "error": err}

    # Step 3: Extraction
    extract_prompt = f"""
    You are a Barcode Extraction Agent.
    Target Product: {search_term}
    
    Search Results from Google:
    {search_context}
    
    Instructions:
    1. Find the **UPC** (12 digits) or **EAN** (13 digits).
    2. If you see an **MPN** (Manufacturer Part Number, e.g., MPXQ2LL/A), capture that as a fallback.
    3. Ignore generic numbers (phone numbers, prices).
    4. If a snippet says "UPC: X", that is high confidence.
    
    Return JSON:
    {{
        "upc": "string or null", 
        "type": "UPC" | "EAN" | "MPN" | null,
        "confidence": "high" | "medium" | "low"
    }}
    """
    
    return _call_pollinations(extract_prompt, model="openai")

def _parse_locally_from_file(tmp_path: str, suffix: str) -> list[dict]:
    # This reads the temporary file and returns raw rows with an added "_raw" key
    rows: list[dict] = []
    ext = suffix.lower().lstrip(".")
    
    if ext in ("csv", "tsv", "txt"):
        with open(tmp_path, "r", encoding="utf-8", errors="ignore") as fh:
            text = fh.read()
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
            reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        except Exception:
            reader = csv.DictReader(io.StringIO(text))
        
        for raw in reader:
            # Map standard keys
            r = {map_header(k): v for k, v in raw.items()}
            # CRITICAL: Preserve raw dict to catch "CPU", "RAM", "SSD" exactly
            r['_raw'] = raw 
            rows.append(r)

    elif ext in ("xlsx", "xlsm", "xltx", "xltm"):
        if load_workbook is None: raise RuntimeError("openpyxl not installed")
        wb = load_workbook(filename=tmp_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Grab headers from first row
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(header_cells)]
        mapped_headers = [map_header(h) for h in headers]

        for rr in ws.iter_rows(min_row=2, values_only=True):
            obj = {}
            raw_obj = {}
            for i, val in enumerate(rr):
                if i < len(headers):
                    key = mapped_headers[i]
                    obj[key] = val
                    # Store raw header -> val
                    raw_obj[headers[i]] = val
            
            # CRITICAL: Preserve raw dict
            obj['_raw'] = raw_obj
            rows.append(obj)
    else: 
        raise RuntimeError(f"no_local_parser_for_ext:{ext}")
    
    return rows
def _rows_to_lines(rows: list[dict]) -> list[dict]:
    out: list[dict] = []
    for r in rows or []:
        # If the row is empty, skip
        if not any((r or {}).values()):
            continue

        # 1. Product Name
        name = (r.get("product_name_raw") or r.get("product") or r.get("name") or r.get("title") or r.get("description") or "").strip()
        
        # Fallback Name Construction (Make + Model + Condition)
        if not name:
            make = str(r.get("make") or "").strip()
            model = str(r.get("model") or "").strip()
            condition = str(r.get("condition") or "").strip()
            
            # If condition is just a number (e.g. 1769), treat it as part of model/name
            is_model_num = len(condition) < 6 and any(c.isdigit() for c in condition)
            
            parts = [p for p in (make, model) if p]
            if is_model_num: parts.append(condition)
            
            name = " ".join(parts).strip()

        if not name:
            continue

        # 2. Quantity
        try:
            qty = int(re.sub(r"[^\d\-]", "", str(r.get("qty") or 1)) or 1)
        except Exception:
            qty = 1

        # 3. Specs & Notes Extraction (from _raw if available, else mapped keys)
        raw_data = r.get('_raw', r) # fallback to r if _raw missing
        specs = {}
        note_parts = []

        # Check Condition column for notes (if it wasn't a model number)
        cond = str(r.get("condition") or "").strip()
        is_model_num = len(cond) < 6 and any(c.isdigit() for c in cond)
        if cond and not is_model_num:
            note_parts.append(f"Condition: {cond}")

        for k, v in raw_data.items():
            if not v: continue
            k_lower = str(k).strip().lower()
            val_str = str(v).strip()

            # SPECS
            if k_lower in ("cpu", "processor", "chip"): specs["processor"] = val_str
            elif k_lower in ("ram", "memory"): specs["ram"] = val_str
            elif k_lower in ("ssd", "hdd", "storage", "hard drive", "disk"): specs["storage"] = val_str
            elif k_lower in ("screen", "display", "size", "monitor"): specs["screen"] = val_str
            elif k_lower in ("color", "colour"): specs["color"] = val_str
            elif "battery" in k_lower: specs["batteryHealth"] = val_str
            
            # NOTES
            elif k_lower in ("notes", "comments", "item_notes", "issues", "defects"):
                note_parts.append(val_str)

        out.append({
            "product_name_raw": name,
            "qty": max(1, qty),
            "unit_cost": to_num(r.get("unit_cost") or r.get("cost")),
            "msrp": to_num(r.get("msrp")),
            "upc": (str(r.get("upc")).strip() if r.get("upc") else None),
            "asin": (str(r.get("asin")).strip() if r.get("asin") else None),
            "category_guess": (str(r.get("category_guess")).strip() if r.get("category_guess") else None),
            "specs": specs,
            "item_notes": " | ".join(note_parts) if note_parts else None
        })
    return out

def _rows_to_csv_text(rows: list[dict], limit: int = 200) -> str:
    rows = rows[:max(1, limit)]
    headers, seen = [], set()
    for r in rows:
        for k in r.keys():
            m = str(k or "").strip()
            if m and m not in seen:
                seen.add(m); headers.append(m)
    sio = io.StringIO()
    w = csv.DictWriter(sio, fieldnames=headers)
    w.writeheader()
    for r in rows:
        w.writerow({k: r.get(k, "") for k in headers})
    return sio.getvalue()

def _norm_key(name: str | None, upc: str | None, asin: str | None) -> tuple:
    def _clean_id(x):
        if not x: return ""
        s = str(x).strip()
        if re.fullmatch(r"[0-9E+.\-]+", s):
            try:
                from decimal import Decimal
                s = str(int(Decimal(s)))
            except Exception:
                s = re.sub(r"\D", "", s)
        return s

    def _clean_name(x):
        return re.sub(r"[\s\W]+", " ", (x or "").lower()).strip()

    return (_clean_id(upc), _clean_id(asin), _clean_name(name))

def _merge_ai_with_local(ai_lines: list[dict], local_lines: list[dict]) -> list[dict]:
    # With Pollinations, we can just use the AI results as they are usually cleaner.
    # But for safety, if AI failed completely, we fallback.
    # Here we just pass through AI lines for now as it's a replacement.
    if ai_lines:
        return ai_lines
    return local_lines

def _tag_ai_response(payload: dict, via: str, model: str | None = None) -> JSONResponse:
    """
    Wraps a JSON payload and annotates whether Gemini was used.
    Adds both a field in the JSON (`via`) and HTTP headers (`X-AI-Used`, `X-AI-Model`).
    """
    out = dict(payload)
    out["via"] = via
    resp = JSONResponse(out)
    resp.headers["X-AI-Used"] = "1" if via == "gemini" else "0"
    if model:
        resp.headers["X-AI-Model"] = model
    return resp

def ai_health():
    return {
        "genai_imported": True,
        "has_key": True,
        "ai_first": (os.getenv("AI_FIRST", "0") == "1"),
        "configured": True,
    }

# ---------------------------------------------------------
# NEW: Helpers for eBay Extension & Listing Generation
# ---------------------------------------------------------

def normalize_functionality(raw_bullets, category, condition, tester_comment):
    """
    Clean up the AI's bullet points to match house style.
    """
    condition = (condition or "").lower()
    category = (category or "").lower()
    note = (tester_comment or "").strip()
    
    # 1. Filter trivial/spec items
    filtered = []
    for b in raw_bullets:
        s = str(b).strip()
        # Skip trivial
        if re.search(r"power button works|volume button|wifi works|bluetooth works only|camera opens only", s, re.IGNORECASE):
            continue
        # Skip specs (GB, RAM, CPU, etc)
        if re.search(r"(^|\s)(\d+\.?\d*)\s?gb\b|ssd|hdd|ram\b|ddr\d|\bstorage\b|\bcpu\b|\bprocessor\b|\bscreen\b|\bdisplay\b|\binch\b|\bips\b|\bhz\b", s, re.IGNORECASE):
            continue
        if s:
            filtered.append(s)

    # 2. Determine Lead Bullet
    is_parts = re.search(r"part|as[\s-]?is|non[-\s]?working|doesn'?t\s*power|won'?t\s*boot", condition)
    
    if is_parts:
        lead = "NOT FULLY FUNCTIONAL — SOLD AS-IS / FOR PARTS"
    else:
        lead = f"FULLY FUNCTIONAL BUT {note}" if note else "FULLY FUNCTIONAL"

    # 3. Deduplicate
    final_list = [lead]
    seen = {lead.lower()}
    
    for b in filtered:
        if b.lower() not in seen:
            # Phone specific tweaks
            if "phone" in category:
                b = re.sub(r"screen burn\b", "screen burn-in", b, flags=re.IGNORECASE)
                b = re.sub(r"ghost touch", "ghost-touch", b, flags=re.IGNORECASE)
                b = re.sub(r"battery (bad|poor)", "battery weak", b, flags=re.IGNORECASE)
            final_list.append(b)
            seen.add(b.lower())

    # 4. Ensure minimum content
    if len(final_list) == 1:
        fallback = "Primary defect noted above; see photos for exact condition." if is_parts else "All core functions tested (calls, audio, display, charge)."
        final_list.append(fallback)

    return final_list[:5] # Cap at 5



def generate_ebay_listing_content(item_data: dict) -> dict:
    """
    Generates Title and the specific HTML Template requested.
    Now uses Pollinations instead of Gemini for robust free access.
    """
    product_name = item_data.get('product_name_raw', '')
    condition = f"Grade {item_data.get('grade', '')} ({item_data.get('condition', '')})"
    specs = item_data.get('specs', {})
    notes = item_data.get('tester_comment', '')
    category = item_data.get('category_label', '')

    # --- 1. Generate Title ---
    title_prompt = f"""
    You are a professional e-commerce title writer for eBay.
    Write a concise title (≤ 80 chars), no emojis, no ALL CAPS.
    Include brand/model, RAM, storage; optionally one key condition/cosmetic note.
    Do not use quotes. No SKU.
    Use ` | ` to separate specs.
    Put specific cosmetic defects in **DOUBLE ASTERISKS** and CAPS at the end.

    Product: {product_name}
    Condition: {condition}
    Specs: {specs}
    Notes: {notes}
    
    Return ONLY the title text as JSON: {{ "title": "..." }}
    """
    
    try:
        title_resp = _call_pollinations(title_prompt)
        title = title_resp.get("title", product_name[:80])
    except Exception as e:
        print(f"Title Gen Error: {e}")
        title = product_name[:80]

    # --- 2. Generate Description Data ---
    desc_prompt = f"""
    Context: Selling used tech on eBay.
    Product: {product_name}
    Condition: {condition}
    Notes: {notes}
    Specs: {specs}

    Return JSON with keys: 
    - condition (1 concise sentence)
    - functionality (array of short bullet strings, NO SPECS like ram/cpu)
    - included (string, e.g. "Device only")
    """
    
    try:
        data = _call_pollinations(desc_prompt)
    except Exception as e:
        print(f"Desc Gen Error: {e}")
        data = {"condition": condition, "functionality": [], "included": "Device only"}

    # --- 3. Normalize & Format ---
    
    raw_func = data.get("functionality", [])
    if not raw_func and notes: raw_func = [notes]
    
    normalized_func = normalize_functionality(raw_func, category, condition, notes)
    
    func_html = "\n".join([f"  <li>{html.escape(s)}</li>" for s in normalized_func])
    
    inc_text = data.get("included", "Device only")
    if not inc_text:
        lower_cat = category.lower()
        if "phone" in lower_cat: inc_text = "Phone only"
        elif "laptop" in lower_cat: inc_text = "Laptop only (charger if pictured)"
        else: inc_text = "Device only"

    # --- 4. Assemble THE Template ---
    html_output = f"""
<div style="text-align:center; font-weight:700; font-size:20px; margin-bottom:10px;">
  ({html.escape(title)})
</div>

<div style="font-weight:700; margin:6px 0 2px;">Condition:</div>
<ul style="margin:0 0 10px 22px; padding:0; list-style:disc;">
  <li>{html.escape(data.get('condition', condition))}</li>
</ul>

<div style="font-weight:700; margin:6px 0 2px;">Functionality:</div>
<ul style="margin:0 0 10px 22px; padding:0; list-style:disc;">
{func_html}
</ul>

<div style="font-weight:700; margin:6px 0 2px;">Included:</div>
<ul style="margin:0 0 16px 22px; padding:0; list-style:disc;">
  <li>{html.escape(inc_text)}</li>
</ul>

<div style="text-align:center; margin:10px 0; font-weight:700;">
  ****International customers outside of the US may be subject to additional customs or duty fees.****
</div>

<div style="text-align:center; margin:10px 0;">
  Thanks for viewing our product. Don't forget to check out our eBay store for more great deals!
</div>

<div style="text-align:center; margin:10px 0;">
  <a href="https://www.ebay.com/str/discountharddrivesupply" rel="noopener noreferrer">click here for more great deals!</a>
</div>

<div style="margin-top:12px;">
  <img src="https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTOXbPUn7ArDbbFLj5oz4bz75tLImeJH0gERQ&s"
       alt="click here for more great deals!"
       style="display:block; margin:0 auto;" />
</div>
""".strip()

    return {
        "title": title,
        "html": html_output
    }